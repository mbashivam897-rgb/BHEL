"""
Propagate granular current vs non-current BS treatment + FY2026 audited splits
into the full 22-sheet BHEL_Financial_Model.xlsx.

- Rebuild Forecast Financial Statements BS (rows 24+) into granular current/
  non-current line items. Historical FY20-26 = same consolidated series as
  BHEL FM.xlsx (FY26 audited; FY20-25 derived from FY26 composition).
- Expand Working Capital Schedule to drive every granular operating line,
  seeded FY26 audited, forecast via Assumptions.
- Add granular WC driver block to Assumptions (with rationale).
- Re-point Cash Flow dNWC + opening-cash seed; rebuild Other A&L memo.
- Update all external references to moved Forecast BS rows.
"""
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- read history
src = openpyxl.load_workbook('BHEL/BHEL FM.xlsx', data_only=False)
bs = src['BS']
# BHEL FM cols C..I = FY2020..FY2026  ->  full model cols B..H
SRC_COLS = ['C','D','E','F','G','H','I']
def hist(row):
    out=[]
    for col in SRC_COLS:
        v=bs[f'{col}{row}'].value
        out.append(float(v) if isinstance(v,(int,float)) else 0.0)
    return out
H = {
 'sharecap':hist(8),'reserve':hist(9),
 'ltbor':hist(13),'tpnc':hist(14),'clnc':hist(15),'ltprov':hist(16),'othncliab':hist(17),
 'stbor':hist(21),'tpcur':hist(22),'clcur':hist(23),'provcur':hist(24),'othcurliab':hist(25),
 'ppe':hist(33),'cwip':hist(34),'invest':hist(35),'trnc':hist(36),'canc':hist(37),'dta':hist(38),'othncasset':hist(39),
 'trcur':hist(43),'inv':hist(44),'cacur':hist(45),'othcurasset':hist(46),'cash':hist(47),
}
# FY26 audited seeds (last element)
S = {k:v[-1] for k,v in H.items()}

# ---------------------------------------------------------------- open target
wb = openpyxl.load_workbook('BHEL_Financial_Model.xlsx')
FFS = wb['Forecast Financial Statements']
WC  = wb['Working Capital Schedule']
ASM = wb['Assumptions']
CFS = wb['Cash Flow Statement']
OAL = wb['Other Assets & Liabilities']
EQ  = wb['Equity Schedule']

HCOLS=['B','C','D','E','F','G','H']     # FY2020-FY2026
FCOLS=['I','J','K','L','M','N','O']     # FY2027-FY2033
ALL = HCOLS+FCOLS
NUMFMT='#,##0;(#,##0)'
BOLD=Font(name='Calibri',size=10,bold=True)
SECF=PatternFill('solid',fgColor='1F3864'); SEC=Font(size=10,bold=True,color='FFFFFF')
SUB=Font(size=10,bold=True,italic=True,color='1F3864')
TOTF=PatternFill('solid',fgColor='D9E1F2')
BLK=Font(name='Calibri',size=10,color='000000')
GRN=Font(name='Calibri',size=10,color='008000')
EST=Font(name='Calibri',size=10,color='C55A11')   # derived historical estimate
topb=Border(top=Side(style='thin'))

# ============================================================ 1. REBUILD BS
# clear old BS block rows 24..80
for r in range(24,81):
    for c in range(1, FFS.max_column+1):
        FFS.cell(row=r,column=c).value=None
        FFS.cell(row=r,column=c).fill=PatternFill(fill_type=None)

def put(row,col,val,font,fmt=NUMFMT):
    c=FFS.cell(row=row,column=col); c.value=val; c.font=font
    if fmt:c.number_format=fmt
    c.alignment=Alignment(horizontal='right')

def hdr(row,text):
    FFS.cell(row=row,column=1).value=text; FFS.cell(row=row,column=1).font=SEC
    for cc in range(1,len(ALL)+2): FFS.cell(row=row,column=cc).fill=SECF
def sub(row,text):
    FFS.cell(row=row,column=1).value=text; FFS.cell(row=row,column=1).font=SUB
def label(row,text,bold=False):
    c=FFS.cell(row=row,column=1); c.value=text; c.font=BOLD if bold else BLK

# col index helper: A=1, B=2 ... so FY col 'B'->2
def ci(col): return openpyxl.utils.column_index_from_string(col)

# line: hist_key for B-H, forecast formula builder fcst(col)->str
def line(row, text, hist_key, fcst=None, est_hist=False):
    label(row,text)
    vals=H[hist_key]
    for j,col in enumerate(HCOLS):
        # FY26 (last col H) audited -> black; FY20-25 derived -> orange (for the 8 split lines)
        f = BLK if (col=='H' or not est_hist) else EST
        put(row,ci(col),round(vals[j],2),f)
    if fcst:
        for col in FCOLS:
            FFS.cell(row=row,column=ci(col)).value=fcst(col)
            FFS.cell(row=row,column=ci(col)).font=GRN
            FFS.cell(row=row,column=ci(col)).number_format=NUMFMT
            FFS.cell(row=row,column=ci(col)).alignment=Alignment(horizontal='right')

def total(row,text,r1,r2):
    label(row,text,bold=True)
    for col in ALL:
        cc=ci(col)
        FFS.cell(row=row,column=cc).value=f"=SUM({col}{r1}:{col}{r2})"
        FFS.cell(row=row,column=cc).font=BOLD; FFS.cell(row=row,column=cc).number_format=NUMFMT
        FFS.cell(row=row,column=cc).fill=TOTF; FFS.cell(row=row,column=cc).alignment=Alignment(horizontal='right')
def addrow(row,text,a,b):
    label(row,text,bold=True)
    for col in ALL:
        cc=ci(col); FFS.cell(row=row,column=cc).value=f"={col}{a}+{col}{b}"
        FFS.cell(row=row,column=cc).font=BOLD; FFS.cell(row=row,column=cc).number_format=NUMFMT
        FFS.cell(row=row,column=cc).fill=TOTF; FFS.cell(row=row,column=cc).alignment=Alignment(horizontal='right')

FFS.cell(row=24,column=1).value='B.  BALANCE SHEET  (FY2020-26 reported; FY2027-33 forecast)  -  current vs non-current'
FFS.cell(row=24,column=1).font=Font(bold=True,color='1F3864')
# header row 25
FFS.cell(row=25,column=1).value='Particulars'; FFS.cell(row=25,column=1).font=BOLD
yrs=['FY2020','FY2021','FY2022','FY2023','FY2024','FY2025','FY2026','FY2027','FY2028','FY2029','FY2030','FY2031','FY2032','FY2033']
for col,y in zip(ALL,yrs):
    c=FFS.cell(row=25,column=ci(col)); c.value=y; c.font=BOLD; c.alignment=Alignment(horizontal='right')

hdr(26,'EQUITY & LIABILITIES')
sub(27,"Shareholders' Fund")
line(28,'   Share Capital','sharecap', lambda c:f"='Equity Schedule'!{c}12")
line(29,'   Reserve & Surplus','reserve', lambda c:f"='Equity Schedule'!{c}10")
addrow(30,"Total Shareholders' Fund",28,29)
sub(32,'Non-Current Liabilities')
line(33,'   Long-term Borrowings','ltbor', lambda c:f"='Assumptions'!{c}51*'Debt Schedule'!{c}9")
line(34,'   Trade Payables - non-current','tpnc', lambda c:f"='Working Capital Schedule'!{c}18", est_hist=True)
line(35,'   Contract Liabilities - non-current','clnc', lambda c:f"='Working Capital Schedule'!{c}20", est_hist=True)
line(36,'   Long-term Provisions','ltprov', lambda c:f"='Working Capital Schedule'!{c}22")
line(37,'   Other Non-Current Liabilities','othncliab', lambda c:f"='Working Capital Schedule'!{c}24")
total(38,'Total Non-Current Liabilities',33,37)
sub(40,'Current Liabilities')
line(41,'   Short-term Borrowings','stbor', lambda c:f"='Debt Schedule'!{c}9-{c}33")
line(42,'   Trade Payables - current','tpcur', lambda c:f"='Working Capital Schedule'!{c}17")
line(43,'   Contract Liabilities - current','clcur', lambda c:f"='Working Capital Schedule'!{c}19", est_hist=True)
line(44,'   Provisions - current','provcur', lambda c:f"='Working Capital Schedule'!{c}21", est_hist=True)
line(45,'   Other Current Liabilities','othcurliab', lambda c:f"='Working Capital Schedule'!{c}23")
total(46,'Total Current Liabilities',41,45)
addrow(47,'Total Liabilities',38,46)
addrow(49,'TOTAL EQUITY & LIABILITIES',30,47)

hdr(51,'ASSETS')
sub(52,'Non-Current Assets')
line(53,'   Property, Plant & Equipment (incl. intangibles)','ppe', lambda c:f"='Fixed Asset Schedule'!{c}10")
line(54,'   Capital Work in Progress','cwip', lambda c:f"='Assumptions'!{c}44")
line(55,'   Investments','invest', lambda c:f"='Assumptions'!{c}45")
line(56,'   Trade Receivables - non-current','trnc', lambda c:f"='Working Capital Schedule'!{c}9", est_hist=True)
line(57,'   Contract Assets - non-current','canc', lambda c:f"='Working Capital Schedule'!{c}12", est_hist=True)
line(58,'   Deferred Tax Assets','dta', lambda c:f"='Working Capital Schedule'!{c}13", est_hist=True)
line(59,'   Other Non-Current Assets','othncasset', lambda c:f"='Working Capital Schedule'!{c}15")
total(60,'Total Non-Current Assets',53,59)
sub(62,'Current Assets')
line(63,'   Trade Receivables - current','trcur', lambda c:f"='Working Capital Schedule'!{c}8")
line(64,'   Inventories','inv', lambda c:f"='Working Capital Schedule'!{c}10")
line(65,'   Contract Assets - current','cacur', lambda c:f"='Working Capital Schedule'!{c}11", est_hist=True)
line(66,'   Other Current Assets','othcurasset', lambda c:f"='Working Capital Schedule'!{c}14")
line(67,'   Cash & Bank Balances','cash', lambda c:f"='Cash Flow Statement'!{c}25")
total(68,'Total Current Assets',63,67)
addrow(70,'TOTAL ASSETS',60,68)
# balance check 71
label(71,'Balance check (TA - TE&L)')
for col in ALL:
    cc=ci(col); FFS.cell(row=71,column=cc).value=f"={col}70-{col}49"
    FFS.cell(row=71,column=cc).font=Font(size=9,italic=True,color='C00000'); FFS.cell(row=71,column=cc).number_format='0.0'
# borrowings memo 72
label(72,'Memo: total borrowings (LT+ST)')
for col in ALL:
    cc=ci(col); FFS.cell(row=72,column=cc).value=f"={col}33+{col}41"
    FFS.cell(row=72,column=cc).font=Font(size=9,italic=True); FFS.cell(row=72,column=cc).number_format=NUMFMT
FFS.cell(row=73,column=1).value=('Note: FY2026 current/non-current line items are audited actuals (annual-report / BSE Integrated Filing). '
  'FY2020-25 splits of contract assets & liabilities, non-current trade receivables, current provisions and deferred tax '
  '(orange) are derived by applying the FY2026 audited composition to each year\'s reported aggregate; reported section '
  'totals remain audited and the balance sheet ties every year.')
FFS.cell(row=73,column=1).font=Font(size=8,italic=True,color='808080')

# ============================================================ 2. WORKING CAPITAL SCHEDULE
for r in range(5,30):
    for c in range(1,WC.max_column+1):
        WC.cell(row=r,column=c).value=None
WC.cell(row=2,column=1).value='Granular operating working capital (current & non-current). FY26 = audited seed; FY27-33 driven by Assumptions.'
def wlabel(row,text,bold=False):
    c=WC.cell(row=row,column=1); c.value=text; c.font=BOLD if bold else BLK
def wseed(row,key):  # FY26 seed in col H
    c=WC.cell(row=row,column=ci('H')); c.value=round(S[key],2); c.font=BLK; c.number_format=NUMFMT
def wfc(row,col,formula,font=GRN):
    c=WC.cell(row=row,column=ci(col)); c.value=formula; c.font=font; c.number_format=NUMFMT; c.alignment=Alignment(horizontal='right')

WC.cell(row=5,column=1).value='Operating working capital'; WC.cell(row=5,column=1).font=SUB
wlabel(6,'Revenue from operations')
wlabel(7,'Cost of sales (total operating cost)')
for col in FCOLS:
    wfc(6,col,f"='Cost Forecast'!{col}6"); wfc(7,col,f"='Cost Forecast'!{col}24")
# assets
specs_assets=[
 (8,'   Trade receivables - current','trcur',  lambda c:f"='Assumptions'!{c}72*{c}6"),
 (9,'   Trade receivables - non-current','trnc',lambda c:f"='Assumptions'!{c}73*{c}6"),
 (10,'   Inventories','inv',                    lambda c:f"='Assumptions'!{c}74*{c}6"),
 (11,'   Contract assets - current','cacur',    lambda c:f"='Assumptions'!{c}75*{c}6"),
 (12,'   Contract assets - non-current','canc', lambda c:f"='Assumptions'!{c}76*{c}6"),
 (13,'   Deferred tax assets','dta',            lambda c:None),   # flat
 (14,'   Other current assets (residual)','othcurasset', lambda c:f"='Assumptions'!{c}78*{c}6"),
 (15,'   Other non-current assets (residual)','othncasset', lambda c:None), # growth
]
for row,txt,key,fc in specs_assets:
    wlabel(row,txt); wseed(row,key)
    for col in FCOLS:
        prev=ALL[ALL.index(col)-1]
        if row==13: wfc(row,col,f"={prev}13",font=BLK)
        elif row==15: wfc(row,col,f"={prev}15*(1+'Assumptions'!{col}79)")
        else: wfc(row,col,fc(col))
total_w=lambda row,r1,r2,txt:(WC.cell(row=row,column=1).__setattr__('value',txt) or [ (WC.cell(row=row,column=ci(col)).__setattr__('value',f"=SUM({col}{r1}:{col}{r2})"), WC.cell(row=row,column=ci(col)).__setattr__('font',BOLD), WC.cell(row=row,column=ci(col)).__setattr__('number_format',NUMFMT)) for col in (['H']+FCOLS)])
total_w(16,8,15,'Total operating assets'); WC.cell(row=16,column=1).font=BOLD
# liabilities
specs_liab=[
 (17,'   Trade payables - current','tpcur',  lambda c:f"='Assumptions'!{c}80*{c}6"),
 (18,'   Trade payables - non-current','tpnc',lambda c:None),  # flat
 (19,'   Contract liabilities - current','clcur', lambda c:f"='Assumptions'!{c}82*{c}6"),
 (20,'   Contract liabilities - non-current','clnc', lambda c:f"='Assumptions'!{c}83*{c}6"),
 (21,'   Provisions - current','provcur',    lambda c:f"='Assumptions'!{c}84*{c}6"),
 (22,'   Long-term provisions','ltprov',     lambda c:None),  # growth
 (23,'   Other current liabilities (residual)','othcurliab', lambda c:f"='Assumptions'!{c}86*{c}6"),
 (24,'   Other non-current liabilities (residual)','othncliab', lambda c:None), # growth
]
for row,txt,key,fc in specs_liab:
    wlabel(row,txt); wseed(row,key)
    for col in FCOLS:
        prev=ALL[ALL.index(col)-1]
        if row==18: wfc(row,col,f"={prev}18",font=BLK)
        elif row==22: wfc(row,col,f"={prev}22*(1+'Assumptions'!{col}85)")
        elif row==24: wfc(row,col,f"={prev}24*(1+'Assumptions'!{col}87)")
        else: wfc(row,col,fc(col))
WC.cell(row=25,column=1).value='Total operating liabilities'; WC.cell(row=25,column=1).font=BOLD
for col in (['H']+FCOLS):
    WC.cell(row=25,column=ci(col)).value=f"=SUM({col}17:{col}24)"; WC.cell(row=25,column=ci(col)).font=BOLD; WC.cell(row=25,column=ci(col)).number_format=NUMFMT
WC.cell(row=26,column=1).value='Net working capital'; WC.cell(row=26,column=1).font=BOLD
for col in (['H']+FCOLS):
    WC.cell(row=26,column=ci(col)).value=f"={col}16-{col}25"; WC.cell(row=26,column=ci(col)).font=BOLD; WC.cell(row=26,column=ci(col)).number_format=NUMFMT
WC.cell(row=27,column=1).value='(Increase)/decrease in NWC'
for col in FCOLS:
    prev=ALL[ALL.index(col)-1]
    WC.cell(row=27,column=ci(col)).value=f"=-({col}26-{prev}26)"; WC.cell(row=27,column=ci(col)).font=GRN; WC.cell(row=27,column=ci(col)).number_format=NUMFMT

# ============================================================ 3. ASSUMPTIONS WC block
ASM.cell(row=70,column=1).value='C.  WORKING-CAPITAL DETAIL  (current vs non-current; % of revenue unless noted)'
ASM.cell(row=70,column=1).font=Font(bold=True,color='1F3864')
ASM.cell(row=71,column=1).value='driver'; ASM.cell(row=71,column=1).font=BOLD
for col,y in zip(FCOLS,range(2027,2034)):
    ASM.cell(row=71,column=ci(col)).value=y; ASM.cell(row=71,column=ci(col)).font=BOLD
ASM.cell(row=71,column=ci('P')).value='Rationale'; ASM.cell(row=71,column=ci('P')).font=BOLD
def drv(row,name,vals,rat,pct=True):
    ASM.cell(row=row,column=1).value=name
    for col,v in zip(FCOLS,vals):
        c=ASM.cell(row=row,column=ci(col)); c.value=v; c.font=BLK
        c.number_format='0.0%' if pct else '#,##0'; c.alignment=Alignment(horizontal='right')
    ASM.cell(row=row,column=ci('P')).value=rat; ASM.cell(row=row,column=ci('P')).font=Font(size=9,italic=True)
drv(72,'Trade receivables - current (% rev)',[0.20]*7,'FY26 actual 20.1% of revenue (~73 days).')
drv(73,'Trade receivables - non-current (% rev)',[0.072,0.067,0.062,0.057,0.052,0.047,0.042],'Long-dated/disputed (incl. Sudan); FY26 7.2%, runs down.')
drv(74,'Inventories (% rev)',[0.395,0.39,0.385,0.38,0.375,0.37,0.365],'FY26 39.5% of revenue; eases with execution.')
drv(75,'Contract assets - current (% rev)',[0.45,0.44,0.43,0.42,0.41,0.40,0.39],'Unbilled billable <12m; FY26 45.0%.')
drv(76,'Contract assets - non-current (% rev)',[0.40,0.38,0.36,0.34,0.32,0.30,0.28],'Long-cycle unbilled; FY26 42.0%, declines.')
drv(77,'(Deferred tax assets held flat at FY26)','',' ',pct=False); ASM.cell(row=77,column=ci('I')).value=round(S['dta'],0); ASM.cell(row=77,column=ci('I')).number_format='#,##0'
ASM.cell(row=77,column=ci('P')).value='Held flat at FY26 audited level (Rs 3,533 cr).'; ASM.cell(row=77,column=ci('P')).font=Font(size=9,italic=True)
drv(78,'Other current assets residual (% rev)',[0.126,0.124,0.122,0.120,0.118,0.116,0.114],'Deposits, current tax & advances ex-contract; FY26 12.6%.')
drv(79,'Other non-current assets residual growth %',[0.03]*7,'LT deposits & loans; slow growth.')
drv(80,'Trade payables - current (% rev)',[0.311,0.311,0.305,0.305,0.30,0.30,0.295],'FY26 31.1% of revenue (~167 days COGS).')
drv(81,'(Trade payables non-current held flat)','',' ',pct=False); ASM.cell(row=81,column=ci('I')).value=round(S['tpnc'],0); ASM.cell(row=81,column=ci('I')).number_format='#,##0'
ASM.cell(row=81,column=ci('P')).value='Held flat at FY26 (Rs 1,343 cr).'; ASM.cell(row=81,column=ci('P')).font=Font(size=9,italic=True)
drv(82,'Contract liabilities - current (% rev)',[0.27,0.27,0.26,0.26,0.25,0.25,0.24],'Customer advances <12m; FY26 27.0%.')
drv(83,'Contract liabilities - non-current (% rev)',[0.38,0.36,0.34,0.32,0.30,0.28,0.26],'LT customer advances (interest-free WC); FY26 39.7%.')
drv(84,'Provisions - current (% rev)',[0.057,0.057,0.056,0.056,0.055,0.055,0.054],'Warranty/LD/employee; FY26 5.7%.')
drv(85,'Long-term provisions growth %',[0.03]*7,'Warranty & employee-benefit provisions.')
drv(86,'Other current liabilities residual (% rev)',[0.082,0.082,0.08,0.08,0.078,0.078,0.076],'Statutory/employee dues, deposits ex-contract & provisions; FY26 8.2%.')
drv(87,'Other non-current liabilities residual growth %',[0.04]*7,'LT deposits & deferred grants.')

# ============================================================ 4. CASH FLOW re-point
for col in FCOLS:
    CFS.cell(row=8,column=ci(col)).value=f"='Working Capital Schedule'!{col}27"  # dNWC
# opening cash FY27 = FY26 BS cash (audited); closing FY26 display = audited
CFS.cell(row=25,column=ci('H')).value=round(S['cash'],2); CFS.cell(row=25,column=ci('H')).number_format=NUMFMT
CFS.cell(row=24,column=ci('I')).value="='Forecast Financial Statements'!H67"

# ============================================================ 5. Other A&L memo re-point
OAL.cell(row=8,column=1).value='Contract assets - current (memo, from WC)'
OAL.cell(row=10,column=1).value='Contract liabilities - current (memo, from WC)'
for col in FCOLS:
    OAL.cell(row=8,column=ci(col)).value=f"='Working Capital Schedule'!{col}11"
    OAL.cell(row=9,column=ci(col)).value=f"='Working Capital Schedule'!{col}14"
    OAL.cell(row=10,column=ci(col)).value=f"='Working Capital Schedule'!{col}19"
    OAL.cell(row=11,column=ci(col)).value=f"='Working Capital Schedule'!{col}23"

# Equity Schedule: share capital constant -> 696.41
for col in FCOLS:
    EQ.cell(row=12,column=ci(col)).value=696.41

# ============================================================ 6. update external refs to moved BS rows
rowmap={38:42,42:49,47:53,55:63,56:64,58:67,61:70,62:71,63:72}
sheets_to_fix=['Ratio Analysis','Error Checks','DCF Valuation','Relative Valuation','Scenario Analysis','Dashboard']
pat=re.compile(r"(Forecast Financial Statements'!\$?[A-Z]{1,2}\$?)(\d+)")
def repl(m):
    pre,r=m.group(1),int(m.group(2))
    return pre+str(rowmap.get(r,r)) if r in rowmap else m.group(0)
for sh in sheets_to_fix:
    ws=wb[sh]
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and 'Forecast Financial Statements' in c.value:
                c.value=pat.sub(repl,c.value)

wb.save('BHEL_Financial_Model.xlsx')
print("Full model restructured. FY26 seeds:",{k:round(v,1) for k,v in S.items()})
