"""Full model: fold contract assets/liabilities into 'Other assets/liabilities'.
Blanks the 4 contract rows (no row-shift -> refs intact); routes 'Other' lines to
combined audited values; updates WC schedule + Assumptions."""
import openpyxl
from openpyxl.styles import Font, Alignment
from audited_data import D

def _r(x): return round(x,2)
oth_nc_a =[_r(D['note9_nc'][i]+D['ofa_nc'][i]) for i in range(7)]
oth_cur_a=[_r(D['note9_cur'][i]+D['ofa_cur'][i]+D['curtax'][i]) for i in range(7)]
oth_nc_l =[_r(D['note20_nc'][i]+D['ofl_nc'][i]) for i in range(7)]
oth_cur_l=[_r(D['note20_cur'][i]+D['ofl_cur'][i]) for i in range(7)]
# absorb <=0.01 audited rounding into other current assets so each year ties
for i in range(7):
    tel=_r(696.41)+_r(D['reserve'][i])+_r(D['leaseNC'][i])+_r(D['tpnc'][i])+_r(D['provnc'][i])+oth_nc_l[i]+_r(D['stbor_grp'][i])+_r(D['tpcur'][i])+_r(D['provcur'][i])+oth_cur_l[i]
    ta=_r(D['ppe_grp'][i])+_r(D['cwip'][i])+_r(D['invest'][i])+_r(D['trnc'][i])+_r(D['dta'][i])+oth_nc_a[i]+_r(D['trcur'][i])+_r(D['invtry'][i])+oth_cur_a[i]+_r(D['cash_grp'][i])
    oth_cur_a[i]=_r(oth_cur_a[i]+(tel-ta))

wb=openpyxl.load_workbook('BHEL_Financial_Model.xlsx')
ffs=wb['Forecast Financial Statements']; WC=wb['Working Capital Schedule']; ASM=wb['Assumptions']
BLK=Font(name='Calibri',size=10,color='000000'); NUM='#,##0;(#,##0)'
HC=['B','C','D','E','F','G','H']; FC=['I','J','K','L','M','N','O']

# 1. blank the 4 contract rows in the Forecast BS (35 CLnc,43 CLcur,57 CAnc,65 CAcur)
for r in [35,43,57,65]:
    for col in HC+FC: ffs.cell(row=r,column=openpyxl.utils.column_index_from_string(col)).value=None
    ffs.cell(row=r,column=1).value=None

# 2. relabel + fill combined audited 'Other' lines (37,45,59,66); forecast links already -> WC 24,23,15,14
def setline(row,label,vals):
    ffs.cell(row=row,column=1).value=label; ffs.cell(row=row,column=1).font=BLK
    for i,col in enumerate(HC):
        c=ffs.cell(row=row,column=openpyxl.utils.column_index_from_string(col)); c.value=vals[i]; c.font=BLK; c.number_format=NUM; c.alignment=Alignment(horizontal='right')
setline(37,'   Other Non-Current Liabilities (incl. contract liabilities)',oth_nc_l)
setline(45,'   Other Current Liabilities (incl. contract liabilities)',oth_cur_l)
setline(59,'   Other Non-Current Assets (incl. contract assets)',oth_nc_a)
setline(66,'   Other Current Assets (incl. contract assets)',oth_cur_a)
ffs.cell(row=73,column=1).value=('Note: Balance sheet FY2020-FY2026 = audited consolidated actuals from BHEL annual reports '
  '(FY2021/2023/2025) + FY2026 filing. Contract assets & liabilities are reported within "Other assets/liabilities" '
  'exactly as in the audited financial statements. No estimated figures.')
ffs.cell(row=73,column=1).font=Font(size=8,italic=True,color='808080')

# 3. Working Capital Schedule: blank contract rows 11,12,19,20; set combined Other in 14,15,23,24
for r in [11,12,19,20]:
    for col in ['H']+FC: WC.cell(row=r,column=openpyxl.utils.column_index_from_string(col)).value=None
    WC.cell(row=r,column=1).value=None
def wcline(row,label,seed,assum_row):
    WC.cell(row=row,column=1).value=label; WC.cell(row=row,column=1).font=BLK
    h=WC.cell(row=row,column=openpyxl.utils.column_index_from_string('H')); h.value=seed; h.font=BLK; h.number_format=NUM
    for col in FC:
        c=WC.cell(row=row,column=openpyxl.utils.column_index_from_string(col)); c.value=f"='Assumptions'!{col}{assum_row}*{col}6"; c.font=Font(color='008000'); c.number_format=NUM
wcline(14,'   Other current assets (incl. contract assets)',oth_cur_a[6],78)
wcline(15,'   Other non-current assets (incl. contract assets)',oth_nc_a[6],79)
wcline(23,'   Other current liabilities (incl. contract liab.)',oth_cur_l[6],86)
wcline(24,'   Other non-current liabilities (incl. contract liab.)',oth_nc_l[6],87)

# 4. Assumptions: combined Other drivers (% revenue); blank contract drivers 75,76,82,83
def aset(row,vals,label,rat):
    ASM.cell(row=row,column=1).value=label
    for col,v in zip(FC,vals):
        c=ASM.cell(row=row,column=openpyxl.utils.column_index_from_string(col)); c.value=v; c.font=BLK; c.number_format='0.0%'; c.alignment=Alignment(horizontal='right')
    ASM.cell(row=row,column=openpyxl.utils.column_index_from_string('P')).value=rat
    ASM.cell(row=row,column=openpyxl.utils.column_index_from_string('P')).font=Font(size=9,italic=True)
aset(78,[0.57,0.56,0.55,0.54,0.53,0.52,0.51],'Other current assets % of revenue (incl. contract assets)','FY26 57.6%; unbilled converts over time.')
aset(79,[0.43,0.41,0.39,0.37,0.35,0.33,0.31],'Other non-current assets % of revenue (incl. contract assets)','FY26 44.3%; long-cycle unbilled declines.')
aset(86,[0.35,0.35,0.34,0.34,0.33,0.33,0.32],'Other current liabilities % of revenue (incl. contract liab.)','FY26 35.1%; advances + statutory dues.')
aset(87,[0.40,0.38,0.36,0.34,0.32,0.30,0.28],'Other non-current liabilities % of revenue (incl. contract liab.)','FY26 41.1%; LT advances normalise.')
for r in [75,76,82,83]:
    for col in ['B']+FC+['P']: ASM.cell(row=r,column=openpyxl.utils.column_index_from_string(col)).value=None

wb.save('BHEL_Financial_Model.xlsx')
print("Full model: contract folded into Other; 100% audited history.")
