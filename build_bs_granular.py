"""
Enhance BHEL FM.xlsx:
 1. Rebuild the BS sheet into a clearly differentiated current vs non-current
    structure with granular line items.
 2. FY2026 (col I) populated with exact audited splits from the uploaded
    BSE Integrated Filing PDF (consolidated).
 3. FY2020-25 retained from the user's balancing consolidated series, with the
    BHEL-specific cross-cutting items kept inside clearly-labelled residual
    'Other' lines (no fabricated historical granularity).
 4. Forecast FY2027-33 driven granularly off the FY26 seed via the Forecast
    Engine + Assumptions (contract assets / liabilities, NC receivables,
    current provisions, deferred tax modelled explicitly).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = 'BHEL/BHEL FM.xlsx'
wb = openpyxl.load_workbook(PATH)

# ---------- styles ----------
BLACK = Font(name='Calibri', size=10, color='000000')           # hardcoded input
BLUE  = Font(name='Calibri', size=10, color='0000CC')           # in-sheet formula
GREEN = Font(name='Calibri', size=10, color='008000')           # cross-sheet link
BOLD  = Font(name='Calibri', size=10, bold=True)
TITLE = Font(name='Calibri', size=12, bold=True, color='1F3864')
SEC   = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
SUBSEC= Font(name='Calibri', size=10, bold=True, italic=True, color='1F3864')
SECFILL = PatternFill('solid', fgColor='1F3864')
TOTFILL = PatternFill('solid', fgColor='D9E1F2')
GREY    = PatternFill('solid', fgColor='F2F2F2')
NUMFMT = '#,##0.0;(#,##0.0)'
topb = Border(top=Side(style='thin'))
dblb = Border(top=Side(style='thin'), bottom=Side(style='double'))

HCOLS = ['C','D','E','F','G','H','I']           # 2020-2026
FCOLS = ['J','K','L','M','N','O','P']           # 2027-2033
ALLCOLS = HCOLS + FCOLS

def num(ws, cell, val, font=BLACK):
    c = ws[cell]; c.value = val; c.font = font; c.number_format = NUMFMT
    c.alignment = Alignment(horizontal='right')

# =========================================================================
# 1. REBUILD BS SHEET
# =========================================================================
ws = wb['BS']
# clear existing used range
for r in range(1, ws.max_row+2):
    for c in range(1, ws.max_column+2):
        ws.cell(row=r, column=c).value = None

ws['B2'] = 'Balance Sheet  (Consolidated, INR Crore)  -  current vs non-current'
ws['B2'].font = TITLE

# year header row 4
ws['B4'] = 'Particulars'; ws['B4'].font = BOLD
ws['C4'] = 2020; ws['C4'].font = BOLD
for i, col in enumerate(['D','E','F','G','H','I','J','K','L','M','N','O','P']):
    prev = ['C','D','E','F','G','H','I','J','K','L','M','N','O'][i]
    ws[f'{col}4'] = f'={prev}4+1'; ws[f'{col}4'].font = BOLD
for col in ALLCOLS:
    ws[f'{col}4'].alignment = Alignment(horizontal='right')
# tag historical vs forecast
ws['C5']=None

def section(row, label):
    ws[f'B{row}'] = label; ws[f'B{row}'].font = SEC
    for col in ['A','B']+ALLCOLS:
        ws[f'{col}{row}'].fill = SECFILL
def subsec(row, label):
    ws[f'B{row}'] = label; ws[f'B{row}'].font = SUBSEC
def total(row, label, formula_cols=True, rng=None):
    ws[f'B{row}'] = label; ws[f'B{row}'].font = BOLD
    for col in ALLCOLS:
        if rng:
            ws[f'{col}{row}'] = f'=SUM({col}{rng[0]}:{col}{rng[1]})'
        ws[f'{col}{row}'].font = BOLD
        ws[f'{col}{row}'].number_format = NUMFMT
        ws[f'{col}{row}'].fill = TOTFILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

# ---- line writer: hist values list (len up to 7 for C..I), fe row for forecast ----
def line(row, label, hist, fe_row=None, fe_sign=1, fcst_const=None, indent=True):
    ws[f'B{row}'] = ('   '+label) if indent else label
    ws[f'B{row}'].font = BLACK if indent else BOLD
    # historical C..I
    for j, col in enumerate(HCOLS):
        v = hist[j] if (hist is not None and j < len(hist)) else None
        if v is not None:
            num(ws, f'{col}{row}', v, BLACK)
    # forecast J..P
    for col in FCOLS:
        if fe_row is not None:
            sign = '' if fe_sign==1 else '-'
            ws[f'{col}{row}'] = f"={sign}'Forecast Engine'!{col}{fe_row}"
            ws[f'{col}{row}'].font = GREEN
        elif fcst_const is not None:
            ws[f'{col}{row}'] = fcst_const
            ws[f'{col}{row}'].font = BLACK
        ws[f'{col}{row}'].number_format = NUMFMT
        ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

# ===================== EQUITY & LIABILITIES =====================
ws['B6'] = 'EQUITY & LIABILITIES'; ws['B6'].font = SEC
for col in ['A','B']+ALLCOLS: ws[f'{col}6'].fill = SECFILL
ws['A6'] = 'x'

subsec(7, "Shareholders' Fund")
line(8, 'Share Capital', [696.41,696.41,696.41,696.41,696.41,696.41,696.41], fcst_const=696.41)
line(9, 'Reserve & Surplus (Other Equity)', [27964.31,25287.25,25810.19,23681.85,23742.24,24025.75,25450.19], fe_row=30)
total(10, "Total Shareholders' Fund", rng=(8,9))

subsec(12, 'Non-Current Liabilities')
line(13, 'Long-term Borrowings', [0,0,0,0,0,0,168.18], fe_row=20)
line(14, 'Trade Payables - non-current', [None,None,None,None,None,None,1343.44], fe_row=60)
line(15, 'Contract Liabilities - non-current', [None,None,None,None,None,None,13413.24], fe_row=59)
line(16, 'Long-term Provisions', [4225.16,3925.56,3771.21,4101.02,2489.08,2585.56,2354.62], fe_row=39)
line(17, 'Other Non-Current Liabilities', [4263.27,4982.75,4594.80,5089.29,6744.96,12549.87,457.35], fe_row=38)
total(18, 'Total Non-Current Liabilities', rng=(13,17))

subsec(20, 'Current Liabilities')
line(21, 'Short-term Borrowings', [4947.92,4849.28,4745,5385,8808,8795,8018.77], fe_row=21)
line(22, 'Trade Payables - current', [8829.16,6683.51,7749.59,9895.83,8696.24,9540.92,10491.60], fe_row=34)
line(23, 'Contract Liabilities - current', [None,None,None,None,None,None,9110.39], fe_row=58)
line(24, 'Provisions - current', [None,None,None,None,None,None,1918.91], fe_row=61)
line(25, 'Other Current Liabilities', [8831.40,8827.11,8876.56,8070.62,7828.57,9889.67,2762.51], fe_row=36)
total(26, 'Total Current Liabilities', rng=(21,25))

total(27, 'Total Liabilities')
for col in ALLCOLS:
    ws[f'{col}27'] = f'={col}18+{col}26'; ws[f'{col}27'].font=BOLD; ws[f'{col}27'].number_format=NUMFMT; ws[f'{col}27'].fill=TOTFILL

ws['A29']='x'; ws['B29']='TOTAL EQUITY & LIABILITIES'; ws['B29'].font=BOLD
for col in ALLCOLS:
    ws[f'{col}29'] = f'={col}10+{col}27'; ws[f'{col}29'].font=BOLD; ws[f'{col}29'].number_format=NUMFMT
    ws[f'{col}29'].border=topb; ws[f'{col}29'].fill=TOTFILL; ws[f'{col}29'].alignment=Alignment(horizontal='right')

# ===================== ASSETS =====================
ws['B31'] = 'ASSETS'; ws['B31'].font = SEC
for col in ['A','B']+ALLCOLS: ws[f'{col}31'].fill = SECFILL
ws['A31']='x'

subsec(32, 'Non-Current Assets')
line(33, 'Property, Plant & Equipment (incl. intangibles)', [2817,2491,2398,2476,2574,2947,3094.38], fe_row=14)
line(34, 'Capital Work in Progress', [314,420,431,354,308,195,399.20], fe_row=40)
line(35, 'Investments (equity method)', [162,185,205,235,256,276,302.06], fe_row=41)
line(36, 'Trade Receivables - non-current', [None,None,None,None,None,None,2426.93], fe_row=52)
line(37, 'Contract Assets - non-current', [None,None,None,None,None,None,14196.72], fe_row=54)
line(38, 'Deferred Tax Assets (net)', [None,None,None,None,None,None,3532.80], fe_row=55)
line(39, 'Other Non-Current Assets', [23737.12,23784.50,25339.12,23763.48,21295.77,21871.68,778.17], fe_row=37)
total(40, 'Total Non-Current Assets', rng=(33,39))

subsec(42, 'Current Assets')
line(43, 'Trade Receivables - current', [7108.60,4035.07,3024.75,3128.35,4785.38,5884.35,6796.27], fe_row=32)
line(44, 'Inventories', [8908.23,7194.45,6560.21,6755.90,7220.57,9869.49,13334.58], fe_row=33)
line(45, 'Contract Assets - current', [None,None,None,None,None,None,15192.89], fe_row=53)
line(46, 'Other Current Assets', [10292.09,10440.40,11131.99,13564.71,16408.31,19427.25,4264.99], fe_row=35)
line(47, 'Cash & Bank Balances', [6418.59,6701.45,7153.69,6642.58,6157.47,7612.41,11866.62], fe_row=50)
total(48, 'Total Current Assets', rng=(43,47))

ws['B50']='TOTAL ASSETS'; ws['B50'].font=BOLD
for col in ALLCOLS:
    ws[f'{col}50'] = f'={col}40+{col}48'; ws[f'{col}50'].font=BOLD; ws[f'{col}50'].number_format=NUMFMT
    ws[f'{col}50'].border=topb; ws[f'{col}50'].fill=TOTFILL; ws[f'{col}50'].alignment=Alignment(horizontal='right')

ws['B52']='Balance check (TA - TE&L)'; ws['B52'].font=Font(size=9,italic=True)
for col in ALLCOLS:
    ws[f'{col}52'] = f'={col}50-{col}29'; ws[f'{col}52'].font=Font(size=9,italic=True,color='C00000'); ws[f'{col}52'].number_format='0.00'

# note
ws['B54']=('Note: Granular current/non-current line-item detail (contract assets & liabilities, non-current '
           'trade receivables, current provisions, deferred tax) is shown from FY2026 - the audited year sourced '
           'from the BSE Integrated Filing. FY2020-25 retain the consolidated series within the labelled "Other" '
           'residual lines; section totals and balancing are preserved for every year.')
ws['B54'].font=Font(size=8,italic=True,color='808080')

ws.column_dimensions['A'].width=3
ws.column_dimensions['B'].width=44
for col in ALLCOLS: ws.column_dimensions[col].width=10.5
ws.freeze_panes='C5'

# =========================================================================
# 2. ASSUMPTIONS  (repurpose 39-42 residual drivers + add granular block)
# =========================================================================
a = wb['Assumptions']
def aset(cell, val, font=BLACK, fmt='0.0%'):
    c=a[cell]; c.value=val; c.font=font; c.number_format=fmt; c.alignment=Alignment(horizontal='right')

# repurpose existing residual drivers (39-42) with FY26-anchored values + rationale
a['B39']='Other current assets residual % of sales'
for col,v in zip(FCOLS,[0.120,0.118,0.115,0.112,0.110,0.108,0.105]): aset(f'{col}39',v)
a['Q39']='Security deposits, current tax assets & advances (ex-contract assets). FY26 actual 12.6% of sales.'
a['B40']='Other current liabilities residual % of sales'
for col,v in zip(FCOLS,[0.082,0.082,0.080,0.080,0.078,0.078,0.076]): aset(f'{col}40',v)
a['Q40']='Statutory dues, employee dues & deposits (ex-contract liabilities & current provisions). FY26 actual 8.2%.'
a['B41']='Other non-current assets residual growth %'
for col in FCOLS: aset(f'{col}41',0.03)
a['Q41']='LT security deposits & loans (ex-deferred tax, NC trade receivables, NC contract assets). Slow growth.'
a['B42']='Other non-current liabilities residual growth %'
for col in FCOLS: aset(f'{col}42',0.04)
a['Q42']='LT deposits from contractors & deferred grants (ex-contract liabilities & NC trade payables).'

# new granular block (rows 67+)
a['B67']='D.  WORKING-CAPITAL DETAIL  (current vs non-current split - fixed drivers)'
a['B67'].font=Font(bold=True, color='1F3864')
a['B68']='ratios / INR Crore'; a['B68'].font=BOLD
for col,yr in zip(FCOLS,range(2027,2034)):
    a[f'{col}68']=yr; a[f'{col}68'].font=BOLD; a[f'{col}68'].alignment=Alignment(horizontal='right')
a['Q68']='Rationale'; a['Q68'].font=BOLD

def adrv(row, label, vals, rationale):
    a[f'B{row}']=label
    for col,v in zip(FCOLS,vals): aset(f'{col}{row}',v)
    a[f'Q{row}']=rationale; a[f'Q{row}'].font=Font(size=9,italic=True)

adrv(69,'Contract assets - current % of sales',[0.45,0.44,0.43,0.42,0.41,0.40,0.39],
     'Unbilled revenue billable within 12m. FY26 actual 45.0% of sales; eases as execution/billing cycle improves.')
adrv(70,'Contract assets - non-current % of sales',[0.40,0.38,0.36,0.34,0.32,0.30,0.28],
     'Long-cycle unbilled revenue (>12m). FY26 actual 42.0%; declines as legacy power projects bill out.')
adrv(71,'Trade receivables - non-current % of sales',[0.070,0.065,0.060,0.055,0.050,0.045,0.040],
     'Long-dated/retention & disputed receivables (incl. Sudan STPG). FY26 actual 7.2%; runs down over time.')
adrv(72,'Contract liabilities - current % of sales',[0.27,0.27,0.26,0.26,0.25,0.25,0.24],
     'Customer advances/excess billing released <12m. FY26 actual 27.0%; tracks ordering momentum.')
adrv(73,'Contract liabilities - non-current % of sales',[0.38,0.36,0.34,0.32,0.30,0.28,0.26],
     'Long-term customer advances - key interest-free WC funding. FY26 actual 39.7%; normalises as orders execute.')
adrv(74,'Provisions - current % of sales',[0.055,0.055,0.054,0.054,0.053,0.053,0.052],
     'Warranty, liquidated damages & employee provisions. FY26 actual 5.7% of sales.')

# refresh a couple of fixed rationales with FY26 actuals
a['Q36']='Largely fixed cost base; declines as % as sales scale. FY26 actual 19.1% of sales.'
a['Q38']='Supplier credit ~167 days of COGS (FY26 actual).'

# =========================================================================
# 3. FORECAST ENGINE  (re-point seeds to new BS rows + add granular rows)
# =========================================================================
fe = wb['Forecast Engine']
def feset(cell, formula, font=GREEN):
    c=fe[cell]; c.value=formula; c.font=font; c.number_format=NUMFMT

# update FY26 seeds (col I) to new BS row addresses
fe['I4']="='PL'!I5"
fe['I11']="='BS'!I33"; fe['I14']="='BS'!I33"
fe['I17']="='BS'!I13+'BS'!I21"; fe['I20']="='BS'!I13"
fe['I30']="='BS'!I9"; fe['I29']="='BS'!I9"
fe['I32']="='BS'!I43"; fe['I33']="='BS'!I44"; fe['I34']="='BS'!I22"
fe['I35']="='BS'!I46"; fe['I36']="='BS'!I25"; fe['I37']="='BS'!I39"
fe['I38']="='BS'!I17"; fe['I39']="='BS'!I16"
fe['I40']="='BS'!I34"; fe['I41']="='BS'!I35"
fe['I49']="='BS'!I47"; fe['I50']="='BS'!I47"

# residual forecast formulas (rows 35,36,37,38) already reference Assumptions 39/40/41/42 - keep
# OCAresid 35 = sales * J39 ; OCLresid 36 = sales * J40 ; ONCAresid 37 growth J41 ; ONCLresid 38 growth J42
for col in FCOLS:
    fe[f'{col}35']=f"={col}4*'Assumptions'!{col}39"; fe[f'{col}35'].font=GREEN; fe[f'{col}35'].number_format=NUMFMT
    fe[f'{col}36']=f"={col}4*'Assumptions'!{col}40"; fe[f'{col}36'].font=GREEN; fe[f'{col}36'].number_format=NUMFMT

# ---- new granular engine rows ----
fe['B52']='TR non-current'; fe['I52']="='BS'!I36"
fe['B53']='CA current';     fe['I53']="='BS'!I45"
fe['B54']='CA non-current'; fe['I54']="='BS'!I37"
fe['B55']='Deferred tax';   fe['I55']="='BS'!I38"
fe['B58']='CL current';     fe['I58']="='BS'!I23"
fe['B59']='CL non-current'; fe['I59']="='BS'!I15"
fe['B60']='TP non-current'; fe['I60']="='BS'!I14"
fe['B61']='Provisions cur'; fe['I61']="='BS'!I24"
for rr in [52,53,54,55,58,59,60,61]:
    fe[f'B{rr}'].font=BLACK
    fe[f'I{rr}'].font=GREEN; fe[f'I{rr}'].number_format=NUMFMT
for col in FCOLS:
    feset(f'{col}52', f"={col}4*'Assumptions'!{col}71")
    feset(f'{col}53', f"={col}4*'Assumptions'!{col}69")
    feset(f'{col}54', f"={col}4*'Assumptions'!{col}70")
    pcol = ALLCOLS[ALLCOLS.index(col)-1]
    feset(f'{col}55', f"={pcol}55", font=BLUE)          # deferred tax held flat
    feset(f'{col}58', f"={col}4*'Assumptions'!{col}72")
    feset(f'{col}59', f"={col}4*'Assumptions'!{col}73")
    feset(f'{col}60', f"={pcol}60", font=BLUE)          # TP non-current held flat
    feset(f'{col}61', f"={col}4*'Assumptions'!{col}74")

# ---- redefine NWC (row 42) to include all granular operating items ----
for col in ALLCOLS:
    assets = f"{col}32+{col}52+{col}33+{col}53+{col}54+{col}55+{col}35+{col}37"
    liabs  = f"{col}34+{col}60+{col}58+{col}59+{col}61+{col}39+{col}36+{col}38"
    fe[f'{col}42']=f"=({assets})-({liabs})"; fe[f'{col}42'].font=BLUE; fe[f'{col}42'].number_format=NUMFMT

wb.save(PATH)
print("Saved enhanced workbook.")
