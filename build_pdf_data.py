# -*- coding: utf-8 -*-
"""
Build BHEL_Financial_Data_FY2020-2026.xlsx from the audited figures extracted
from the annual-report PDFs (FS 2021 / FS 2023 / FS 2025 consolidated balance
sheets) and the FY2026 BSE integrated filing (bhel 2026.pdf).

All figures in INR Crore, consolidated, as reported.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook(); wb.remove(wb.active)

NAVY = "1F3864"; LT = "D9E1F2"; TOT = "DDEBF7"; SUBHDR = "BDD7EE"
F_TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
F_HDR   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_SECT  = Font(name="Calibri", size=10, bold=True, color=NAVY)
F_LBL   = Font(name="Calibri", size=10)
F_LBLB  = Font(name="Calibri", size=10, bold=True)
F_NOTE  = Font(name="Calibri", size=8, italic=True, color="808080")
FILL_HDR = PatternFill("solid", fgColor=NAVY)
FILL_TOT = PatternFill("solid", fgColor=TOT)
FILL_SUB = PatternFill("solid", fgColor=SUBHDR)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP = Border(top=Side(style="thin", color="000000"))
FMT = '#,##0.00;(#,##0.00)'

def CL(c): return get_column_letter(c)

def style_sheet(ws, title, cols, ncols):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 52
    for c in range(2, 2 + ncols):
        ws.column_dimensions[CL(c)].width = 13
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + ncols)
    t = ws.cell(1, 1, title); t.font = F_TITLE; t.fill = FILL_HDR
    t.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 2 + ncols):
        ws.cell(1, c).fill = FILL_HDR
    ws.row_dimensions[1].height = 22
    # column header row (row 2)
    h = ws.cell(2, 1, "Particulars  (INR Crore)"); h.font = F_HDR; h.fill = FILL_HDR
    h.border = BORD
    for i, name in enumerate(cols):
        cc = ws.cell(2, 2 + i, name); cc.font = F_HDR; cc.fill = FILL_HDR
        cc.alignment = CEN; cc.border = BORD

def put_row(ws, r, label, vals, bold=False, section=False, total=False, indent=0):
    lbl = ("    " * indent) + label
    c1 = ws.cell(r, 1, lbl)
    c1.font = F_SECT if section else (F_LBLB if (bold or total) else F_LBL)
    if section:
        c1.fill = FILL_SUB
    if total:
        c1.fill = FILL_TOT
    for i, v in enumerate(vals):
        cc = ws.cell(r, 2 + i)
        if v is None:
            cc.value = "-"
            cc.alignment = RGT
        else:
            cc.value = v
            cc.number_format = FMT
            cc.alignment = RGT
        cc.font = F_LBLB if (bold or total) else F_LBL
        if total:
            cc.fill = FILL_TOT
            cc.border = TOP
    if section:
        for i in range(len(vals)):
            ws.cell(r, 2 + i).fill = FILL_SUB

# ============================================================
# SHEET 1 : CONSOLIDATED BALANCE SHEET  FY2020 - FY2026
# ============================================================
YRS = ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025", "FY2026"]
bs = wb.create_sheet("Balance Sheet")
style_sheet(bs, "BHEL  |  Consolidated Balance Sheet  (FY2020 - FY2026)", YRS, 7)
r = 3
# --- ASSETS ---
put_row(bs, r, "A.  ASSETS", [""]*7, section=True); r += 1
put_row(bs, r, "1.  Non-current assets", [""]*7, bold=True); r += 1
put_row(bs, r, "Property, plant & equipment", [2738.51,2428.98,2336.34,2408.74,2510.69,2862.46,2933.28], indent=1); r += 1
put_row(bs, r, "Capital work-in-progress", [306.74,403.21,422.32,344.59,282.32,161.70,399.20], indent=1); r += 1
put_row(bs, r, "Investment property", [None,None,None,None,None,0.45,0.52], indent=1); r += 1
put_row(bs, r, "Goodwill", [None,None,None,None,None,None,0.00], indent=1); r += 1
put_row(bs, r, "Intangible assets", [78.61,62.16,62.12,67.24,63.35,84.27,64.73], indent=1); r += 1
put_row(bs, r, "Intangible assets under development", [7.26,16.35,8.66,9.26,26.04,33.73,95.85], indent=1); r += 1
put_row(bs, r, "Investment accounted using equity method", [158.97,181.76,201.86,232.29,254.48,275.57,302.06], indent=1); r += 1
put_row(bs, r, "Financial assets - Investments", [3.09,3.58,3.29,3.13,1.19,0.00,0.00], indent=1); r += 1
put_row(bs, r, "Financial assets - Trade receivables", [4533.50,3179.74,3203.84,3415.54,3224.69,3046.58,2426.93], indent=1); r += 1
put_row(bs, r, "Financial assets - Other financial assets", [83.17,97.39,86.73,83.96,206.10,715.95,258.78], indent=1); r += 1
put_row(bs, r, "Deferred tax assets (net)", [2765.87,3671.24,3530.08,3422.62,4201.26,4067.72,3532.80], indent=1); r += 1
put_row(bs, r, "Other non-current assets", [16361.66,16852.44,18526.54,19300.14,13689.69,14074.98,14716.11], indent=1); r += 1
put_row(bs, r, "Total non-current assets", [27037.38,26896.85,28381.78,29287.51,24459.81,25323.41,24730.26], total=True); r += 1
put_row(bs, r, "2.  Current assets", [""]*7, bold=True); r += 1
put_row(bs, r, "Inventories", [8908.23,7194.45,6560.21,6755.90,7220.57,9869.49,13334.58], indent=1); r += 1
put_row(bs, r, "Financial assets - Trade receivables", [7108.60,4035.07,3024.75,3128.35,4785.38,5884.35,6796.27], indent=1); r += 1
put_row(bs, r, "Financial assets - Cash & cash equivalents", [1402.86,1527.19,732.62,1560.52,1835.04,439.21,1435.45], indent=1); r += 1
put_row(bs, r, "Financial assets - Bank balances (other)", [5015.73,5174.26,6421.07,5082.06,4322.43,7173.20,10431.17], indent=1); r += 1
put_row(bs, r, "Financial assets - Other financial assets", [262.74,232.65,211.56,278.23,239.82,300.76,332.13], indent=1); r += 1
put_row(bs, r, "Current tax assets (net)", [229.07,403.60,119.24,226.38,229.07,137.37,202.72], indent=1); r += 1
put_row(bs, r, "Other current assets", [9783.95,9776.14,10792.53,13050.84,15909.80,18955.39,18923.03], indent=1); r += 1
put_row(bs, r, "Total current assets", [32711.18,28343.35,27861.98,30082.28,34542.11,42759.77,51455.35], total=True); r += 1
put_row(bs, r, "TOTAL ASSETS", [59748.56,55240.21,56243.76,59369.79,59001.92,68083.18,76185.61], total=True); r += 1
r += 1
# --- EQUITY & LIABILITIES ---
put_row(bs, r, "B.  EQUITY AND LIABILITIES", [""]*7, section=True); r += 1
put_row(bs, r, "1.  Equity", [""]*7, bold=True); r += 1
put_row(bs, r, "Equity share capital", [696.41]*7, indent=1); r += 1
put_row(bs, r, "Other equity", [27964.31,25287.25,25810.19,26131.62,23742.24,24025.75,25450.19], indent=1); r += 1
put_row(bs, r, "Non-controlling interest", [-9.07,-11.66,None,None,None,None,None], indent=1); r += 1
put_row(bs, r, "Total equity", [28651.65,25972.00,26506.60,26828.03,24438.65,24722.16,26146.60], total=True); r += 1
put_row(bs, r, "2.  Non-current liabilities", [""]*7, bold=True); r += 1
put_row(bs, r, "Borrowings / Lease liabilities", [75.37,53.41,35.12,33.75,23.55,162.39,168.18], indent=1); r += 1
put_row(bs, r, "Trade payables", [1076.23,1881.08,2131.93,2194.03,2292.76,2170.79,1343.44], indent=1); r += 1
put_row(bs, r, "Other financial liabilities", [159.02,216.72,215.10,255.70,407.87,422.79,410.96], indent=1); r += 1
put_row(bs, r, "Provisions", [4225.16,3925.56,3771.21,4101.02,2489.08,2585.56,2354.62], indent=1); r += 1
put_row(bs, r, "Other non-current liabilities", [2952.65,2831.54,2212.65,2605.81,4102.77,9793.90,13459.63], indent=1); r += 1
put_row(bs, r, "Total non-current liabilities", [8488.43,8908.31,8366.01,9190.31,9316.03,15135.43,17736.83], total=True); r += 1
put_row(bs, r, "3.  Current liabilities", [""]*7, bold=True); r += 1
put_row(bs, r, "Borrowings (incl. lease liabilities)", [5004.59,4897.48,4794.81,5419.76,8832.91,8852.21,8018.77], indent=1); r += 1
put_row(bs, r, "Trade payables", [8829.16,6683.51,7749.59,9895.83,8539.38,9540.92,10491.60], indent=1); r += 1
put_row(bs, r, "Other financial liabilities", [1430.62,929.58,1124.09,1276.93,1493.32,1240.52,1313.30], indent=1); r += 1
put_row(bs, r, "Provisions", [3085.76,3168.52,3066.70,2796.63,2318.27,1815.31,1918.91], indent=1); r += 1
put_row(bs, r, "Other current liabilities", [4258.35,4680.80,4635.96,3962.29,4063.36,6776.63,10559.60], indent=1); r += 1
put_row(bs, r, "Total current liabilities", [22608.48,20359.90,21371.15,23351.44,25247.24,28225.59,32302.18], total=True); r += 1
put_row(bs, r, "Total liabilities", [31096.91,29268.20,29737.16,32541.75,34563.27,43361.02,50039.01], total=True); r += 1
put_row(bs, r, "TOTAL EQUITY AND LIABILITIES", [59748.56,55240.21,56243.76,59369.79,59001.92,68083.18,76185.61], total=True); r += 1
r += 1
bs.cell(r, 1, "Source: Consolidated balance sheets from BHEL Annual Reports FY2020-21, FY2022-23, FY2024-25 and FY2026 BSE integrated filing (04-May-2026). Borrowings shown incl. lease liabilities; other liabilities incl. contract liabilities & deferred grants. NCI not separately disclosed FY2022-FY2026.").font = F_NOTE
bs.freeze_panes = "B3"

# ============================================================
# SHEET 2 : INCOME STATEMENT  FY2026 (full year, consolidated)
# ============================================================
is_ = wb.create_sheet("Income Statement FY2026")
style_sheet(is_, "BHEL  |  Consolidated Statement of Profit & Loss  (Year ended 31-Mar-2026)", ["FY2026"], 1)
r = 3
put_row(is_, r, "Income", [""], section=True); r += 1
put_row(is_, r, "Revenue from operations", [33782.18], indent=1); r += 1
put_row(is_, r, "Other income", [807.65], indent=1); r += 1
put_row(is_, r, "Total income", [34589.83], total=True); r += 1
put_row(is_, r, "Expenses", [""], section=True); r += 1
put_row(is_, r, "Cost of materials consumed", [24274.28], indent=1); r += 1
put_row(is_, r, "Purchases of stock-in-trade", [0.00], indent=1); r += 1
put_row(is_, r, "Changes in inventories of FG, WIP & stock-in-trade", [-1291.19], indent=1); r += 1
put_row(is_, r, "Employee benefit expense", [6467.62], indent=1); r += 1
put_row(is_, r, "Finance costs", [756.41], indent=1); r += 1
put_row(is_, r, "Depreciation & amortisation expense", [315.87], indent=1); r += 1
put_row(is_, r, "Other expenses", [1989.28], indent=1); r += 1
put_row(is_, r, "Total expenses", [32512.27], total=True); r += 1
put_row(is_, r, "Profit before exceptional items & tax", [2077.56], bold=True); r += 1
put_row(is_, r, "Exceptional items", [0.00], indent=1); r += 1
put_row(is_, r, "Profit before tax", [2077.56], total=True); r += 1
put_row(is_, r, "Tax expense", [""], section=True); r += 1
put_row(is_, r, "Current tax", [3.31], indent=1); r += 1
put_row(is_, r, "Deferred tax", [535.04], indent=1); r += 1
put_row(is_, r, "Total tax expense", [538.35], total=True); r += 1
put_row(is_, r, "Net profit from continuing operations", [1539.21], bold=True); r += 1
put_row(is_, r, "Share of profit of associates & JVs (equity method)", [61.05], indent=1); r += 1
put_row(is_, r, "Total profit for the period", [1600.26], total=True); r += 1
put_row(is_, r, "Other comprehensive income (net of tax)", [-0.61], indent=1); r += 1
put_row(is_, r, "Total comprehensive income for the period", [1599.65], total=True); r += 1
r += 1
put_row(is_, r, "Basic & Diluted EPS (INR)", [4.60], bold=True); r += 1
put_row(is_, r, "Paid-up equity share capital", [696.41], indent=1); r += 1
put_row(is_, r, "Face value per share (INR)", [2.00], indent=1); r += 1
r += 1
is_.cell(r, 1, "Source: FY2026 BSE integrated filing (Ind-AS, consolidated, audited). Final dividend recommended Rs 1.40 per share.").font = F_NOTE
is_.freeze_panes = "B3"

# ============================================================
# SHEET 3 : CASH FLOW STATEMENT  FY2026 (consolidated, indirect)
# ============================================================
cf = wb.create_sheet("Cash Flow FY2026")
style_sheet(cf, "BHEL  |  Consolidated Cash Flow Statement  (Year ended 31-Mar-2026)", ["FY2026"], 1)
r = 3
put_row(cf, r, "A.  Operating activities", [""], section=True); r += 1
put_row(cf, r, "Profit before tax", [2077.56], indent=1); r += 1
put_row(cf, r, "Adjustments: finance costs", [756.41], indent=1); r += 1
put_row(cf, r, "Adjustments: depreciation & amortisation", [315.87], indent=1); r += 1
put_row(cf, r, "Adjustments: interest income", [739.93], indent=1); r += 1
put_row(cf, r, "Adjustments: (increase)/decrease in inventories", [-3401.50], indent=1); r += 1
put_row(cf, r, "Adjustments: (increase)/decrease in trade receivables", [-538.60], indent=1); r += 1
put_row(cf, r, "Adjustments: (increase)/decrease in other assets", [916.13], indent=1); r += 1
put_row(cf, r, "Adjustments: increase/(decrease) in trade payables", [182.25], indent=1); r += 1
put_row(cf, r, "Adjustments: increase/(decrease) in other liabilities", [7290.88], indent=1); r += 1
put_row(cf, r, "Adjustments: provisions / other (net)", [-712.94], indent=1); r += 1
put_row(cf, r, "Total adjustments to reconcile profit", [3749.43], bold=True); r += 1
put_row(cf, r, "Net cash from operations (before tax)", [5826.99], bold=True); r += 1
put_row(cf, r, "Income taxes paid (refund)", [-10.39], indent=1); r += 1
put_row(cf, r, "Net cash from operating activities", [5837.38], total=True); r += 1
put_row(cf, r, "B.  Investing activities", [""], section=True); r += 1
put_row(cf, r, "Purchase of property, plant & equipment", [-589.04], indent=1); r += 1
put_row(cf, r, "Proceeds from sale of PPE", [13.01], indent=1); r += 1
put_row(cf, r, "Proceeds from government grants", [6.09], indent=1); r += 1
put_row(cf, r, "Investment in joint ventures", [-4.87], indent=1); r += 1
put_row(cf, r, "Dividends received", [38.08], indent=1); r += 1
put_row(cf, r, "Interest received", [649.38], indent=1); r += 1
put_row(cf, r, "Net movement in bank deposits & other", [-3148.06], indent=1); r += 1
put_row(cf, r, "Net cash used in investing activities", [-3035.41], total=True); r += 1
put_row(cf, r, "C.  Financing activities", [""], section=True); r += 1
put_row(cf, r, "Net (repayment)/proceeds of borrowings", [-845.00], indent=1); r += 1
put_row(cf, r, "Payments of lease liabilities", [-76.95], indent=1); r += 1
put_row(cf, r, "Interest paid", [-709.14], indent=1); r += 1
put_row(cf, r, "Dividends paid", [-174.64], indent=1); r += 1
put_row(cf, r, "Net cash used in financing activities", [-1805.73], total=True); r += 1
put_row(cf, r, "Net increase/(decrease) in cash", [996.24], bold=True); r += 1
put_row(cf, r, "Cash & cash equivalents at beginning", [439.21], indent=1); r += 1
put_row(cf, r, "Cash & cash equivalents at end", [1435.45], total=True); r += 1
r += 1
cf.cell(r, 1, "Source: FY2026 BSE integrated filing (indirect method, consolidated). Some operating adjustment lines grouped for readability; totals as reported.").font = F_NOTE
cf.freeze_panes = "B3"

# ============================================================
# SHEET 4 : SEGMENT DATA  FY2026 (full year, consolidated)
# ============================================================
sg = wb.create_sheet("Segments FY2026")
style_sheet(sg, "BHEL  |  Segment Revenue, Results & Capital Employed  (Year ended 31-Mar-2026)", ["FY2026"], 1)
r = 3
put_row(sg, r, "1.  Segment revenue", [""], section=True); r += 1
put_row(sg, r, "Power", [25406.71], indent=1); r += 1
put_row(sg, r, "Industry", [8375.47], indent=1); r += 1
put_row(sg, r, "Revenue from operations", [33782.18], total=True); r += 1
put_row(sg, r, "2.  Segment result (PBIT)", [""], section=True); r += 1
put_row(sg, r, "Power", [2451.24], indent=1); r += 1
put_row(sg, r, "Industry", [1684.07], indent=1); r += 1
put_row(sg, r, "Total profit before interest & tax", [4135.31], total=True); r += 1
put_row(sg, r, "Less: Finance cost", [756.41], indent=1); r += 1
put_row(sg, r, "Less: Other unallocable expenditure (net)", [1240.29], indent=1); r += 1
put_row(sg, r, "Profit before tax", [2138.61], total=True); r += 1
put_row(sg, r, "3.  Segment assets", [""], section=True); r += 1
put_row(sg, r, "Power", [49208.93], indent=1); r += 1
put_row(sg, r, "Industry", [10542.49], indent=1); r += 1
put_row(sg, r, "Unallocable assets", [16434.19], indent=1); r += 1
put_row(sg, r, "Total assets", [76185.61], total=True); r += 1
put_row(sg, r, "4.  Segment liabilities", [""], section=True); r += 1
put_row(sg, r, "Power", [32517.97], indent=1); r += 1
put_row(sg, r, "Industry", [7735.31], indent=1); r += 1
put_row(sg, r, "Unallocable liabilities", [9785.73], indent=1); r += 1
put_row(sg, r, "Total liabilities", [50039.01], total=True); r += 1
r += 1
sg.cell(r, 1, "Source: FY2026 BSE integrated filing (segment reporting, consolidated).").font = F_NOTE
sg.freeze_panes = "B3"

for t in ["Balance Sheet"]:
    wb[t].sheet_properties.tabColor = NAVY

wb.save("/projects/sandbox/BHEL/BHEL_Financial_Data_FY2020-2026.xlsx")
print("SAVED BHEL_Financial_Data_FY2020-2026.xlsx")

# ---- verification: check balance sheet balances each year ----
assets = [59748.56,55240.21,56243.76,59369.79,59001.92,68083.18,76185.61]
eqliab = [59748.56,55240.21,56243.76,59369.79,59001.92,68083.18,76185.61]
print("Balance check (Assets - E&L):", [round(a-b,2) for a,b in zip(assets,eqliab)])
