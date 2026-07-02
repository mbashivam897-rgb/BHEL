# -*- coding: utf-8 -*-
"""
Institutional-grade integrated financial model for Bharat Heavy Electricals Ltd (BHEL)
Historical: FY2020-FY2026 (consolidated, sourced from BSE filings via Screener.in)
Forecast:   FY2027-FY2033 (order-book-driven, fully linked 3-statement model)
Income statement presented in gross-profit format:
  Sales - COGS = Gross Profit - Operating Expenses = EBITDA - D&A = EBIT
  + Other Income - Interest = PBT - Tax + Adjustments = PAT
Conventions: BLACK=input, BLUE=formula, GREEN=cross-sheet link, gold fill=editable input.
All figures INR Crore unless stated; per-share in INR.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
# AUDITED historical data (FY2020-FY2026) extracted & verified from the annual reports
from audited_full import IS as AIS, BS as ABS, CF as ACF
# Contract Accounting Schedule (Ind AS 115) - historical data, driver analysis
# (coefficient-of-variation-based selection) and forecast glide paths
import contract_schedule as CS

wb = Workbook(); wb.remove(wb.active)

HIST = ['FY2020','FY2021','FY2022','FY2023','FY2024','FY2025','FY2026']
FCST = ['FY2027','FY2028','FY2029','FY2030','FY2031','FY2032','FY2033']
YEARS = HIST + FCST
HCOLS = list(range(2, 9)); FCOLS = list(range(9, 16)); ACOLS = HCOLS + FCOLS
def CL(c): return get_column_letter(c)

NAVY='1F3864'; LTBL='D9E1F2'; INPUTF='FFF2CC'; GREEN_F='E2EFDA'
F_BLACK=Font(name='Calibri',size=10,color='000000')
F_BLUE =Font(name='Calibri',size=10,color='0000CC')
F_GREEN=Font(name='Calibri',size=10,color='008000')
F_LBL  =Font(name='Calibri',size=10,color='000000')
F_LBLB =Font(name='Calibri',size=10,color='000000',bold=True)
F_HDRW =Font(name='Calibri',size=10,color='FFFFFF',bold=True)
F_TITLE=Font(name='Calibri',size=16,color='FFFFFF',bold=True)
F_SUB  =Font(name='Calibri',size=11,color='FFFFFF')
F_SECT =Font(name='Calibri',size=11,color=NAVY,bold=True)
F_NOTE =Font(name='Calibri',size=8, color='808080',italic=True)
F_ITAL =Font(name='Calibri',size=9, color='595959',italic=True)
FILL_HDR=PatternFill('solid',fgColor=NAVY); FILL_SUB=PatternFill('solid',fgColor=LTBL)
FILL_IN =PatternFill('solid',fgColor=INPUTF); FILL_TOT=PatternFill('solid',fgColor='DDEBF7')
FILL_OK =PatternFill('solid',fgColor=GREEN_F)
thin=Side(style='thin',color='BFBFBF')
B_ALL=Border(left=thin,right=thin,top=thin,bottom=thin)
B_TOP=Border(top=Side(style='thin',color='000000'))
B_TB =Border(top=Side(style='thin',color='000000'),bottom=Side(style='double',color='000000'))
FMT_CR='#,##0;(#,##0)'; FMT_CR1='#,##0.0;(#,##0.0)'; FMT_PCT='0.0%'
FMT_PS='0.00'; FMT_X='0.0"x"'; FMT_DAYS='0'
CEN=Alignment(horizontal='center',vertical='center'); RGT=Alignment(horizontal='right')

REG={}
def reg(ws,key,row): REG.setdefault(ws.title,{})[key]=row
def XR(title,key,c): return "'%s'!%s%d"%(title,CL(c),REG[title][key])

def put(ws,r,c,v,font=F_BLACK,fmt=None,fill=None,align=None,border=None,bold=False):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name=font.name,size=font.size,color=font.color.rgb if font.color else '000000',bold=(bold or font.bold),italic=font.italic)
    if fmt: cell.number_format=fmt
    if fill: cell.fill=fill
    if align: cell.alignment=align
    if border: cell.border=border
    return cell
def section(ws,r,text,span=15):
    for c in range(1,span+1): ws.cell(row=r,column=c).fill=FILL_HDR
    put(ws,r,1,text,font=F_HDRW,fill=FILL_HDR); return r
def ah_label(ws,r,text,bold=False,indent=0,italic=False):
    f=F_LBLB if bold else (F_ITAL if italic else F_LBL)
    put(ws,r,1,('   '*indent)+text,font=f)
def std_widths(ws,first=46):
    ws.column_dimensions['A'].width=first
    for c in ACOLS: ws.column_dimensions[CL(c)].width=11.5
def freeze(ws,cell='B5'): ws.freeze_panes=cell
def newsheet(title):
    ws=wb.create_sheet(title); ws.sheet_view.showGridLines=False; std_widths(ws)
    put(ws,1,1,title.upper(),font=F_HDRW,fill=FILL_HDR)
    for c in range(2,16): ws.cell(row=1,column=c).fill=FILL_HDR
    return ws
def yhdr(ws,r,a='INR Crore'):
    put(ws,r,1,a,font=F_ITAL,fill=FILL_SUB)
    for i,c in enumerate(ACOLS):
        put(ws,r,c,YEARS[i],font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
def lnk(c,key,sheet):
    return "%s!%s%d"%("'"+sheet+"'",CL(c),REG[sheet][key])
print("Framework loaded.")


# ===== 1. COVER PAGE =====
cv=wb.create_sheet('Cover Page'); cv.sheet_view.showGridLines=False
cv.column_dimensions['A'].width=3
for c in ['B','C','D','E','F','G']: cv.column_dimensions[c].width=24
cv.column_dimensions['B'].width=32
for r in range(1,4):
    for c in range(1,8): cv.cell(row=r,column=c).fill=FILL_HDR
put(cv,2,2,'BHARAT HEAVY ELECTRICALS LIMITED (BHEL)',font=F_TITLE,fill=FILL_HDR)
put(cv,3,2,'Integrated Financial Model & Equity Valuation  |  NSE: BHEL  |  BSE: 500103',font=F_SUB,fill=FILL_HDR)
meta=[('Company','Bharat Heavy Electricals Ltd (Maharatna CPSE, Govt. of India)'),
 ('Sector','Capital Goods / Heavy Electrical Equipment'),
 ('Reporting basis','Consolidated, Indian Accounting Standards (Ind-AS)'),
 ('Currency / Units','INR Crore (per-share in INR)'),
 ('Historical period','FY2020 - FY2026 (reported)'),
 ('Forecast period','FY2027 - FY2033 (7 years, analyst projections)'),
 ('Valuation date','27-Jun-2026'),('Last traded price','INR 403 (25-Jun-2026)'),
 ('Shares outstanding','~348.0 crore (Face value INR 2)'),('Promoter (Govt. of India)','58.17%')]
r=6
for k,v in meta: put(cv,r,2,k,font=F_LBLB); put(cv,r,3,v); r+=1
put(cv,r+1,2,'Income statement format',font=F_SECT); r+=2
put(cv,r,2,'Sales - COGS = Gross Profit; less Operating Expenses (Employee, Selling & admin) = EBITDA;',font=F_ITAL); r+=1
put(cv,r,2,'less D&A = EBIT; +Other Income -Interest = PBT; less Tax +/- Adjustments = PAT.',font=F_ITAL); r+=2
put(cv,r,2,'Formatting legend',font=F_SECT); r+=1
put(cv,r,2,'Black',font=F_BLACK); put(cv,r,3,'Hardcoded input / reported data')
put(cv,r+1,2,'Blue',font=F_BLUE); put(cv,r+1,3,'In-sheet formula')
put(cv,r+2,2,'Green',font=F_GREEN); put(cv,r+2,3,'Link to another worksheet')
put(cv,r+3,2,'Gold fill',font=F_BLACK,fill=FILL_IN); put(cv,r+3,3,'Editable assumption / input')
r+=5
put(cv,r,2,'Data sources & caveats',font=F_SECT); r+=1
for s in ['Historical financials FY2020-FY2026: AUDITED consolidated statements from BHEL Annual Reports',
 '  (FY20-21, FY22-23, FY24-25) and the FY2026 BSE integrated filing - full P&L, balance sheet & cash flow.',
 'By-nature P&L costs are mapped to the gross-profit format; each year reproduces audited PBT & PAT exactly.',
 'Order book, inflow & segment data: BHEL results releases & Directors\' Report (standalone basis).',
 'Peer multiples on Relative Valuation are indicative; refresh with live data before use.']:
    put(cv,r,2,s,font=F_ITAL); r+=1
freeze(cv,'A1')

# ===== 2. ASSUMPTIONS =====
ASM='Assumptions'; asm=newsheet(ASM)
from openpyxl.worksheet.datavalidation import DataValidation
put(asm,2,2,'Scenario Switch',font=F_LBLB)
sw=put(asm,2,4,'Base',font=F_LBLB,fill=FILL_IN,align=CEN,border=B_ALL)
_dv=DataValidation(type='list',formula1='"Base,Best,Bear"',allow_blank=False); asm.add_data_validation(_dv); _dv.add('D2')
put(asm,2,6,'<- toggles the VARIABLE drivers below; revenue, margins, cash flow & valuation recalculate',font=F_NOTE)
SW='$D$2'
put(asm,3,1,'A.  VARIABLE ASSUMPTIONS  (Base / Best / Bear - selected by Scenario Switch)',font=F_SECT)
put(asm,3,9,'Sales growth is DERIVED (Revenue Build-up): Revenue = Execution rate x Opening order book; Closing OB = Opening + Inflow - Revenue.',font=F_NOTE)
yhdr(asm,4,'ratios / INR Crore')
rowA=5
def aset(key,label,vals7,fmt,note='',indent=0):
    global rowA; r=rowA; ah_label(asm,r,label,indent=indent)
    for j,c in enumerate(FCOLS): put(asm,r,c,vals7[j],font=F_BLACK,fmt=fmt,fill=FILL_IN,align=RGT,border=B_ALL)
    if note: put(asm,r,16,note,font=F_NOTE)
    reg(asm,key,r); rowA+=1; return r
def vaset(key,label,base,best,bear,fmt,note=''):
    global rowA; r=rowA; ah_label(asm,r,label,bold=True)
    for j,c in enumerate(FCOLS):
        f='=IF(%s="Best",%s%d,IF(%s="Bear",%s%d,%s%d))'%(SW,CL(c),r+2,SW,CL(c),r+3,CL(c),r+1)
        put(asm,r,c,f,font=F_BLUE,fmt=fmt,fill=FILL_TOT,align=RGT,border=B_ALL)
    if note: put(asm,r,16,note,font=F_NOTE)
    for jj,(nm,vals) in enumerate([('Base',base),('Best',best),('Bear',bear)]):
        r2=r+1+jj; ah_label(asm,r2,'   '+nm,italic=True)
        for j,c in enumerate(FCOLS): put(asm,r2,c,vals[j],font=F_BLACK,fmt=fmt,fill=FILL_IN,align=RGT,border=B_ALL)
    reg(asm,key,r); rowA=r+5; return r
vaset('order_inflow','New order inflow (INR Cr)  [revenue driver]',
 [78000,82000,86000,90000,92000,94000,96000],[90000,95000,100000,105000,108000,110000,112000],[62000,64000,66000,68000,70000,72000,74000],FMT_CR,
 'Demand driver: thermal/T&D/defence ordering. Best=accelerated power capex cycle; Bear=delayed ordering.')
vaset('exec_rate','Execution rate (revenue / opening order book)  [revenue driver]',
 [0.165,0.170,0.175,0.180,0.185,0.190,0.195],[0.175,0.182,0.190,0.197,0.205,0.210,0.215],[0.150,0.152,0.155,0.158,0.160,0.162,0.165],FMT_PCT,
 'Order-book-to-revenue conversion; rises with capacity utilisation & supply-chain normalisation.')
vaset('ebitda_margin','EBITDA margin (% revenue)',
 [0.075,0.085,0.095,0.105,0.110,0.115,0.120],[0.090,0.100,0.110,0.120,0.125,0.130,0.135],[0.060,0.065,0.072,0.080,0.085,0.090,0.095],FMT_PCT,
 'Operating leverage on a largely fixed cost base; commodity (steel/copper) & pricing risk.')
put(asm,rowA,1,'B.  FIXED ASSUMPTIONS  (structural / policy - constant across scenarios)',font=F_SECT); rowA+=1
yhdr(asm,rowA,'ratios / INR Crore'); rowA+=1
aset('consol_uplift','Consolidation uplift (consol vs standalone rev) %',[0.02]*7,FMT_PCT,note='Subsidiary/JV contribution to revenue; historical 0-6% (FY26 ~0%).')
aset('oth_inc_pct','Other income (% revenue)',[0.024]*7,FMT_PCT)
aset('emp_pct','Employee cost (% revenue)',[0.185,0.178,0.170,0.163,0.155,0.148,0.140],FMT_PCT)
aset('subc_pct','Subcontracting & erection (% revenue, in COGS)',[0.085]*7,FMT_PCT)
aset('sga_pct','Selling & admin (% revenue)',[0.075]*7,FMT_PCT)
aset('steel_pct','  COGS: Steel (% revenue)',[0.12]*7,FMT_PCT)
aset('copper_pct','  COGS: Copper (% revenue)',[0.05]*7,FMT_PCT)
aset('elec_pct','  COGS: Electrical & electronic components (% rev)',[0.10]*7,FMT_PCT)
aset('imported_pct','  COGS: Imported components (% revenue)',[0.08]*7,FMT_PCT)
aset('selling_pct','  OpEx: Selling & distribution (% revenue)',[0.025]*7,FMT_PCT)
aset('rnd_pct','  OpEx: Research & development (% revenue)',[0.025]*7,FMT_PCT)
aset('power_mix','Revenue mix: Power segment %',[0.75]*7,FMT_PCT)
aset('industry_mix','Revenue mix: Industry segment %',[0.22]*7,FMT_PCT)
aset('capex','Capital expenditure',[600,700,750,800,850,900,950],FMT_CR)
aset('dep_rate','Depreciation rate (% opening net block)',[0.095]*7,FMT_PCT)
aset('recv_days','Trade receivable days (vs revenue)',[74,73,72,71,70,69,68],FMT_DAYS)
aset('inv_days','Inventory days (vs COGS)',[210,207,204,201,198,195,192],FMT_DAYS)
aset('pay_days','Trade payable days (vs COGS)',[167,166,165,164,163,162,160],FMT_DAYS)
aset('ca_to_rev','Contract assets (total) / Revenue  [see Contract Accounting Schedule]',CS.CA_TO_REV_FCST,FMT_PCT,
     note='CV-selected driver (0.139, lowest); CA = unbilled revenue, scales with revenue recognised.')
aset('cl_to_ob','Contract liabilities (total) / Closing order book  [see Contract Accounting Schedule]',CS.CL_TO_OB_FCST,FMT_PCT,
     note='CV-selected driver (0.178, lowest order-book-based); CL = customer advances, scales with unexecuted backlog.')
aset('ca_cur_split','Contract assets - current % of total',[CS.CA_CUR_SPLIT_RECENT]*7,FMT_PCT,
     note='Recent 3yr (FY24-26) average; structural shift post-FY23 per Q4FY24 concall WC-improvement guidance.')
aset('cl_cur_split','Contract liabilities - current % of total',[CS.CL_CUR_SPLIT_RECENT]*7,FMT_PCT,
     note='Recent 3yr (FY24-26) average; structural shift post-FY23 per Q4FY24 concall WC-improvement guidance.')
aset('oca_g','Other current/non-curr assets growth %',[0.03]*7,FMT_PCT)
aset('ol_g','Provisions & other liabilities growth %',[0.03]*7,FMT_PCT)
aset('cwip','Capital work-in-progress (closing)',[400]*7,FMT_CR)
aset('investments','Investments (closing)',[310,320,330,340,350,360,370],FMT_CR)
aset('tax_rate','Effective tax rate',[0.25]*7,FMT_PCT)
aset('div_payout','Dividend payout (% PAT)',[0.30]*7,FMT_PCT)
aset('new_borrow','New borrowings drawn',[0]*7,FMT_CR)
aset('debt_repay','Debt repayment',[0]*7,FMT_CR)
aset('int_rate','Interest rate on opening debt',[0.085]*7,FMT_PCT)
aset('lt_debt_pct','LT borrowings (% of total debt)',[0.02]*7,FMT_PCT)
aset('other_nca','Other non-current assets - NON-CONTRACT residual (DTA, legacy receivables, other)',[6738]*7,FMT_CR,
     note='FY26 audited residual (other_nca total less non-current contract assets); held flat - no clear trend FY24-26.')
aset('other_ncl','Other non-current liabilities - NON-CONTRACT residual (LT provisions, other)',[4155]*7,FMT_CR,
     note='FY26 audited residual (other_ncl total less non-current contract liabilities); held flat - no clear trend FY24-26.')
wr=rowA+1; put(asm,wr,1,'WACC & valuation inputs',font=F_SECT); wr+=1
def asingle(key,label,value,fmt,note=''):
    global wr; ah_label(asm,wr,label)
    put(asm,wr,3,value,font=F_BLACK,fmt=fmt,fill=FILL_IN,align=RGT,border=B_ALL)
    if note: put(asm,wr,5,note,font=F_NOTE)
    reg(asm,key,wr); wr+=1
asingle('rf','Risk-free rate (10Y India G-Sec)',0.068,FMT_PCT)
asingle('beta','Levered beta',1.25,'0.00')
asingle('erp','Equity risk premium',0.065,FMT_PCT)
asingle('kd','Pre-tax cost of debt',0.085,FMT_PCT)
asingle('tax_w','Marginal tax rate (WACC)',0.25,FMT_PCT)
asingle('wd','Target weight of debt',0.15,FMT_PCT)
asingle('we','Target weight of equity',0.85,FMT_PCT)
asingle('term_g','Terminal growth rate',0.045,FMT_PCT)
ah_label(asm,wr,'Cost of equity (Ke = Rf + B x ERP)',bold=True)
put(asm,wr,3,"=%s+%s*%s"%(XR(ASM,'rf',3),XR(ASM,'beta',3),XR(ASM,'erp',3)),font=F_BLUE,fmt=FMT_PCT,align=RGT,border=B_ALL); reg(asm,'ke',wr); wr+=1
ah_label(asm,wr,'WACC (We*Ke + Wd*Kd*(1-t))',bold=True)
put(asm,wr,3,"=%s*%s+%s*%s*(1-%s)"%(XR(ASM,'we',3),XR(ASM,'ke',3),XR(ASM,'wd',3),XR(ASM,'kd',3),XR(ASM,'tax_w',3)),font=F_BLUE,fmt=FMT_PCT,fill=FILL_TOT,align=RGT,border=B_ALL); reg(asm,'wacc',wr); wr+=1
asingle('shares','Shares outstanding (crore)',348.05,FMT_CR1)
asingle('price','Current market price (INR)',403.0,FMT_PS)
asingle('fv','Face value (INR)',2.0,FMT_PS)
freeze(asm,'B5')
print("Cover + Assumptions built.")


# ===== 3. HISTORICAL FINANCIAL STATEMENTS (consolidated, reported) =====
HFS='Historical Financial Statements'; hf=newsheet(HFS)
def hrow(r,label,vals,fmt=FMT_CR,font=F_BLACK,bold=False,indent=0,fill=None,key=None,border=None):
    ah_label(hf,r,label,bold=bold,indent=indent)
    for j,c in enumerate(HCOLS): put(hf,r,c,vals[j],font=font,fmt=fmt,fill=fill,align=RGT,border=border)
    if key: reg(hf,key,r)
def hform(r,label,fn,fmt=FMT_CR,bold=False,indent=0,fill=None,key=None,border=None,italic=False):
    ah_label(hf,r,label,bold=bold,indent=indent,italic=italic)
    for c in HCOLS: put(hf,r,c,fn(c),font=F_BLUE,fmt=fmt,fill=fill,align=RGT,border=border)
    if key: reg(hf,key,r)
section(hf,3,'A.  INCOME STATEMENT'); yhdr(hf,4,'Particulars')
# AUDITED consolidated P&L FY2020-FY2026 (annual reports), mapped to gross-profit format
rev=AIS['rev']
emp=AIS['emp']
sga=AIS['sga']
cogs=AIS['cogs']
oi=AIS['oi']; dep=AIS['dep']
fin=AIS['fin']; tax=AIS['tax']
div=AIS['div']
hrow(5,'Sales',rev,key='rev',bold=True)
hrow(6,'Total COGS',cogs,key='cogs',font=F_BLACK)
hform(7,'Gross Profit',lambda c:"=%s5-%s6"%(CL(c),CL(c)),key='gross',bold=True,border=B_TOP)
ah_label(hf,8,'Operating Expenses')
hrow(9,'Employee Cost',emp,key='emp',indent=1)
hrow(10,'Selling and admin',sga,key='sga',indent=1)
hform(11,'Total Operating Expenses',lambda c:"=%s9+%s10"%(CL(c),CL(c)),key='totopex',bold=True,border=B_TOP)
hform(12,'EBITDA',lambda c:"=%s7-%s11"%(CL(c),CL(c)),key='ebitda',bold=True,fill=FILL_TOT,border=B_TOP)
hrow(13,'Depreciation & Amortization',dep,key='dep')
hform(14,'EBIT',lambda c:"=%s12-%s13"%(CL(c),CL(c)),key='ebit',bold=True)
hrow(15,'Other Income',oi,key='oi')
hrow(16,'Interest Paid',fin,key='fin')
hform(17,'Profit Before Tax (PBT)',lambda c:"=%s14+%s15-%s16"%(CL(c),CL(c),CL(c)),key='pbt',bold=True,border=B_TOP)
hrow(18,'Tax',tax,key='tax')
hrow(19,'Adjustments (share of profit of JV/associates)',AIS['adj'],key='adj')
hform(20,'Profit After Tax (PAT)',lambda c:"=%s17-%s18+%s19"%(CL(c),CL(c),CL(c)),key='pat',bold=True,fill=FILL_TOT,border=B_TB)
hform(21,'EPS (INR)',lambda c:"=%s20/%s"%(CL(c),XR(ASM,'shares',3)),fmt=FMT_PS,key='eps')
hrow(22,'Dividend',div,key='div')
hform(23,'Memo: total operating cost (COGS+OpEx)',lambda c:"=%s6+%s11"%(CL(c),CL(c)),key='totcost',italic=True)
put(hf,24,1,'AUDITED consolidated figures (Annual Reports FY20-21/FY22-23/FY24-25 + FY26 BSE filing). By-nature costs mapped to gross-profit format: COGS = materials + inventory change; Selling/admin = other expenses (+ FX & provisions in FY20/21); Adjustments = share of JV/associate profit. Each year reproduces audited PBT & PAT exactly.',font=F_NOTE)
# Balance sheet
section(hf,26,'B.  BALANCE SHEET'); yhdr(hf,27,'Particulars')
# AUDITED consolidated balance sheet FY2020-FY2026 (ties exactly, no plug).
ppe=ABS['ppe']; cwip=ABS['cwip']
invv=ABS['invst']
otha=[ABS['recv'][i]+ABS['invy'][i]+ABS['oca'][i]+ABS['cash'][i]+ABS['onca'][i] for i in range(7)]
eqc=ABS['eqc']; res=ABS['res']
bor=[ABS['ltbor'][i]+ABS['stbor'][i] for i in range(7)]
othl=[ABS['oncl'][i]+ABS['pay'][i]+ABS['ocl'][i] for i in range(7)]
hrow(28,'Net property, plant & equipment',ppe,key='ppe')
hrow(29,'Capital work-in-progress',cwip,key='cwip')
hrow(30,'Investments',invv,key='inv')
hrow(31,'Other assets (inventory, receivables, cash, contract & other)',otha,key='otha')
hform(32,'TOTAL ASSETS',lambda c:"=SUM(%s28:%s31)"%(CL(c),CL(c)),key='ta',bold=True,fill=FILL_TOT,border=B_TB)
hrow(33,'Equity share capital',eqc,key='eqc')
hrow(34,'Other equity (reserves & surplus)',res,key='res')
hrow(35,'Borrowings (incl. lease liabilities)',bor,key='bor')
hrow(36,'Other liabilities (payables, advances, provisions)',othl,key='othl')
hform(37,'TOTAL EQUITY & LIABILITIES',lambda c:"=SUM(%s33:%s36)"%(CL(c),CL(c)),key='tle',bold=True,fill=FILL_TOT,border=B_TB)
hform(38,'Balance check (TA - TE&L)',lambda c:"=%s32-%s37"%(CL(c),CL(c)),key='bschk')
# Cash flow
section(hf,40,'C.  CASH FLOW STATEMENT (reported summary)'); yhdr(hf,41,'Particulars')
hrow(42,'Cash flow from operating activities',ACF['cfo'],key='cfo')
hrow(43,'Cash flow from investing activities',ACF['cfi'],key='cfi')
hrow(44,'Cash flow from financing activities',ACF['cff'],key='cff')
hform(45,'Net increase/(decrease) in cash',lambda c:"=SUM(%s42:%s44)"%(CL(c),CL(c)),key='netcf',bold=True,fill=FILL_TOT,border=B_TB)
freeze(hf,'B5')
print("Historical Financial Statements (gross-profit IS) built.")


# ===== 4. HISTORICAL RATIO ANALYSIS =====
hr=newsheet('Historical Ratio Analysis'); yhdr(hr,3)
def rrow(r,label,fn,fmt,bold=False,start=2):
    ah_label(hr,r,label,bold=bold)
    for c in HCOLS:
        if c<start: continue
        put(hr,r,c,fn(c),font=F_GREEN,fmt=fmt,align=RGT)
section(hr,4,'Growth & profitability')
rrow(5,'Revenue growth %',lambda c:"=%s/%s-1"%(lnk(c,'rev',HFS),lnk(c-1,'rev',HFS)),FMT_PCT,start=3)
rrow(6,'Gross margin %',lambda c:"=%s/%s"%(lnk(c,'gross',HFS),lnk(c,'rev',HFS)),FMT_PCT)
rrow(7,'EBITDA margin %',lambda c:"=%s/%s"%(lnk(c,'ebitda',HFS),lnk(c,'rev',HFS)),FMT_PCT)
rrow(8,'EBIT margin %',lambda c:"=%s/%s"%(lnk(c,'ebit',HFS),lnk(c,'rev',HFS)),FMT_PCT)
rrow(9,'PAT margin %',lambda c:"=%s/%s"%(lnk(c,'pat',HFS),lnk(c,'rev',HFS)),FMT_PCT)
section(hr,11,'Returns')
rrow(12,'ROE %',lambda c:"=%s/((%s+%s+%s+%s)/2)"%(lnk(c,'pat',HFS),lnk(c,'eqc',HFS),lnk(c,'res',HFS),lnk(c-1,'eqc',HFS),lnk(c-1,'res',HFS)),FMT_PCT,start=3)
rrow(13,'ROCE %',lambda c:"=%s/(%s+%s+%s)"%(lnk(c,'ebit',HFS),lnk(c,'eqc',HFS),lnk(c,'res',HFS),lnk(c,'bor',HFS)),FMT_PCT)
rrow(14,'ROA %',lambda c:"=%s/%s"%(lnk(c,'pat',HFS),lnk(c,'ta',HFS)),FMT_PCT)
section(hr,16,'Leverage & liquidity')
rrow(17,'Debt / equity (x)',lambda c:"=%s/(%s+%s)"%(lnk(c,'bor',HFS),lnk(c,'eqc',HFS),lnk(c,'res',HFS)),FMT_X)
rrow(18,'Interest coverage (x)',lambda c:"=%s/%s"%(lnk(c,'ebit',HFS),lnk(c,'fin',HFS)),FMT_X)
rrow(19,'Asset turnover (x)',lambda c:"=%s/%s"%(lnk(c,'rev',HFS),lnk(c,'ta',HFS)),FMT_X)
section(hr,21,'Working capital (reported day ratios)')
def drow_h(r,label,vals):
    ah_label(hr,r,label)
    for j,c in enumerate(HCOLS): put(hr,r,c,vals[j],font=F_BLACK,fmt=FMT_DAYS,align=RGT)
drow_h(22,'Debtor days (vs revenue)',[121,85,52,49,73,76,73])
drow_h(23,'Inventory days (vs COGS)',[232,221,162,151,157,191,212])
drow_h(24,'Payable days (vs COGS)',[230,205,191,222,185,185,167])
drow_h(25,'Cash conversion cycle (days)',[123,101,23,-22,45,82,118])
freeze(hr,'B4')

# ===== 5. OPERATIONAL DRIVERS =====
od=newsheet('Operational Drivers')
put(od,2,1,'Order data on standalone basis. FY20-22 figures are estimates (flagged). The "Adjustments" line captures order cancellations, de-scoping, FX revaluation, price/scope revisions and tax-basis differences so the roll ties to the reported closing order book.',font=F_NOTE)
yhdr(od,4)
def odrow(r,label,vals,fmt=FMT_CR,font=F_BLACK,bold=False,key=None,note=''):
    ah_label(od,r,label,bold=bold)
    for j,c in enumerate(HCOLS): put(od,r,c,vals[j],font=font,fmt=fmt,align=RGT)
    if key: reg(od,key,r)
    if note: put(od,r,16,note,font=F_NOTE)
def odf(r,label,fn,fmt,key=None):
    ah_label(od,r,label)
    for c in HCOLS: put(od,r,c,fn(c),font=F_BLUE,fmt=fmt,align=RGT)
    if key: reg(od,key,r)
section(od,5,'Order book dynamics (INR Crore, standalone)')
odrow(6,'Opening order book',[115000,110000,102000,98000,91336,131598,196328],key='ob_open',note='FY20-22 est.')
odrow(7,'Add: new order inflow',[18000,15000,22000,23548,63000,92535,75000],key='ob_inflow',note='some est.')
odrow(8,'Less: revenue executed (standalone)',[-20495,-16296,-20153,-22136,-22921,-27355,-33782],key='ob_exec',note='standalone rev')
odf(9,'+/- Adjustments (cancellations, FX, scope, taxes)',lambda c:"=%s10-%s6-%s7-%s8"%(CL(c),CL(c),CL(c),CL(c)),FMT_CR,key='ob_adj')
odrow(10,'Closing order book (reported)',[110000,102000,98000,91336,131598,196328,240000],key='ob_close',bold=True)
odf(11,'  Roll check (open+inflow-rev+adj-close)',lambda c:"=%s6+%s7+%s8+%s9-%s10"%(CL(c),CL(c),CL(c),CL(c),CL(c)),FMT_CR)
odf(12,'Book-to-bill (closing OB / revenue, x)',lambda c:"=%s10/-%s8"%(CL(c),CL(c)),FMT_X)
odf(13,'Execution rate (revenue / opening OB)',lambda c:"=-%s8/%s6"%(CL(c),CL(c)),FMT_PCT)
section(od,15,'Segment revenue mix (INR Crore, est.)')
odrow(16,'Power segment',[16800,13500,16500,18100,18600,20937,25407],key='seg_pow')
odrow(17,'Industry segment',[3900,3200,4000,4600,4700,6400,7400],key='seg_ind')
odrow(18,'Others / exports / renewables',[763,609,711,665,593,1002,975],key='seg_oth')
odf(19,'Total (check vs revenue)',lambda c:"=SUM(%s16:%s18)"%(CL(c),CL(c)),FMT_CR)
odf(20,'Power as % of revenue',lambda c:"=%s16/%s"%(CL(c),lnk(c,'rev',HFS)),FMT_PCT)
section(od,22,'Other KPIs')
odrow(23,'Approx. workforce (number)',[33500,31500,29800,28500,27000,26000,25500],fmt='#,##0',note='approx.')
odrow(24,'R&D expenditure',[1180,1090,1050,1100,1150,1300,1500],note='approx.')
odf(25,'R&D as % of revenue',lambda c:"=%s24/%s"%(CL(c),lnk(c,'rev',HFS)),FMT_PCT)
odrow(26,'Capacity utilisation (mfg.)',[0.55,0.42,0.50,0.55,0.58,0.65,0.72],fmt=FMT_PCT,note='approx.')
freeze(od,'B5')
print("Historical Ratios + Operational Drivers built.")


# ===== helper for forecast schedule rows =====
def frow(ws,r,label,hist=None,fcst=None,fmt=FMT_CR,bold=False,indent=0,fill=None,key=None,
         histfont=F_GREEN,fcstfont=F_BLUE,seed=None,seedcol=8):
    ah_label(ws,r,label,bold=bold,indent=indent)
    if hist is not None:
        for c in HCOLS:
            v=hist(c)
            if v is None: continue
            put(ws,r,c,v,font=histfont,fmt=fmt,fill=fill,align=RGT)
    if seed is not None:
        put(ws,r,seedcol,seed,font=F_BLACK,fmt=fmt,fill=FILL_IN,align=RGT)
    if fcst is not None:
        for c in FCOLS: put(ws,r,c,fcst(c),font=fcstfont,fmt=fmt,fill=fill,align=RGT)
    if key: reg(ws,key,r)

# ===== 6. REVENUE BUILD-UP =====
RB='Revenue Build-up'; rb=newsheet(RB)
put(rb,2,1,'Order book rolled on STANDALONE basis: Closing OB = Opening + Inflow - Standalone revenue + Adjustments (cancellations/FX/scope/taxes). Consolidated revenue (P&L) is bridged from standalone via a consolidation uplift.',font=F_NOTE)
yhdr(rb,4); section(rb,5,'A.  Order book roll-forward  (standalone basis, as disclosed)')
obc=REG['Operational Drivers']['ob_close']; obo=REG['Operational Drivers']['ob_open']; obi=REG['Operational Drivers']['ob_inflow']; obx=REG['Operational Drivers']['ob_exec']; obadj=REG['Operational Drivers']['ob_adj']
frow(rb,6,'Opening order book (standalone)',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),obo),
     fcst=lambda c:("='Operational Drivers'!H%d"%obc) if c==9 else ("=%s10"%CL(c-1)),key='open_ob',bold=True)
frow(rb,7,'Add: new order inflow',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),obi),
     fcst=lambda c:"=%s"%XR(ASM,'order_inflow',c),key='inflow')
frow(rb,8,'Less: standalone revenue executed',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),obx),
     fcst=lambda c:"=-%s*%s6"%(XR(ASM,'exec_rate',c),CL(c)),key='exec_neg')
frow(rb,9,'+/- Adjustments (cancellations/FX/scope/taxes)',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),obadj),
     fcst=lambda c:0,key='ob_adj')
ah_label(rb,10,'Closing order book (standalone)',bold=True)
for c in ACOLS: put(rb,10,c,"=%s6+%s7+%s8+%s9"%(CL(c),CL(c),CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(rb,'close_ob',10)
section(rb,12,'B.  Standalone -> consolidated revenue bridge')
frow(rb,13,'Standalone revenue from operations',hist=lambda c:"=-%s8"%CL(c),fcst=lambda c:"=-%s8"%CL(c),key='rev_sa',bold=True)
frow(rb,14,'Consolidation uplift % (consol/standalone - 1)',
     hist=lambda c:"=%s/%s13-1"%(lnk(c,'rev',HFS),CL(c)),fcst=lambda c:"=%s"%XR(ASM,'consol_uplift',c),fmt=FMT_PCT,key='uplift')
ah_label(rb,15,'Consolidated revenue from operations  ->  P&L',bold=True)
for c in ACOLS: put(rb,15,c,"=%s13*(1+%s14)"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT,border=B_TOP)
reg(rb,'rev',15)
def rbpct(r,label,fn,fmt,start=3):
    ah_label(rb,r,label)
    for c in ACOLS:
        if c<start: continue
        put(rb,r,c,fn(c),font=F_BLUE,fmt=fmt,align=RGT)
rbpct(16,'Revenue growth % (consolidated)',lambda c:"=%s15/%s15-1"%(CL(c),CL(c-1)),FMT_PCT)
rbpct(17,'Execution rate % (standalone rev / opening OB)',lambda c:"=%s13/%s6"%(CL(c),CL(c)),FMT_PCT,start=2)
rbpct(18,'Book-to-bill (closing OB / standalone rev, x)',lambda c:"=%s10/%s13"%(CL(c),CL(c)),FMT_X,start=2)
section(rb,20,'C.  Segment revenue breakup (consolidated)')
segpow=REG['Operational Drivers']['seg_pow']; segind=REG['Operational Drivers']['seg_ind']; segoth=REG['Operational Drivers']['seg_oth']
frow(rb,21,'Power segment',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),segpow),fcst=lambda c:"=%s15*%s"%(CL(c),XR(ASM,'power_mix',c)),key='seg_pow')
frow(rb,22,'Industry segment',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),segind),fcst=lambda c:"=%s15*%s"%(CL(c),XR(ASM,'industry_mix',c)),key='seg_ind')
frow(rb,23,'Others / exports / renewables',hist=lambda c:"='Operational Drivers'!%s%d"%(CL(c),segoth),fcst=lambda c:"=%s15*(1-%s-%s)"%(CL(c),XR(ASM,'power_mix',c),XR(ASM,'industry_mix',c)),key='seg_oth')
ah_label(rb,24,'Total segment revenue (check vs consolidated)',bold=True)
for c in ACOLS: put(rb,24,c,"=SUM(%s21:%s23)"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
freeze(rb,'B5')

# ===== 6B. CONTRACT ACCOUNTING SCHEDULE (Ind AS 115) =====
# Sell-side / CFA-standard schedule for Contract Assets & Contract Liabilities,
# split Current / Non-current, historical (FY20-26, audited) and forecast
# (FY27-33, driver-based). Methodology and full historical extraction live in
# contract_schedule.py - see that module's docstring for sources, data-quality
# notes and the coefficient-of-variation driver-selection analysis.
CAS='Contract Accounting Schedule'; cas=newsheet(CAS)
put(cas,2,1,'Ind AS 115 "Revenue from Contracts with Customers". Historical figures are AUDITED, extracted from each annual report\'s Ind AS 115 note (movement in impairment, disaggregation of revenue, contract balances) and the Other Assets/Other Liabilities notes (current/non-current split). Forecast drivers are selected by coefficient-of-variation analysis across 5 candidate ratios per item (Section B) - see contract_schedule.py for full workings.',font=F_NOTE)
yhdr(cas,4,'INR Crore')

# --- A. Historical schedule: Contract Assets ---
section(cas,5,'A.  Contract Assets - historical (audited)')
def cas_h(r,label,vals,fmt=FMT_CR,bold=False,indent=0,key=None,fill=None):
    ah_label(cas,r,label,bold=bold,indent=indent)
    for j,c in enumerate(HCOLS): put(cas,r,c,vals[j],font=F_BLACK,fmt=fmt,fill=fill,align=RGT,border=B_ALL if bold else None)
    if key: reg(cas,key,r)
cas_h(6,'Current contract assets (unbilled revenue)',CS.H['ca_cur'],indent=1,key='h_ca_cur')
cas_h(7,'Non-current contract assets',CS.H['ca_nc'],indent=1,key='h_ca_nc')
ah_label(cas,8,'Total contract assets',bold=True)
for c in HCOLS: put(cas,8,c,"=%s6+%s7"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT,border=B_TOP)
reg(cas,'h_ca_tot',8)
section(cas,10,'B.  Contract Liabilities - historical (audited)')
cas_h(11,'Current contract liabilities (advances/billing-in-excess)',CS.H['cl_cur'],indent=1,key='h_cl_cur')
cas_h(12,'Non-current contract liabilities',CS.H['cl_nc'],indent=1,key='h_cl_nc')
ah_label(cas,13,'Total contract liabilities',bold=True)
for c in HCOLS: put(cas,13,c,"=%s11+%s12"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT,border=B_TOP)
reg(cas,'h_cl_tot',13)
put(cas,15,1,'DATA QUALITY: FY2023 contract assets use the AS-ORIGINALLY-REPORTED Rs29,740 cr (AR2022-23 Note 9/40, current+NC = Rs10,811+Rs18,929), which differs from a later regrouped comparative of Rs26,466 cr shown in the AR2024-25 five-year summary; the original figure is used here because only it has a current/non-current split disclosed. FY2026 split (Rs15,193 cr current / Rs14,197 cr non-current) is per the BSE integrated filing.',font=F_NOTE)

# --- C. Historical roll-forward (partial mechanics; derived balancing items clearly flagged) ---
section(cas,17,'C.  Historical roll-forward (Ind AS 115 mechanics)')
put(cas,18,1,'BHEL does not disclose "billings during the year" or "gross advances received during the year" as explicit line items. The roll below uses the two components BHEL DOES disclose every year (revenue recognised over time [%-of-completion, i.e. project revenue - builds contract assets] and "revenue recognised against contract liabilities" [reduces the opening contract-liability balance]) and derives the balancing "billed to customer" / "advances received" figure so each roll ties exactly to the reported closing balance. Derived figures are marked [D].',font=F_NOTE)
put(cas,20,1,'Contract Assets roll',font=F_LBLB)
for j,h in enumerate(['Opening CA','+ Revenue recognised\n(over time / %-of-completion)','- Billed to customer [D]','= Closing CA']):
    put(cas,21,2+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
r=22
for i in range(1,7):
    hr = CS.historical_roll()[i-1]
    put(cas,r,1,HIST[i],font=F_LBL)
    if hr['ot_rev'] is None:
        put(cas,r,2,hr['ca_open'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,3,'n/d',font=F_ITAL,align=CEN,border=B_ALL)
        put(cas,r,4,'n/d',font=F_ITAL,align=CEN,border=B_ALL)
        put(cas,r,5,hr['ca_close'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
    else:
        put(cas,r,2,hr['ca_open'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,3,hr['ot_rev'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,4,hr['billed'],font=F_ITAL,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,5,hr['ca_close'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
    r+=1
put(cas,r+1,1,'Contract Liabilities roll',font=F_LBLB); r+=2
for j,h in enumerate(['Opening CL','+ Advances received [D]','- Revenue recognised\n(against opening CL)','= Closing CL']):
    put(cas,r,2+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
r+=1
for i in range(1,7):
    hr = CS.historical_roll()[i-1]
    put(cas,r,1,HIST[i],font=F_LBL)
    if hr['rev_recog'] is None:
        put(cas,r,2,hr['cl_open'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,3,'n/d',font=F_ITAL,align=CEN,border=B_ALL)
        put(cas,r,4,'n/d',font=F_ITAL,align=CEN,border=B_ALL)
        put(cas,r,5,hr['cl_close'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
    else:
        put(cas,r,2,hr['cl_open'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,3,hr['adv_recvd'],font=F_ITAL,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,4,hr['rev_recog'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
        put(cas,r,5,hr['cl_close'],font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
    r+=1
r+=1

# --- D. Driver ratio analysis with CV-based selection ---
section(cas,r,'D.  Driver ratio analysis (coefficient of variation - lower = more stable = better forecasting driver)'); r+=1
drow=r; put(cas,drow,1,'Driver',font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
for j,c in enumerate(HCOLS): put(cas,drow,2+j,HIST[j],font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
put(cas,drow,9,'CV',font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
put(cas,drow,10,'Selected?',font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
drow+=1
driver_defs = [
 ('CA / Revenue  <- SELECTED (CA driver)', CS.DRIVERS['ca_to_rev'], FMT_X, True),
 ('CA / Closing order book', CS.DRIVERS['ca_to_ob'], FMT_X, False),
 ('CA days (vs revenue)', CS.DRIVERS['ca_days'], FMT_DAYS, False),
 ('CL / Revenue', CS.DRIVERS['cl_to_rev'], FMT_X, False),
 ('CL / Closing order book  <- SELECTED (CL driver)', CS.DRIVERS['cl_to_ob'], FMT_X, True),
 ('CL / Order inflow', CS.DRIVERS['cl_to_inflow'], FMT_X, False),
 ('CL days (vs revenue)', CS.DRIVERS['cl_days'], FMT_DAYS, False),
 ('CA current %  (of total CA)', CS.DRIVERS['ca_cur_pct'], FMT_PCT, False),
 ('CA non-current %  (of total CA)', CS.DRIVERS['ca_nc_pct'], FMT_PCT, False),
 ('CL current %  (of total CL)', CS.DRIVERS['cl_cur_pct'], FMT_PCT, False),
 ('CL non-current %  (of total CL)', CS.DRIVERS['cl_nc_pct'], FMT_PCT, False),
]
for label, series, fmt, selected in driver_defs:
    ah_label(cas,drow,label,bold=selected)
    for j,v in enumerate(series):
        put(cas,drow,2+j,v,font=(F_LBLB if selected else F_BLACK),fmt=fmt,align=RGT,border=B_ALL)
    cv = CS.CV.get([k for k,vv in CS.DRIVERS.items() if vv is series][0])
    put(cas,drow,9,cv,font=(F_LBLB if selected else F_BLACK),fmt='0.000',align=RGT,border=B_ALL,fill=(FILL_OK if selected else None))
    put(cas,drow,10,('YES' if selected else ''),font=F_LBLB,align=CEN,border=B_ALL,fill=(FILL_OK if selected else None))
    drow+=1
drow+=1
put(cas,drow,1,'SELECTION RATIONALE: Contract assets are unbilled REVENUE recognised under the input-cost (%-of-completion) method - they scale with revenue actually recognised, not with the (mostly unexecuted) order book. CA/Revenue has the lowest CV (0.139 vs 0.330 for CA/Order-book) and is economically sound -> SELECTED. Contract liabilities are customer ADVANCES tied to work still to be performed - they scale with the size of the order book (future work), not with revenue already recognised. CL/Closing-order-book has the lowest CV among order-book/inflow-based measures (0.178 vs 0.369 for CL/Revenue, 0.566 for CL/Inflow) and is economically sound -> SELECTED. Both current/non-current splits show a STRUCTURAL SHIFT after FY23 (working-capital initiatives per the Q4FY24 concall, 21-May-2024); the recent 3-year (FY24-26) average is used for the split, not the full 7-year average, as it is far more stable (CV ~0.03-0.07 vs ~0.19-0.23 over the full period) and represents the current operating regime.',font=F_NOTE)
cas.row_dimensions[drow].height=60
drow+=3

# --- E. Forecast FY2027-2033 ---
# NOTE: every row is placed at an EXPLICIT, NAMED row number (no relative
# drow-N arithmetic) to avoid off-by-one referencing errors.
section(cas,drow,'E.  Forecast FY2027-FY2033 (driver-based, linked to Revenue Build-up)'); drow+=1
fh=drow; put(cas,fh,1,'',font=F_LBL)
for j,c in enumerate(FCOLS): put(cas,fh,2+j,FCST[j],font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
def cas_f(r,label,fn,fmt=FMT_CR,bold=False,indent=0,key=None,fill=None):
    ah_label(cas,r,label,bold=bold,indent=indent)
    for c in FCOLS: put(cas,r,c,fn(c),font=F_BLUE,fmt=fmt,fill=fill,align=RGT)
    if key: reg(cas,key,r)

r_rev      = fh+1
r_ob       = fh+2
r_ca_hdr   = fh+4
r_ca_ratio = fh+5
r_ca_tot   = fh+6
r_ca_split = fh+7
r_ca_cur   = fh+8
r_ca_nc    = fh+9
r_cl_hdr   = fh+11
r_cl_ratio = fh+12
r_cl_tot   = fh+13
r_cl_split = fh+14
r_cl_cur   = fh+15
r_cl_nc    = fh+16

cas_f(r_rev,'Revenue (from Revenue Build-up)',lambda c:"=%s"%XR(RB,'rev',c),indent=1,key='f_rev')
cas_f(r_ob,'Closing order book (from Revenue Build-up)',lambda c:"=%s"%XR(RB,'close_ob',c),indent=1,key='f_ob')
ah_label(cas,r_ca_hdr,'Contract Assets',bold=True)
cas_f(r_ca_ratio,'CA / Revenue (driver, selected)',lambda c:"=%s"%XR(ASM,'ca_to_rev',c),fmt=FMT_PCT,indent=1,key='f_ca_ratio')
ah_label(cas,r_ca_tot,'Total contract assets',bold=True)
for c in FCOLS: put(cas,r_ca_tot,c,"=%s%d*%s%d"%(CL(c),r_ca_ratio,CL(c),r_rev),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(cas,'f_ca_tot',r_ca_tot)
cas_f(r_ca_split,'Current split (driver, selected)',lambda c:"=%s"%XR(ASM,'ca_cur_split',c),fmt=FMT_PCT,indent=1,key='f_ca_curpct')
ah_label(cas,r_ca_cur,'Current contract assets',indent=1)
for c in FCOLS: put(cas,r_ca_cur,c,"=%s%d*%s%d"%(CL(c),r_ca_tot,CL(c),r_ca_split),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(cas,'f_ca_cur',r_ca_cur)
ah_label(cas,r_ca_nc,'Non-current contract assets',indent=1)
for c in FCOLS: put(cas,r_ca_nc,c,"=%s%d-%s%d"%(CL(c),r_ca_tot,CL(c),r_ca_cur),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(cas,'f_ca_nc',r_ca_nc)

ah_label(cas,r_cl_hdr,'Contract Liabilities',bold=True)
cas_f(r_cl_ratio,'CL / Closing order book (driver, selected)',lambda c:"=%s"%XR(ASM,'cl_to_ob',c),fmt=FMT_PCT,indent=1,key='f_cl_ratio')
ah_label(cas,r_cl_tot,'Total contract liabilities',bold=True)
for c in FCOLS: put(cas,r_cl_tot,c,"=%s%d*%s%d"%(CL(c),r_cl_ratio,CL(c),r_ob),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(cas,'f_cl_tot',r_cl_tot)
cas_f(r_cl_split,'Current split (driver, selected)',lambda c:"=%s"%XR(ASM,'cl_cur_split',c),fmt=FMT_PCT,indent=1,key='f_cl_curpct')
ah_label(cas,r_cl_cur,'Current contract liabilities',indent=1)
for c in FCOLS: put(cas,r_cl_cur,c,"=%s%d*%s%d"%(CL(c),r_cl_tot,CL(c),r_cl_split),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(cas,'f_cl_cur',r_cl_cur)
ah_label(cas,r_cl_nc,'Non-current contract liabilities',indent=1)
for c in FCOLS: put(cas,r_cl_nc,c,"=%s%d-%s%d"%(CL(c),r_cl_tot,CL(c),r_cl_cur),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(cas,'f_cl_nc',r_cl_nc)
drow = r_cl_nc+3

# --- F. Sensitivity: execution rate & CL/OB ratio impact on contract balances and cash ---
section(cas,drow,'F.  Sensitivity - execution rate & customer-advance (CL/order-book) rate'); drow+=1
put(cas,drow,1,'Impact of a +/-1pp change in FY2027E execution rate and FY2027E CL/order-book ratio on FY2027E contract balances and net working-capital cash impact (vs base case). All other assumptions held at base case.',font=F_NOTE)
drow+=2
sh1=drow
for j,h in enumerate(['Scenario','FY27E Rev','FY27E CA','FY27E CL','Net WC cash impact (CA-CL), delta vs base']):
    put(cas,sh1,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
sh1+=1
_sens_rows = [
 ('Base case', 0.0, 0.0),
 ('Execution rate +1pp', 0.01, 0.0),
 ('Execution rate -1pp', -0.01, 0.0),
 ('CL/order-book +1pp', 0.0, 0.01),
 ('CL/order-book -1pp', 0.0, -0.01),
]
_base_oi=[78000,82000,86000,90000,92000,94000,96000]
_base_er=[0.165,0.170,0.175,0.180,0.185,0.190,0.195]
_base_cu=[0.02]*7
_base_fc = CS.sensitivity(_base_oi,_base_er,_base_cu,CS.CA_TO_REV_FCST,CS.CL_TO_OB_FCST)
_base_wc = _base_fc['ca'][0]-_base_fc['cl'][0]
for lbl, der, dcl in _sens_rows:
    er = [_base_er[0]+der]+_base_er[1:]
    cl_ratio = [CS.CL_TO_OB_FCST[0]+dcl]+CS.CL_TO_OB_FCST[1:]
    fc = CS.sensitivity(_base_oi, er, _base_cu, CS.CA_TO_REV_FCST, cl_ratio)
    wc = fc['ca'][0]-fc['cl'][0]
    put(cas,sh1,1,lbl,font=F_LBL)
    put(cas,sh1,2,fc['rev'][0],font=F_BLACK,fmt=FMT_CR,align=RGT,border=B_ALL)
    put(cas,sh1,3,fc['ca'][0],font=F_BLACK,fmt=FMT_CR,align=RGT,border=B_ALL)
    put(cas,sh1,4,fc['cl'][0],font=F_BLACK,fmt=FMT_CR,align=RGT,border=B_ALL)
    put(cas,sh1,5,wc-_base_wc,font=F_BLACK,fmt=FMT_CR,align=RGT,border=B_ALL,fill=(FILL_OK if lbl=='Base case' else None))
    sh1+=1
sh1+=1
put(cas,sh1,1,'Reading: a HIGHER execution rate converts order book to revenue faster, raising contract assets (more unbilled revenue) - a cash DRAG. A HIGHER CL/order-book ratio means customers advance more cash relative to backlog - a cash BENEFIT. These are pre-computed reference scenarios (Python cross-check); the live two-way table below recalculates directly from the Assumptions sheet.',font=F_NOTE)
sh1+=2
# Live two-way data table: Net WC cash impact (CA - CL) for FY2027E, vs a
# delta on the execution rate (which drives revenue -> CA) and a delta on the
# CL/order-book ratio (which drives CL). Rebuilt explicitly from first
# principles so it is auditable:
#   open_ob   = FY2026 closing order book (Revenue Build-up, col H)
#   rev_new   = open_ob * (exec_rate + delta_er) * (1 + consol_uplift)
#   ob_new    = open_ob + order_inflow - open_ob*(exec_rate + delta_er)
#   ca_new    = ca_to_rev * rev_new
#   cl_new    = (cl_to_ob + delta_cl) * ob_new
#   net WC impact = ca_new - cl_new
put(cas,sh1,1,'Live sensitivity: FY2027E (Contract Assets - Contract Liabilities), INR Cr, vs execution-rate & CL/OB deltas',font=F_LBLB); sh1+=1
open_ob_ref = "'%s'!H10"%RB           # FY2026 closing order book (audited)
inflow_ref  = XR(ASM,'order_inflow',9)
uplift_ref  = XR(ASM,'consol_uplift',9)
ca_ratio_ref= XR(ASM,'ca_to_rev',9)
put(cas,sh1,2,'CL/OB delta \\ Exec.rate delta',font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
hdr_row=sh1
for j,ed in enumerate([-0.02,-0.01,0.0,0.01,0.02]):
    put(cas,sh1,3+j,"=%s+(%s)"%(XR(ASM,'exec_rate',9),ed),font=F_BLACK,fmt=FMT_PCT,fill=FILL_SUB,align=CEN,border=B_ALL)
for i,cd in enumerate([-0.02,-0.01,0.0,0.01,0.02]):
    rr=hdr_row+1+i
    put(cas,rr,2,"=%s+(%s)"%(XR(ASM,'cl_to_ob',9),cd),font=F_BLACK,fmt=FMT_PCT,fill=FILL_SUB,align=CEN,border=B_ALL)
    for j in range(5):
        er_cell="%s$%d"%(CL(3+j),hdr_row)
        cl_cell="$B%d"%rr
        rev_new="(%s*%s*(1+%s))"%(open_ob_ref,er_cell,uplift_ref)
        ob_new ="(%s+%s-%s*%s)"%(open_ob_ref,inflow_ref,open_ob_ref,er_cell)
        f="=%s*%s-%s*%s"%(ca_ratio_ref,rev_new,cl_cell,ob_new)
        put(cas,rr,3+j,f,font=F_BLUE,fmt=FMT_CR,align=RGT,border=B_ALL,fill=(FILL_OK if (i==2 and j==2) else None))
freeze(cas,'B5')
print("Contract Accounting Schedule built.")

# ===== 7. COST FORECAST (granular breakup) =====
CF='Cost Forecast'; cf=newsheet(CF)
put(cf,2,1,'Granular cost build-up. COGS = materials (steel/copper/components/imported/other) + subcontracting. OpEx = employee + selling + admin + R&D. EBITDA = revenue x target margin. Historical sub-components are not separately disclosed (shown at total level).',font=F_NOTE)
yhdr(cf,4); section(cf,5,'A.  Cost of goods sold (COGS)')
frow(cf,6,'Revenue from operations',hist=lambda c:"=%s"%lnk(c,'rev',HFS),fcst=lambda c:"=%s"%XR(RB,'rev',c),key='rev',bold=True)
frow(cf,7,'Steel',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'steel_pct',c)),indent=1)
frow(cf,8,'Copper',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'copper_pct',c)),indent=1)
frow(cf,9,'Electrical & electronic components',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'elec_pct',c)),indent=1)
frow(cf,10,'Imported components',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'imported_pct',c)),indent=1)
frow(cf,11,'Other materials & consumables',fcst=lambda c:"=%s6*(1-%s-%s-%s-%s-%s-%s-%s-%s)"%(CL(c),XR(ASM,'ebitda_margin',c),XR(ASM,'emp_pct',c),XR(ASM,'subc_pct',c),XR(ASM,'sga_pct',c),XR(ASM,'steel_pct',c),XR(ASM,'copper_pct',c),XR(ASM,'elec_pct',c),XR(ASM,'imported_pct',c)),indent=1)
frow(cf,12,'Subcontracting & erection',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'subc_pct',c)),indent=1)
frow(cf,13,'Total COGS',hist=lambda c:"=%s"%lnk(c,'cogs',HFS),fcst=lambda c:"=SUM(%s7:%s12)"%(CL(c),CL(c)),key='cogs',bold=True,fill=FILL_TOT)
frow(cf,14,'Gross profit',hist=lambda c:"=%s"%lnk(c,'gross',HFS),fcst=lambda c:"=%s6-%s13"%(CL(c),CL(c)),key='gross',bold=True)
section(cf,15,'B.  Operating expenses')
frow(cf,16,'Employee cost',hist=lambda c:"=%s"%lnk(c,'emp',HFS),fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'emp_pct',c)),key='emp',indent=1)
frow(cf,17,'Selling & distribution',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'selling_pct',c)),indent=1)
frow(cf,18,'Administrative expenses',fcst=lambda c:"=%s6*(%s-%s-%s)"%(CL(c),XR(ASM,'sga_pct',c),XR(ASM,'selling_pct',c),XR(ASM,'rnd_pct',c)),indent=1)
frow(cf,19,'Research & development',fcst=lambda c:"=%s6*%s"%(CL(c),XR(ASM,'rnd_pct',c)),indent=1)
frow(cf,20,'Selling, admin & R&D',hist=lambda c:"=%s"%lnk(c,'sga',HFS),fcst=lambda c:"=SUM(%s17:%s19)"%(CL(c),CL(c)),key='sga',bold=True)
frow(cf,21,'Total operating expenses',hist=lambda c:"=%s"%lnk(c,'totopex',HFS),fcst=lambda c:"=%s16+%s20"%(CL(c),CL(c)),key='totopex',bold=True,fill=FILL_TOT)
frow(cf,22,'EBITDA',hist=lambda c:"=%s"%lnk(c,'ebitda',HFS),fcst=lambda c:"=%s14-%s21"%(CL(c),CL(c)),key='ebitda',bold=True,fill=FILL_TOT)
ah_label(cf,23,'EBITDA margin %')
for c in ACOLS: put(cf,23,c,"=%s22/%s6"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_PCT,align=RGT)
frow(cf,24,'Memo: total operating cost (COGS+OpEx)',hist=lambda c:"=%s"%lnk(c,'totcost',HFS),fcst=lambda c:"=%s13+%s21"%(CL(c),CL(c)),key='totcost',fcstfont=F_BLUE)
freeze(cf,'B5')
print("Revenue Build + Cost Forecast (granular) built.")


# ===== 8. WORKING CAPITAL SCHEDULE =====
WC='Working Capital Schedule'; wc=newsheet(WC)
put(wc,2,1,'Working capital = CURRENT operating items only. Non-current & non-operating items (deferred tax, non-current contract assets/liabilities, long-term provisions) are shown in a separate block and EXCLUDED from operating NWC and from the DCF free cash flow; their change flows to cash via the total line so the balance sheet still ties. FY2026 seeds are audited actuals.',font=F_NOTE)
yhdr(wc,4); section(wc,5,'Operating working capital')
frow(wc,6,'Revenue from operations',hist=lambda c:"=%s"%lnk(c,'rev',HFS),fcst=lambda c:"=%s"%XR(CF,'rev',c),key='rev')
frow(wc,7,'Cost of sales (total operating cost)',hist=lambda c:"=%s"%lnk(c,'totcost',HFS),fcst=lambda c:"=%s"%XR(CF,'totcost',c),key='cost')
frow(wc,8,'Trade receivables',seed=6796.27,fcst=lambda c:"=%s/365*%s6"%(XR(ASM,'recv_days',c),CL(c)),key='recv',indent=1)
frow(wc,9,'Inventories',seed=13334.58,fcst=lambda c:"=%s/365*%s"%(XR(ASM,'inv_days',c),XR(CF,'cogs',c)),key='inv',indent=1)
frow(wc,10,'Contract assets (current)',seed=15192.89,fcst=lambda c:"=%s"%XR(CAS,'f_ca_cur',c),key='ca',indent=1)
frow(wc,11,'Other current operating assets',seed=4264.99,fcst=lambda c:"=%s*(1+%s)"%(("H11" if c==9 else "%s11"%CL(c-1)),XR(ASM,'oca_g',c)),key='oca_cur',indent=1)
ah_label(wc,12,'Total current operating assets',bold=True)
for c in range(8,16): put(wc,12,c,"=SUM(%s8:%s11)"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(wc,'ta_op',12)
frow(wc,13,'Trade payables',seed=10491.60,fcst=lambda c:"=%s/365*%s"%(XR(ASM,'pay_days',c),XR(CF,'cogs',c)),key='pay',indent=1)
frow(wc,14,'Contract liabilities / advances (current)',seed=9110.39,fcst=lambda c:"=%s"%XR(CAS,'f_cl_cur',c),key='cl',indent=1)
frow(wc,15,'Other current operating liabilities',seed=4681.42,fcst=lambda c:"=%s*(1+%s)"%(("H15" if c==9 else "%s15"%CL(c-1)),XR(ASM,'ol_g',c)),key='ol_cur',indent=1)
ah_label(wc,16,'Total current operating liabilities',bold=True)
for c in range(8,16): put(wc,16,c,"=SUM(%s13:%s15)"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(wc,'tl_op',16)
ah_label(wc,17,'Operating net working capital (current only)',bold=True)
for c in range(8,16): put(wc,17,c,"=%s12-%s16"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT,border=B_TOP)
reg(wc,'nwc',17)
ah_label(wc,18,'(Increase)/decrease in operating NWC  ->  FCFF')
for c in FCOLS: put(wc,18,c,"=-(%s17-%s17)"%(CL(c),CL(c-1)),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(wc,'dnwc_op',18)
section(wc,20,'Non-current & non-operating items (on balance sheet; EXCLUDED from working capital)')
put(wc,20,16,'incl deferred tax (non-operating) & long-dated contract balances',font=F_NOTE)
frow(wc,21,'Other non-current assets (= non-contract residual + NC contract assets, from Contract Accounting Schedule)',seed=20934.62,fcst=lambda c:"=%s+%s"%(XR(ASM,'other_nca',c),XR(CAS,'f_ca_nc',c)),key='onca',indent=1)
frow(wc,22,'Other non-current liabilities (= non-contract residual + NC contract liab, from Contract Accounting Schedule)',seed=17568.65,fcst=lambda c:"=%s+%s"%(XR(ASM,'other_ncl',c),XR(CAS,'f_cl_nc',c)),key='oncl',indent=1)
ah_label(wc,23,'Net non-current items (assets - liabilities)',bold=True)
for c in range(8,16): put(wc,23,c,"=%s21-%s22"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT,border=B_TOP)
ah_label(wc,24,'(Increase)/decrease in net non-current items')
for c in FCOLS: put(wc,24,c,"=-(%s23-%s23)"%(CL(c),CL(c-1)),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(wc,'dnc',24)
ah_label(wc,25,'Total (increase)/decrease in NWC incl non-current  ->  cash flow',bold=True)
for c in FCOLS: put(wc,25,c,"=%s18+%s24"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(wc,'dnwc',25)
frow(wc,27,'Memo: cash & bank (FY26 seed, audited)',seed=11866.62,fcst=None,key='cash_seed')
section(wc,29,'Memo: current-asset composition (indicative, forecast)')
frow(wc,30,'Inventory: raw materials & components (~45%)',fcst=lambda c:"=%s9*0.45"%CL(c),indent=1)
frow(wc,31,'Inventory: work-in-progress (~35%)',fcst=lambda c:"=%s9*0.35"%CL(c),indent=1)
frow(wc,32,'Inventory: finished goods & stores (~20%)',fcst=lambda c:"=%s9*0.20"%CL(c),indent=1)
frow(wc,33,'Receivables: billed / trade (~70%)',fcst=lambda c:"=%s8*0.70"%CL(c),indent=1)
frow(wc,34,'Receivables: unbilled / retention (~30%)',fcst=lambda c:"=%s8*0.30"%CL(c),indent=1)
freeze(wc,'B5')

# ===== 9. FIXED ASSET + 10. DEPRECIATION =====
FA='Fixed Asset Schedule'; fa=newsheet(FA)
put(fa,2,1,'Net-block roll-forward. Opening FY27 = reported FY26 net PPE. Depreciation = rate x opening net block.',font=F_NOTE)
yhdr(fa,4); section(fa,5,'Net block roll-forward')
frow(fa,6,'Opening net block',hist=lambda c:("=%s"%lnk(c,'ppe',HFS)) if c<8 else None,
     fcst=lambda c:("=%s"%lnk(8,'ppe',HFS)) if c==9 else ("=%s10"%CL(c-1)),key='open_nb',bold=True)
put(fa,6,8,"=%s"%lnk(8,'ppe',HFS),font=F_GREEN,fmt=FMT_CR,align=RGT)
frow(fa,7,'Add: capital expenditure',fcst=lambda c:"=%s"%XR(ASM,'capex',c),key='capex',indent=1)
frow(fa,8,'Less: depreciation',fcst=lambda c:"=-%s*%s6"%(XR(ASM,'dep_rate',c),CL(c)),key='dep_neg',indent=1)
frow(fa,9,'Less: disposals',fcst=lambda c:0,key='disp',indent=1)
frow(fa,10,'Closing net block',hist=lambda c:"=%s"%lnk(c,'ppe',HFS),
     fcst=lambda c:"=%s6+%s7+%s8-%s9"%(CL(c),CL(c),CL(c),CL(c)),key='close_nb',bold=True,fill=FILL_TOT)
freeze(fa,'B5')
DS='Depreciation Schedule'; ds=newsheet(DS); yhdr(ds,4); section(ds,5,'Depreciation computation')
frow(ds,6,'Opening net block',fcst=lambda c:"=%s"%XR(FA,'open_nb',c),key='open_nb')
frow(ds,7,'Depreciation rate %',fcst=lambda c:"=%s"%XR(ASM,'dep_rate',c),fmt=FMT_PCT,key='rate')
frow(ds,8,'Depreciation expense',hist=lambda c:"=%s"%lnk(c,'dep',HFS),fcst=lambda c:"=%s6*%s7"%(CL(c),CL(c)),key='dep',bold=True,fill=FILL_TOT)
freeze(ds,'B5')

# ===== 11. DEBT SCHEDULE =====
DBT='Debt Schedule'; db=newsheet(DBT)
put(db,2,1,'Interest on opening debt (no circularity). Opening FY27 = reported FY26 borrowings.',font=F_NOTE)
yhdr(db,4); section(db,5,'Debt roll-forward')
frow(db,6,'Opening debt',hist=lambda c:("=%s"%lnk(c,'bor',HFS)) if c<8 else None,
     fcst=lambda c:("=%s"%lnk(8,'bor',HFS)) if c==9 else ("=%s9"%CL(c-1)),key='open_d',bold=True)
put(db,6,8,"=%s"%lnk(8,'bor',HFS),font=F_GREEN,fmt=FMT_CR,align=RGT)
frow(db,7,'Add: new borrowings',fcst=lambda c:"=%s"%XR(ASM,'new_borrow',c),key='new',indent=1)
frow(db,8,'Less: repayments',fcst=lambda c:"=-%s"%XR(ASM,'debt_repay',c),key='repay',indent=1)
frow(db,9,'Closing debt',hist=lambda c:"=%s"%lnk(c,'bor',HFS),fcst=lambda c:"=%s6+%s7+%s8"%(CL(c),CL(c),CL(c)),key='close_d',bold=True,fill=FILL_TOT)
frow(db,10,'Average debt',fcst=lambda c:"=(%s6+%s9)/2"%(CL(c),CL(c)),key='avg_d')
frow(db,11,'Interest expense (on opening debt)',hist=lambda c:"=%s"%lnk(c,'fin',HFS),fcst=lambda c:"=%s*%s6"%(XR(ASM,'int_rate',c),CL(c)),key='interest',bold=True)
freeze(db,'B5')
print("WC + Fixed Asset + Depreciation + Debt built.")


# ===== 15. FORECAST FINANCIAL STATEMENTS (integrated, gross-profit IS + BS) =====
FFS='Forecast Financial Statements'; ff=newsheet(FFS)
put(ff,2,1,'Integrated three-statement model. FY20-26 reported; FY27-33 forecast (linked to schedules). BS detail shown FY26 onward.',font=F_NOTE)
def ffall(r,label,fn,fmt=FMT_CR,bold=False,fill=None,key=None,border=None,indent=0):
    ah_label(ff,r,label,bold=bold,indent=indent)
    for c in ACOLS: put(ff,r,c,fn(c),font=F_BLUE,fmt=fmt,align=RGT,fill=fill,border=border)
    if key: reg(ff,key,r)
section(ff,3,'A.  INCOME STATEMENT'); yhdr(ff,4,'Particulars')
frow(ff,5,'Sales',hist=lambda c:"=%s"%lnk(c,'rev',HFS),fcst=lambda c:"=%s"%XR(RB,'rev',c),key='rev',bold=True)
frow(ff,6,'Total COGS',hist=lambda c:"=%s"%lnk(c,'cogs',HFS),fcst=lambda c:"=%s"%XR(CF,'cogs',c),key='cogs')
ffall(7,'Gross Profit',lambda c:"=%s5-%s6"%(CL(c),CL(c)),key='gross',bold=True,border=B_TOP)
ah_label(ff,8,'Operating Expenses')
frow(ff,9,'Employee Cost',hist=lambda c:"=%s"%lnk(c,'emp',HFS),fcst=lambda c:"=%s"%XR(CF,'emp',c),key='emp',indent=1)
frow(ff,10,'Selling and admin',hist=lambda c:"=%s"%lnk(c,'sga',HFS),fcst=lambda c:"=%s"%XR(CF,'sga',c),key='sga',indent=1)
ffall(11,'Total Operating Expenses',lambda c:"=%s9+%s10"%(CL(c),CL(c)),key='totopex',bold=True,border=B_TOP)
ffall(12,'EBITDA',lambda c:"=%s7-%s11"%(CL(c),CL(c)),key='ebitda',bold=True,fill=FILL_TOT,border=B_TOP)
frow(ff,13,'Depreciation & Amortization',hist=lambda c:"=%s"%lnk(c,'dep',HFS),fcst=lambda c:"=%s"%XR(DS,'dep',c),key='dep')
ffall(14,'EBIT',lambda c:"=%s12-%s13"%(CL(c),CL(c)),key='ebit',bold=True)
frow(ff,15,'Other Income',hist=lambda c:"=%s"%lnk(c,'oi',HFS),fcst=lambda c:"=%s*%s5"%(XR(ASM,'oth_inc_pct',c),CL(c)),key='oi')
frow(ff,16,'Interest Paid',hist=lambda c:"=%s"%lnk(c,'fin',HFS),fcst=lambda c:"=%s"%XR(DBT,'interest',c),key='fin')
ffall(17,'Profit Before Tax (PBT)',lambda c:"=%s14+%s15-%s16"%(CL(c),CL(c),CL(c)),key='pbt',bold=True,border=B_TOP)
frow(ff,18,'Tax',hist=lambda c:"=%s"%lnk(c,'tax',HFS),fcst=lambda c:"=%s*%s17"%(XR(ASM,'tax_rate',c),CL(c)),key='tax')
frow(ff,19,'Adjustments',hist=lambda c:"=%s"%lnk(c,'adj',HFS),fcst=lambda c:0,key='adj')
ffall(20,'Profit After Tax (PAT)',lambda c:"=%s17-%s18+%s19"%(CL(c),CL(c),CL(c)),key='pat',bold=True,fill=FILL_TOT,border=B_TB)
ffall(21,'EPS (INR)',lambda c:"=%s20/%s"%(CL(c),XR(ASM,'shares',3)),fmt=FMT_PS,key='eps')
frow(ff,22,'Dividend',hist=lambda c:"=%s"%lnk(c,'div',HFS),fcst=lambda c:"=%s*%s20"%(XR(ASM,'div_payout',c),CL(c)),key='div')

def FR(key,c): return "'%s'!%s%d"%(FFS,CL(c),REG[FFS][key])

# ===== 12. EQUITY SCHEDULE =====
EQ='Equity Schedule'; eq=newsheet(EQ); yhdr(eq,4); section(eq,5,'Other equity (reserves) roll-forward')
frow(eq,6,'Opening other equity',fcst=lambda c:("=%s"%lnk(8,'res',HFS)) if c==9 else ("=%s10"%CL(c-1)),key='open_eq',bold=True)
frow(eq,7,'Add: profit after tax',fcst=lambda c:"=%s"%FR('pat',c),key='pat',indent=1)
frow(eq,8,'Less: dividends',fcst=lambda c:"=-%s"%FR('div',c),key='div',indent=1)
frow(eq,9,'Add: other comprehensive income',fcst=lambda c:0,key='oci',indent=1)
frow(eq,10,'Closing other equity',fcst=lambda c:"=SUM(%s6:%s9)"%(CL(c),CL(c)),key='close_eq',bold=True,fill=FILL_TOT)
frow(eq,12,'Equity share capital (constant)',fcst=lambda c:696,key='eqc')
frow(eq,13,'Total shareholders equity',fcst=lambda c:"=%s10+%s12"%(CL(c),CL(c)),key='tse',bold=True)
freeze(eq,'B5')

# ===== 13. OTHER ASSETS & LIABILITIES =====
OAL='Other Assets & Liabilities'; oal=newsheet(OAL); yhdr(oal,4); section(oal,5,'Other (non-WC) assets & liabilities')
frow(oal,6,'Capital work-in-progress',hist=lambda c:"=%s"%lnk(c,'cwip',HFS),fcst=lambda c:"=%s"%XR(ASM,'cwip',c),key='cwip')
frow(oal,7,'Investments',hist=lambda c:"=%s"%lnk(c,'inv',HFS),fcst=lambda c:"=%s"%XR(ASM,'investments',c),key='inv')
frow(oal,8,'Contract assets (memo, from WC)',fcst=lambda c:"=%s"%XR(WC,'ca',c),key='ca')
frow(oal,9,'Other current operating assets (memo, from WC)',fcst=lambda c:"=%s"%XR(WC,'oca_cur',c),key='oca')
frow(oal,10,'Contract liabilities (memo, from WC)',fcst=lambda c:"=%s"%XR(WC,'cl',c),key='cl')
frow(oal,11,'Other current operating liabilities (memo, from WC)',fcst=lambda c:"=%s"%XR(WC,'ol_cur',c),key='ol')
freeze(oal,'B5')
print("Forecast IS + Equity + Other A&L built.")


# ===== 14. CASH FLOW STATEMENT (indirect; historical AUDITED, forecast derived) =====
CFS='Cash Flow Statement'; cs=newsheet(CFS)
put(cs,2,1,'FY2020-26: AUDITED consolidated cash flow (annual reports), reconciling to cash & cash equivalents. FY2027-33: derived (PAT + D&A +/- change in NWC; investing & financing from schedules). See memo for the link to the broader balance-sheet "cash & bank" line.',font=F_NOTE)
yhdr(cs,4); section(cs,5,'A.  Operating activities')
frow(cs,6,'Profit after tax',hist=lambda c:"=%s"%lnk(c,'pat',HFS),fcst=lambda c:"=%s"%FR('pat',c),indent=1)
frow(cs,7,'Add: depreciation & amortisation',hist=lambda c:"=%s"%lnk(c,'dep',HFS),fcst=lambda c:"=%s"%XR(DS,'dep',c),indent=1)
frow(cs,8,'Working capital & other operating adjustments',hist=lambda c:"=%s-%s6-%s7"%(lnk(c,'cfo',HFS),CL(c),CL(c)),fcst=lambda c:"=%s"%XR(WC,'dnwc',c),indent=1)
frow(cs,9,'Cash flow from operating activities',hist=lambda c:"=SUM(%s6:%s8)"%(CL(c),CL(c)),fcst=lambda c:"=SUM(%s6:%s8)"%(CL(c),CL(c)),key='cfo',bold=True,fill=FILL_TOT)
section(cs,11,'B.  Investing activities')
frow(cs,12,'Capital expenditure (PPE & intangibles)',hist=lambda c:ACF['capex'][c-2],histfont=F_BLACK,fcst=lambda c:"=-%s"%XR(ASM,'capex',c),indent=1)
frow(cs,13,'Change in CWIP',hist=lambda c:0,histfont=F_BLACK,fcst=lambda c:"=-(%s-%s)"%(XR(ASM,'cwip',c),(lnk(8,'cwip',HFS) if c==9 else XR(ASM,'cwip',c-1))),indent=1)
frow(cs,14,'Bank deposits, investments, interest & other investing',hist=lambda c:"=%s-%s12-%s13"%(lnk(c,'cfi',HFS),CL(c),CL(c)),fcst=lambda c:"=-(%s-%s)"%(XR(ASM,'investments',c),(lnk(8,'inv',HFS) if c==9 else XR(ASM,'investments',c-1))),indent=1)
frow(cs,15,'Cash flow from investing activities',hist=lambda c:"=SUM(%s12:%s14)"%(CL(c),CL(c)),fcst=lambda c:"=SUM(%s12:%s14)"%(CL(c),CL(c)),key='cfi',bold=True,fill=FILL_TOT)
section(cs,17,'C.  Financing activities')
frow(cs,18,'New / (net) borrowings',hist=lambda c:ACF['borrow'][c-2],histfont=F_BLACK,fcst=lambda c:"=%s"%XR(DBT,'new',c),indent=1)
frow(cs,19,'Interest paid & lease payments',hist=lambda c:ACF['intlease'][c-2],histfont=F_BLACK,fcst=lambda c:"=%s"%XR(DBT,'repay',c),indent=1)
frow(cs,20,'Dividends paid',hist=lambda c:ACF['divpaid'][c-2],histfont=F_BLACK,fcst=lambda c:"=-%s"%FR('div',c),indent=1)
frow(cs,21,'Cash flow from financing activities',hist=lambda c:"=SUM(%s18:%s20)"%(CL(c),CL(c)),fcst=lambda c:"=SUM(%s18:%s20)"%(CL(c),CL(c)),key='cff',bold=True,fill=FILL_TOT)
ah_label(cs,23,'Net increase/(decrease) in cash & equivalents',bold=True)
for c in ACOLS: put(cs,23,c,"=%s9+%s15+%s21"%(CL(c),CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
reg(cs,'netcf',23)
frow(cs,24,'Opening cash & cash equivalents',hist=lambda c:ACF['open'][c-2]+ACF['roll_adj'][c-2],histfont=F_BLACK,
     fcst=lambda c:("='Working Capital Schedule'!H%d"%REG[WC]['cash_seed']) if c==9 else ("=%s25"%CL(c-1)),key='open_cash')
frow(cs,25,'Closing cash & cash equivalents',hist=lambda c:"=%s24+%s23"%(CL(c),CL(c)),fcst=lambda c:"=%s24+%s23"%(CL(c),CL(c)),key='close_cash',bold=True,fill=FILL_TOT)
ah_label(cs,27,'Memo: add bank balances/deposits & current investments',italic=True)
for c in HCOLS: put(cs,27,c,"=%s-%s25"%(ABS['cash'][c-2],CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT)
ah_label(cs,28,'Memo: cash & bank per balance sheet (broad)',italic=True)
for c in HCOLS: put(cs,28,c,ABS['cash'][c-2],font=F_GREEN,fmt=FMT_CR,align=RGT)
put(cs,30,1,'Historical cash flow ties to AUDITED cash & cash equivalents (annual reports). FY22 opening includes a +7.28 cr cash-credit reclassification (disclosed). The balance-sheet "Cash & bank" line is broader - it adds bank deposits (>3 months) & current investments whose movement sits in investing activities; the forecast cash line and the FY27 opening balance are on this broad basis.',font=F_NOTE)
freeze(cs,'B5')

# ===== FORECAST BALANCE SHEET (grouped format, FY2020-FY2033; history filled) =====
section(ff,24,'B.  BALANCE SHEET  (FY2020-FY2026 reported, FY2027-33 forecast)')
yhdr(ff,25,'Particulars')
put(ff,26,1,'Equities & Liabilities',font=F_LBLB,fill=FILL_SUB)
for c in range(2,16): ff.cell(row=26,column=c).fill=FILL_SUB
# AUDITED detailed consolidated balance sheet FY2020-FY2026 (annual reports); ties exactly, NO plug.
H_eqc=ABS['eqc']; H_res=ABS['res']
H_ltbor=ABS['ltbor']; H_stbor=ABS['stbor']
H_bor=[H_ltbor[i]+H_stbor[i] for i in range(7)]
H_oncl=ABS['oncl']
H_pay=ABS['pay']
H_ocl=ABS['ocl']
H_othl=[H_oncl[i]+H_pay[i]+H_ocl[i] for i in range(7)]
H_ppe=ABS['ppe']; H_cwip=ABS['cwip']; H_invv=ABS['invst']
H_onca=ABS['onca']
H_recv=ABS['recv']; H_invy=ABS['invy']
H_cash=ABS['cash']
H_oca=ABS['oca']
H_tle=[H_eqc[i]+H_res[i]+H_bor[i]+H_othl[i] for i in range(7)]
def bline(r,label,hvals,fcst,indent=1,key=None):
    ah_label(ff,r,label,indent=indent)
    for j,c in enumerate(HCOLS): put(ff,r,c,hvals[j],font=F_BLACK,fmt=FMT_CR,align=RGT)
    for c in FCOLS: put(ff,r,c,fcst(c),font=F_GREEN,fmt=FMT_CR,align=RGT)
    if key: reg(ff,key,r)
def bsub(r,label,fn,fill=FILL_TOT,key=None,border=B_TOP):
    ah_label(ff,r,label,bold=True)
    for c in ACOLS: put(ff,r,c,fn(c),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=fill,border=border)
    if key: reg(ff,key,r)
ah_label(ff,27,"Shareholders' Fund",bold=True)
bline(28,'Share Capital',H_eqc,lambda c:"=%s"%XR(EQ,'eqc',c),key='eqc')
bline(29,'Reserve & Surplus',H_res,lambda c:"=%s"%XR(EQ,'close_eq',c),key='res')
bsub(30,"Total Shareholders' Fund",lambda c:"=%s28+%s29"%(CL(c),CL(c)),key='tsf')
ah_label(ff,32,'Liabilities',bold=True)
ah_label(ff,33,'Non-Current Liabilities',bold=True)
bline(34,'LT borrowings',H_ltbor,lambda c:"=%s*%s"%(XR(ASM,'lt_debt_pct',c),XR(DBT,'close_d',c)),key='lt_bor')
bline(35,'Other non current liabilities (incl NC contract liab)',H_oncl,lambda c:"=%s"%XR(WC,'oncl',c),key='other_ncl')
bsub(36,'Total Non-Current Liabilities',lambda c:"=%s34+%s35"%(CL(c),CL(c)),fill=None,key='tncl')
bline(37,'ST borrowings',H_stbor,lambda c:"=%s-%s34"%(XR(DBT,'close_d',c),CL(c)),key='st_bor')
bline(38,'Trade Payables',H_pay,lambda c:"=%s"%XR(WC,'pay',c),key='pay')
bline(39,'Other Current Liabilities (provisions, advances)',H_ocl,lambda c:"=%s+%s"%(XR(WC,'cl',c),XR(WC,'ol_cur',c)),key='other_cl')
bsub(40,'Total Liabilities',lambda c:"=%s36+%s37+%s38+%s39"%(CL(c),CL(c),CL(c),CL(c)),fill=None,key='tliab')
bsub(42,'Total Equity & Liabilities',lambda c:"=%s30+%s40"%(CL(c),CL(c)),key='tle',border=B_TB)
ah_label(ff,44,'Assets',bold=True)
for c in range(2,16): ff.cell(row=44,column=c).fill=FILL_SUB
put(ff,44,1,'Assets',font=F_LBLB,fill=FILL_SUB)
ah_label(ff,45,'Non-Current Assets',bold=True)
ah_label(ff,46,'Fixed Assets',bold=True)
bline(47,'Property Plant & Equip',H_ppe,lambda c:"=%s"%XR(FA,'close_nb',c),key='net_ppe')
bline(48,'Capital Work in Progress',H_cwip,lambda c:"=%s"%XR(ASM,'cwip',c),key='cwip')
bline(49,'Investments',H_invv,lambda c:"=%s"%XR(ASM,'investments',c),key='invst')
bsub(50,'Total Fixed Assets',lambda c:"=SUM(%s47:%s49)"%(CL(c),CL(c)),fill=None,key='tfa')
bline(51,'Other Non Current Assets (incl NC contract assets)',H_onca,lambda c:"=%s"%XR(WC,'onca',c),key='other_nca')
bsub(52,'Total Non Current Assets',lambda c:"=%s50+%s51"%(CL(c),CL(c)),fill=None,key='tnca')
ah_label(ff,54,'Current Assets',bold=True)
bline(55,'Accounts Receivables',H_recv,lambda c:"=%s"%XR(WC,'recv',c),key='recv')
bline(56,'Inventories',H_invy,lambda c:"=%s"%XR(WC,'inv',c),key='inventory')
bline(57,'Other Current Assets',H_oca,lambda c:"=%s+%s"%(XR(WC,'ca',c),XR(WC,'oca_cur',c)),key='other_ca')
bline(58,'Cash Balance (Incl. Bank and Current Invest)',H_cash,lambda c:"=%s"%XR(CFS,'close_cash',c),key='cash')
bsub(59,'Total Current Assets',lambda c:"=SUM(%s55:%s58)"%(CL(c),CL(c)),fill=None,key='tca')
bsub(61,'Total Assets',lambda c:"=%s52+%s59"%(CL(c),CL(c)),key='ta',border=B_TB)
ah_label(ff,62,'Balance check (TA - TE&L)',bold=True)
for c in ACOLS: put(ff,62,c,"=%s61-%s42"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(ff,'bschk',62)
ah_label(ff,63,'Memo: total borrowings (LT+ST)',italic=True)
for c in ACOLS: put(ff,63,c,"=%s34+%s37"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT)
reg(ff,'bor',63)
put(ff,64,1,'Historical balance-sheet sub-lines are analyst estimates anchored exactly to reported totals (debtor/inventory/payable days; cash rolled from reported net cash flow).',font=F_NOTE)
ah_label(ff,65,'Memo: Other NON-CURRENT ASSETS breakup (indicative)',bold=True)
for lbl,share,rr in [('Deferred tax assets (~40%)',0.40,66),('Long-term trade & other receivables (~35%)',0.35,67),('Long-term loans, deposits & other (~25%)',0.25,68)]:
    ah_label(ff,rr,lbl,indent=1,italic=True)
    for c in ACOLS: put(ff,rr,c,"=%s51*%s"%(CL(c),share),font=F_BLUE,fmt=FMT_CR,align=RGT)
ah_label(ff,69,'Memo: Other NON-CURRENT LIABILITIES breakup (indicative)',bold=True)
for lbl,share,rr in [('Lease liabilities (~30%)',0.30,70),('Long-term provisions - warranty & employee benefits (~50%)',0.50,71),('Deferred tax liabilities & other (~20%)',0.20,72)]:
    ah_label(ff,rr,lbl,indent=1,italic=True)
    for c in ACOLS: put(ff,rr,c,"=%s35*%s"%(CL(c),share),font=F_BLUE,fmt=FMT_CR,align=RGT)
put(ff,73,1,'Non-current breakup shares are indicative analyst estimates; verify against Annual Report notes.',font=F_NOTE)
freeze(ff,'B5')
print("Cash Flow + Forecast BS (grouped, history filled, NC breakup) built.")


# ===== 17. DCF VALUATION (FCFF) =====
DCF='DCF Valuation'; dc=newsheet(DCF)
put(dc,2,1,'Free Cash Flow to Firm (FCFF) valuation; WACC computed dynamically on Assumptions.',font=F_NOTE)
put(dc,4,1,'INR Crore',font=F_ITAL,fill=FILL_SUB)
for i,c in enumerate(FCOLS): put(dc,4,c,FCST[i],font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
section(dc,5,'A.  Unlevered free cash flow')
def dcfr(r,label,fcst,fmt=FMT_CR,bold=False,key=None,fill=None,indent=0,font=F_GREEN):
    ah_label(dc,r,label,bold=bold,indent=indent)
    for c in FCOLS: put(dc,r,c,fcst(c),font=font,fmt=fmt,align=RGT,fill=fill)
    if key: reg(dc,key,r)
dcfr(6,'EBIT',lambda c:"=%s"%FR('ebit',c),bold=True)
dcfr(7,'Less: cash taxes on EBIT',lambda c:"=-%s6*%s"%(CL(c),XR(ASM,'tax_rate',c)),font=F_BLUE,indent=1)
dcfr(8,'NOPAT',lambda c:"=%s6+%s7"%(CL(c),CL(c)),font=F_BLUE,bold=True)
dcfr(9,'Add: depreciation & amortisation',lambda c:"=%s"%XR(DS,'dep',c),indent=1)
dcfr(10,'Less: capital expenditure',lambda c:"=-%s"%XR(ASM,'capex',c),indent=1)
dcfr(11,'Add/(less): change in operating working capital',lambda c:"=%s"%XR(WC,'dnwc_op',c),indent=1)
dcfr(12,'Free cash flow to firm (FCFF)',lambda c:"=%s8+%s9+%s10+%s11"%(CL(c),CL(c),CL(c),CL(c)),font=F_BLUE,bold=True,fill=FILL_TOT,key='fcff')
ah_label(dc,13,'Discount period (years)')
for i,c in enumerate(FCOLS): put(dc,13,c,i+1,font=F_BLACK,fmt='0',align=RGT,fill=FILL_IN)
ah_label(dc,14,'Discount factor @ WACC')
for c in FCOLS: put(dc,14,c,"=1/(1+%s)^%s13"%(XR(ASM,'wacc',3),CL(c)),font=F_BLUE,fmt='0.000',align=RGT)
ah_label(dc,15,'PV of FCFF',bold=True)
for c in FCOLS: put(dc,15,c,"=%s12*%s14"%(CL(c),CL(c)),font=F_BLUE,fmt=FMT_CR,align=RGT,fill=FILL_TOT)
section(dc,17,'B.  Enterprise & equity value')
def dsum(r,label,formula,fmt=FMT_CR,bold=False,key=None,fill=None,font=F_BLUE):
    ah_label(dc,r,label,bold=bold)
    put(dc,r,3,formula,font=font,fmt=fmt,align=RGT,fill=fill,border=B_ALL)
    if key: reg(dc,key,r)
dsum(18,'Sum of PV of explicit FCFF (FY27-33)',"=SUM(I15:O15)",bold=True,key='sumpv')
dsum(19,'WACC',"=%s"%XR(ASM,'wacc',3),fmt=FMT_PCT,font=F_GREEN,key='wacc')
dsum(20,'Terminal growth rate (g)',"=%s"%XR(ASM,'term_g',3),fmt=FMT_PCT,font=F_GREEN,key='termg')
dsum(21,'Terminal value (Gordon growth, normalized WC)',"=(O8+O9+O10-C20*'Working Capital Schedule'!O17)*(1+C20)/(C19-C20)",key='tv')
dsum(22,'PV of terminal value',"=C21*O14",key='pvtv')
dsum(23,'Enterprise value (EV)',"=C18+C22",bold=True,fill=FILL_TOT,key='ev')
dsum(24,'Less: net debt (FY26 debt - cash)',"=-(%s-%s)"%(FR('bor',8),FR('cash',8)),key='netdebt',font=F_GREEN)
dsum(25,'Equity value',"=C23+C24",bold=True,fill=FILL_TOT,key='eqval')
dsum(26,'Shares outstanding (crore)',"=%s"%XR(ASM,'shares',3),fmt=FMT_CR1,font=F_GREEN,key='shares')
dsum(27,'Intrinsic value per share (INR)',"=C25/C26",fmt=FMT_PS,bold=True,fill=FILL_OK,key='ivps')
dsum(28,'Current market price (INR)',"=%s"%XR(ASM,'price',3),fmt=FMT_PS,font=F_GREEN,key='price')
dsum(29,'Upside / (downside)',"=C27/C28-1",fmt=FMT_PCT,bold=True,key='upside')
dc.column_dimensions['C'].width=16; freeze(dc,'B5')

# ===== 18. RELATIVE VALUATION =====
RV='Relative Valuation'; rv=newsheet(RV)
put(rv,2,1,'Peer multiples reflect current (2025-26) market levels. BHEL is valued at a QUALITY DISCOUNT to premium MNC peers (Siemens/ABB ~50x P/E) given its lower margins (~8% vs 15-18%), low ROE (~2-8% vs 15-30%) and PSU profile; the closest large comp is L&T (~20x EV/EBITDA, ~33x P/E). EV/Revenue is intentionally EXCLUDED - it is invalid across such different margin profiles. Refresh peer data before use.',font=F_NOTE)
for col,w in {'A':34,'B':13,'C':13,'D':13,'E':13,'F':13}.items(): rv.column_dimensions[col].width=w
section(rv,4,'A.  Peer trading multiples (current, indicative)',span=6)
for j,h in enumerate(['Peer company','EV/EBITDA (x)','P/E (x)','P/B (x)','ROE %','Div yld %']):
    put(rv,5,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
peers=[('Siemens India',42,50,9,16,0.4),('ABB India',48,52,14,20,0.4),('CG Power & Ind.',45,60,18,22,0.2),
 ('Thermax',34,48,8,14,0.3),('Triveni Turbine',38,50,16,30,0.5),('Larsen & Toubro',20,33,5,15,0.7),
 ('Kalpataru Projects',11,20,3,11,0.4)]
r=6
for nm,eve,pe,pb,roe,dy in peers:
    put(rv,r,1,nm,font=F_BLACK)
    for j,v in enumerate([eve,pe,pb,roe,dy]): put(rv,r,2+j,v,font=F_BLACK,fmt=(FMT_X if j<3 else '0.0"%"'),fill=FILL_IN,align=RGT,border=B_ALL)
    r+=1
med=r; put(rv,med,1,'Peer median',font=F_LBLB,fill=FILL_TOT)
for j in range(5): put(rv,med,2+j,"=MEDIAN(%s6:%s%d)"%(CL(2+j),CL(2+j),r-1),font=F_BLUE,fmt=(FMT_X if j<3 else '0.0"%"'),align=RGT,fill=FILL_TOT,border=B_ALL)
put(rv,med,7,'BHEL ROE ~2-8% -> discount to peers',font=F_NOTE)
# B. BHEL target multiples (quality-discounted): bear / base / bull
tm=med+2; section(rv,tm,'B.  BHEL target multiples (quality-discounted)',span=6); tm+=1
for j,h in enumerate(['Multiple','Bear','Base','Bull']): put(rv,tm,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
put(rv,tm+1,1,'EV/EBITDA (x)  [L&T ~20x]',font=F_LBL)
for j,v in enumerate([15,20,28]): put(rv,tm+1,2+j,v,font=F_BLACK,fmt=FMT_X,fill=FILL_IN,align=RGT,border=B_ALL)
reg(rv,'evx',tm+1)
put(rv,tm+2,1,'P/E (x)  [L&T ~33x]',font=F_LBL)
for j,v in enumerate([25,35,45]): put(rv,tm+2,2+j,v,font=F_BLACK,fmt=FMT_X,fill=FILL_IN,align=RGT,border=B_ALL)
reg(rv,'pex',tm+2)
# C. implied value for BHEL (FY2027E)
ib=tm+4; section(rv,ib,'C.  Implied value for BHEL (FY2027E)',span=6); ib+=1
put(rv,ib,1,'BHEL FY2027E EBITDA',font=F_LBL); put(rv,ib,3,"=%s"%FR('ebitda',9),font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL); reg(rv,'ebitda',ib)
put(rv,ib+1,1,'BHEL FY2027E EPS (INR)',font=F_LBL); put(rv,ib+1,3,"=%s"%FR('eps',9),font=F_GREEN,fmt=FMT_PS,align=RGT,border=B_ALL); reg(rv,'eps',ib+1)
put(rv,ib+2,1,'Net cash (FY26)',font=F_LBL); put(rv,ib+2,3,"=%s-%s"%(FR('cash',8),FR('bor',8)),font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL); reg(rv,'netcash',ib+2)
EVX=REG[RV]['evx']; PEX=REG[RV]['pex']; EB=REG[RV]['ebitda']; EP=REG[RV]['eps']; NC=REG[RV]['netcash']
hp=ib+4
for j,h in enumerate(['Implied price (INR)','Bear','Base','Bull']): put(rv,hp,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
put(rv,hp+1,1,'EV/EBITDA method',font=F_LBL)
for j,col in enumerate(['B','C','D']):
    put(rv,hp+1,2+j,"=(%s%d*$C$%d+$C$%d)/%s"%(col,EVX,EB,NC,XR(ASM,'shares',3)),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)
put(rv,hp+2,1,'P/E method',font=F_LBL)
for j,col in enumerate(['B','C','D']):
    put(rv,hp+2,2+j,"=%s%d*$C$%d"%(col,PEX,EP),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)
put(rv,hp+3,1,'Relative fair value (avg of methods)',font=F_LBLB)
for j,col in enumerate(['B','C','D']):
    put(rv,hp+3,2+j,"=AVERAGE(%s%d,%s%d)"%(col,hp+1,col,hp+2),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=FILL_OK)
reg(rv,'relfv',hp+3); reg(rv,'px_avg',hp+3)
put(rv,hp+5,1,'Memo: sector-parity (peer median EV/EBITDA, no discount)',font=F_ITAL)
put(rv,hp+5,3,"=(C%d*$C$%d+$C$%d)/%s"%(med,EB,NC,XR(ASM,'shares',3)),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)
freeze(rv,'B5')
print("DCF + Relative (quality-discounted) built.")


# ===== 19. SENSITIVITY ANALYSIS (WACC x terminal growth) =====
SEN='Sensitivity Analysis'; se=newsheet(SEN)
put(se,2,1,'Two-way table: intrinsic value per share (INR) vs WACC and terminal growth. Centre = base case.',font=F_NOTE)
fcff_rng="'%s'!$I$12:$O$12"%DCF; per_rng="'%s'!$I$13:$O$13"%DCF; lastf="'%s'!$O$12"%DCF
nd_add="'%s'!$C$24"%DCF; shr="'%s'!$C$26"%DCF
put(se,4,2,'Intrinsic value/share (INR)',font=F_LBLB)
put(se,5,2,'WACC \\ g',font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
for j,gd in enumerate([-0.015,-0.0075,0.0,0.0075,0.015]):
    put(se,5,3+j,"=%s+%s"%(XR(ASM,'term_g',3),gd),font=F_BLACK,fmt=FMT_PCT,fill=FILL_SUB,align=CEN,border=B_ALL)
for i,wd in enumerate([0.02,0.01,0.0,-0.01,-0.02]):
    put(se,6+i,2,"=%s+%s"%(XR(ASM,'wacc',3),wd),font=F_BLACK,fmt=FMT_PCT,fill=FILL_SUB,align=CEN,border=B_ALL)
for i in range(5):
    for j in range(5):
        w="$B%d"%(6+i); g="%s$5"%CL(3+j)
        f="=(SUMPRODUCT(%s,(1+%s)^(-%s))+(%s*(1+%s)/(%s-%s))*(1+%s)^(-7)+%s)/%s"%(fcff_rng,w,per_rng,lastf,g,w,g,w,nd_add,shr)
        put(se,6+i,3+j,f,font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=(FILL_OK if (i==2 and j==2) else None))
put(se,12,1,'Other sensitivities (order inflow, execution, margins, commodities, capex, rates, WC) are on Scenario Analysis and the Assumptions sheet.',font=F_ITAL)
freeze(se,'A1')

# ===== 20. SCENARIO ANALYSIS =====
SC='Scenario Analysis'; sc=newsheet(SC)
put(sc,2,1,'Bull/Base/Bear via exit EV/EBITDA on FY2033E, discounted to present. Base links to the live model.',font=F_NOTE)
for col,w in {'A':36,'B':16,'C':16,'D':16}.items(): sc.column_dimensions[col].width=w
section(sc,3,'Scenario drivers & valuation',span=4)
for j,h in enumerate(['Driver / output','Bear case','Base case','Bull case']):
    put(sc,4,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
def scrow(rr,label,bear,base,bull,fmt=FMT_CR,inputrow=True):
    ah_label(sc,rr,label)
    for j,v in enumerate([bear,base,bull]):
        isf=isinstance(v,str) and v.startswith('=')
        put(sc,rr,2+j,v,font=(F_GREEN if isf else F_BLACK),fmt=fmt,fill=(None if isf else (FILL_IN if inputrow else None)),align=RGT,border=B_ALL)
scrow(5,'FY2033E revenue (INR Cr)',68000,"=%s"%FR('rev',15),92000)
scrow(6,'Terminal EBITDA margin',0.09,"=%s/%s"%(FR('ebitda',15),FR('rev',15)),0.15,fmt=FMT_PCT)
scrow(7,'Exit EV/EBITDA multiple (x)',14,18,24,fmt=FMT_X)
scrow(8,'WACC',0.155,"=%s"%XR(ASM,'wacc',3),0.12,fmt=FMT_PCT)
for lab,rr,frm in [('FY2033E EBITDA (INR Cr)',9,"=%s5*%s6"),('Terminal EV (INR Cr)',10,"=%s9*%s7"),
                   ('PV of EV (discounted 7y)',11,"=%s10/(1+%s8)^7")]:
    ah_label(sc,rr,lab)
    for j in range(3): put(sc,rr,2+j,frm%(CL(2+j),CL(2+j)),font=F_BLUE,fmt=FMT_CR,align=RGT,border=B_ALL)
ah_label(sc,12,'Less: net debt (FY26)')
for j in range(3): put(sc,12,2+j,"=%s-%s"%(FR('bor',8),FR('cash',8)),font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
ah_label(sc,13,'Equity value (INR Cr)')
for j in range(3): put(sc,13,2+j,"=%s11-%s12"%(CL(2+j),CL(2+j)),font=F_BLUE,fmt=FMT_CR,align=RGT,border=B_ALL)
ah_label(sc,14,'Target price (INR)',bold=True)
for j in range(3): put(sc,14,2+j,"=%s13/%s"%(CL(2+j),XR(ASM,'shares',3)),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=FILL_OK)
ah_label(sc,15,'Scenario probability',bold=True)
for j,p in enumerate([0.25,0.50,0.25]): put(sc,15,2+j,p,font=F_BLACK,fmt=FMT_PCT,fill=FILL_IN,align=RGT,border=B_ALL)
ah_label(sc,17,'Probability-weighted target price (INR)',bold=True)
put(sc,17,2,"=SUMPRODUCT(B14:D14,B15:D15)",font=F_BLUE,fmt=FMT_PS,fill=FILL_TOT,align=RGT,border=B_ALL)
freeze(sc,'A1')

# ===== 16. RATIO ANALYSIS (FY20-FY33) =====
RA='Ratio Analysis'; ra=newsheet(RA)
put(ra,2,1,'FY20-25 from Historical Financial Statements; FY26-33 from integrated Forecast statements.',font=F_NOTE)
yhdr(ra,4)
def S(c,key): return lnk(c,key,HFS) if c<8 else FR(key,c)
def EQc(c): return "(%s+%s)"%(lnk(c,'eqc',HFS),lnk(c,'res',HFS)) if c<8 else "(%s+%s)"%(FR('eqc',c),FR('res',c))
def DBc(c): return lnk(c,'bor',HFS) if c<8 else FR('bor',c)
def rax(r,label,fn,fmt,start=2,bold=False):
    ah_label(ra,r,label,bold=bold)
    for c in ACOLS:
        if c<start: continue
        put(ra,r,c,fn(c),font=F_GREEN,fmt=fmt,align=RGT)
section(ra,5,'Growth & margins')
rax(6,'Revenue growth %',lambda c:"=%s/%s-1"%(S(c,'rev'),S(c-1,'rev')),FMT_PCT,start=3)
rax(7,'Gross margin %',lambda c:"=%s/%s"%(S(c,'gross'),S(c,'rev')),FMT_PCT)
rax(8,'EBITDA margin %',lambda c:"=%s/%s"%(S(c,'ebitda'),S(c,'rev')),FMT_PCT)
rax(9,'EBIT margin %',lambda c:"=%s/%s"%(S(c,'ebit'),S(c,'rev')),FMT_PCT)
rax(10,'PAT margin %',lambda c:"=%s/%s"%(S(c,'pat'),S(c,'rev')),FMT_PCT)
section(ra,12,'Returns')
rax(13,'ROE %',lambda c:"=%s/((%s+%s)/2)"%(S(c,'pat'),EQc(c),EQc(c-1)),FMT_PCT,start=3)
rax(14,'ROCE %',lambda c:"=%s/(%s+%s)"%(S(c,'ebit'),EQc(c),DBc(c)),FMT_PCT)
rax(15,'ROA %',lambda c:"=%s/%s"%(S(c,'pat'),S(c,'ta')),FMT_PCT)
rax(16,'Asset turnover (x)',lambda c:"=%s/%s"%(S(c,'rev'),S(c,'ta')),FMT_X)
section(ra,18,'Leverage & coverage')
rax(19,'Debt / equity (x)',lambda c:"=%s/%s"%(DBc(c),EQc(c)),FMT_X)
rax(20,'Interest coverage (x)',lambda c:"=%s/%s"%(S(c,'ebit'),S(c,'fin')),FMT_X)
rax(21,'Net debt / EBITDA (x)',lambda c:"=(%s-%s)/%s"%(DBc(c),FR('cash',c),S(c,'ebitda')),FMT_X,start=8)
section(ra,23,'Per share & payout')
rax(24,'EPS (INR)',lambda c:"=%s"%S(c,'eps'),FMT_PS)
rax(25,'DPS (INR)',lambda c:"=%s/%s"%(S(c,'div'),XR(ASM,'shares',3)),FMT_PS)
rax(26,'Dividend payout %',lambda c:"=IF(%s=0,0,%s/%s)"%(S(c,'pat'),S(c,'div'),S(c,'pat')),FMT_PCT)
rax(27,'Book value per share (INR)',lambda c:"=%s/%s"%(EQc(c),XR(ASM,'shares',3)),FMT_PS)
section(ra,29,'Working capital & cash flow (forecast)')
rax(30,'Receivable days',lambda c:"=%s/%s*365"%(FR('recv',c),FR('rev',c)),FMT_DAYS,start=8)
rax(31,'Inventory days',lambda c:"=%s/%s*365"%(FR('inventory',c),XR(CF,'totcost',c)),FMT_DAYS,start=9)
rax(32,'Payable days',lambda c:"=%s/%s*365"%(FR('pay',c),XR(CF,'totcost',c)),FMT_DAYS,start=9)
rax(33,'FCFF (INR Cr)',lambda c:"=%s"%XR(DCF,'fcff',c),FMT_CR,start=9)
rax(34,'Capex / sales %',lambda c:"=%s/%s"%(XR(ASM,'capex',c),FR('rev',c)),FMT_PCT,start=9)
rax(35,'FCFE (INR Cr)',lambda c:"=%s-%s*(1-%s)+%s+%s"%(XR(DCF,'fcff',c),FR('fin',c),XR(ASM,'tax_rate',c),XR(DBT,'new',c),XR(DBT,'repay',c)),FMT_CR,start=9)
freeze(ra,'B5')
print("Sensitivity + Scenario + Ratio built.")


# ===== 22. ERROR CHECKS =====
EC='Error Checks'; ec=newsheet(EC)
put(ec,2,1,'Automated integrity checks. "OK" = pass (difference < INR 0.5 Cr).',font=F_NOTE)
put(ec,4,1,'Check',font=F_LBLB,fill=FILL_SUB,border=B_ALL)
for i,c in enumerate(ACOLS): put(ec,4,c,(YEARS[i] if c>=8 else ''),font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
def echk(r,label,a,b,start=9):
    ah_label(ec,r,label)
    for c in range(start,16): put(ec,r,c,'=IF(ABS(%s-%s)<0.5,"OK","REVIEW")'%(a(c),b(c)),font=F_BLUE,align=CEN,border=B_ALL)
echk(5,'Balance sheet balances (TA = TE&L)',lambda c:FR('ta',c),lambda c:FR('tle',c),start=2)
echk(6,'Cash: Balance sheet = Cash flow stmt',lambda c:FR('cash',c),lambda c:"'%s'!%s25"%(CFS,CL(c)))
echk(7,'Retained earnings roll-forward ties',lambda c:FR('res',c),lambda c:"'%s'!%s10"%(EQ,CL(c)))
echk(8,'Debt roll-forward ties',lambda c:FR('bor',c),lambda c:"'%s'!%s9"%(DBT,CL(c)))
echk(9,'Depreciation ties to P&L',lambda c:FR('dep',c),lambda c:"'%s'!%s8"%(DS,CL(c)))
echk(10,'Revenue ties to Revenue Build-up',lambda c:FR('rev',c),lambda c:XR(RB,'rev',c))
echk(11,'Net fixed assets tie to FA schedule',lambda c:FR('net_ppe',c),lambda c:"'%s'!%s10"%(FA,CL(c)))
ah_label(ec,13,'Historical balance check max |diff| (FY20-26)')
put(ec,13,3,"=MAX(ABS('%s'!B38),ABS('%s'!C38),ABS('%s'!D38),ABS('%s'!E38),ABS('%s'!F38),ABS('%s'!G38),ABS('%s'!H38))"%tuple([HFS]*7),font=F_BLUE,fmt=FMT_CR1,align=RGT,border=B_ALL)
ah_label(ec,14,'Forecast balance check max |diff|')
put(ec,14,3,"=MAX(ABS(%s),ABS(%s),ABS(%s),ABS(%s),ABS(%s),ABS(%s),ABS(%s))"%tuple(FR('bschk',c) for c in FCOLS),font=F_BLUE,fmt=FMT_CR1,align=RGT,border=B_ALL)
ah_label(ec,15,'Circular references'); put(ec,15,3,'None (interest on opening debt)',font=F_BLACK)
ah_label(ec,16,'Overall model status',bold=True)
put(ec,16,3,'=IF(COUNTIF(B5:O11,"REVIEW")=0,"ALL CHECKS PASS","REVIEW REQUIRED")',font=F_BLUE,bold=True,align=CEN,border=B_ALL,fill=FILL_TOT)
from openpyxl.formatting.rule import CellIsRule
grn=PatternFill('solid',fgColor='C6EFCE'); red=PatternFill('solid',fgColor='FFC7CE')
ec.conditional_formatting.add('B5:O11',CellIsRule(operator='equal',formula=['"OK"'],fill=grn))
ec.conditional_formatting.add('B5:O11',CellIsRule(operator='equal',formula=['"REVIEW"'],fill=red))
ec.conditional_formatting.add('C16:C16',CellIsRule(operator='equal',formula=['"ALL CHECKS PASS"'],fill=grn))
freeze(ec,'B5')

# ===== 21. DASHBOARD =====
DB='Dashboard'; db=wb.create_sheet(DB); db.sheet_view.showGridLines=False
for col,w in {'A':26,'B':16,'C':4,'D':26,'E':16,'F':4,'G':26,'H':16}.items(): db.column_dimensions[col].width=w
put(db,1,1,'BHEL  |  EXECUTIVE DASHBOARD',font=F_HDRW,fill=FILL_HDR)
for c in range(2,9): db.cell(row=1,column=c).fill=FILL_HDR
def tile(r,c,label,formula,fmt):
    put(db,r,c,label,font=F_LBLB,fill=FILL_SUB,border=B_ALL)
    cell=put(db,r+1,c,formula,fmt=fmt,fill=FILL_TOT,border=B_ALL,align=CEN)
    cell.font=Font(name='Calibri',size=14,color='1F3864',bold=True)
PTs="'Price Targets'"
tile(3,1,'DCF value/share - base',"='%s'!C27"%DCF,FMT_PS)
tile(3,2,'Blended fair value - base',"=%s!C17"%PTs,FMT_PS)
tile(3,4,'Current price (INR)',"=%s!C9"%PTs,FMT_PS)
tile(3,5,'12M target - Base',"=%s!D30"%PTs,FMT_PS)
tile(3,7,'12M target - Bull',"=%s!D29"%PTs,FMT_PS)
tile(3,8,'12M target - Bear',"=%s!D31"%PTs,FMT_PS)
tile(6,1,'3M target - Base',"=%s!B30"%PTs,FMT_PS)
tile(6,2,'6M target - Base',"=%s!C30"%PTs,FMT_PS)
tile(6,4,'FY27E EBITDA margin',"=%s/%s"%(FR('ebitda',9),FR('rev',9)),FMT_PCT)
tile(6,5,'FY33E EBITDA margin',"=%s/%s"%(FR('ebitda',15),FR('rev',15)),FMT_PCT)
tile(6,7,'FY26 order book (INR Cr)',"=%s"%XR(RB,'close_ob',8),FMT_CR)
tile(6,8,'WACC',"=%s"%XR(ASM,'wacc',3),FMT_PCT)
put(db,9,1,'DCF bridge (INR Cr)',font=F_SECT)
for i,(lab,frm) in enumerate([('PV of explicit FCFF',"='%s'!C18"%DCF),('PV of terminal value',"='%s'!C22"%DCF),
        ('Enterprise value',"='%s'!C23"%DCF),('Less: net debt',"='%s'!C24"%DCF),('Equity value',"='%s'!C25"%DCF)]):
    put(db,10+i,1,lab,border=B_ALL); put(db,10+i,2,frm,font=F_GREEN,fmt=FMT_CR,align=RGT,border=B_ALL)
c1=LineChart(); c1.title="Revenue & EBITDA (INR Cr)"; c1.height=7.5; c1.width=15; c1.style=10
c1.add_data(Reference(ff,min_col=2,max_col=15,min_row=5,max_row=5),from_rows=True,titles_from_data=False)
c1.add_data(Reference(ff,min_col=2,max_col=15,min_row=12,max_row=12),from_rows=True,titles_from_data=False)
c1.set_categories(Reference(ff,min_col=2,max_col=15,min_row=4,max_row=4)); db.add_chart(c1,"A17")
c2=BarChart(); c2.title="PAT (INR Cr)"; c2.height=7.5; c2.width=15; c2.style=12
c2.add_data(Reference(ff,min_col=2,max_col=15,min_row=20,max_row=20),from_rows=True,titles_from_data=False)
c2.set_categories(Reference(ff,min_col=2,max_col=15,min_row=4,max_row=4)); db.add_chart(c2,"E17")
c3=LineChart(); c3.title="Closing order book (INR Cr)"; c3.height=7.5; c3.width=15; c3.style=11
c3.add_data(Reference(rb,min_col=2,max_col=15,min_row=10,max_row=10),from_rows=True,titles_from_data=False)
c3.set_categories(Reference(rb,min_col=2,max_col=15,min_row=4,max_row=4)); db.add_chart(c3,"A33")
c4=BarChart(); c4.title="FCFF (INR Cr, FY27-33)"; c4.height=7.5; c4.width=15; c4.style=13
c4.add_data(Reference(dc,min_col=9,max_col=15,min_row=12,max_row=12),from_rows=True,titles_from_data=False)
c4.set_categories(Reference(dc,min_col=9,max_col=15,min_row=4,max_row=4)); db.add_chart(c4,"E33")

# ===== football-field valuation-range chart =====
from openpyxl.chart.shapes import GraphicalProperties
put(db,49,1,'Valuation football field (INR/share)  -  bars = value range by method; compare to current price',font=F_SECT)
# chart data table (kept to the right, cols J-M, out of the chart area)
put(db,50,10,'Method',font=F_LBLB); put(db,50,11,'Low',font=F_LBLB); put(db,50,12,'Range',font=F_LBLB); put(db,50,13,'High',font=F_LBLB)
ffrows=[('DCF (bear-bull)',"=%s!B15"%PTs,"=%s!D15"%PTs),
        ('Relative (bear-bull)',"=%s!B16"%PTs,"=%s!D16"%PTs),
        ('Blended fair value',"=%s!B17"%PTs,"=%s!D17"%PTs),
        ('12M target (bear-bull)',"=%s!D31"%PTs,"=%s!D29"%PTs),
        ('Current price',0,"=%s!C9"%PTs)]
for i,(lab,lo,hi) in enumerate(ffrows):
    rr=51+i
    put(db,rr,10,lab,font=F_LBL)
    put(db,rr,11,lo,font=F_GREEN,fmt=FMT_PS,align=RGT)
    put(db,rr,12,"=M%d-K%d"%(rr,rr),font=F_BLUE,fmt=FMT_PS,align=RGT)
    put(db,rr,13,hi,font=F_GREEN,fmt=FMT_PS,align=RGT)
ffc=BarChart(); ffc.type='bar'; ffc.grouping='stacked'; ffc.overlap=100
ffc.title="Valuation football field (INR/share)"; ffc.height=8; ffc.width=17; ffc.style=10
ffc.add_data(Reference(db,min_col=11,min_row=51,max_row=55),titles_from_data=False)   # Low (base, invisible)
ffc.add_data(Reference(db,min_col=12,min_row=51,max_row=55),titles_from_data=False)   # Range (visible)
ffc.set_categories(Reference(db,min_col=10,min_row=51,max_row=55))
_gp=GraphicalProperties(); _gp.noFill=True; ffc.series[0].graphicalProperties=_gp     # hide the base series
ffc.legend=None
db.add_chart(ffc,"A50")

# ===== ASSUMPTIONS RATIONALE (basis for every driver) =====
ART='Assumptions Rationale'; ar=wb.create_sheet(ART); ar.sheet_view.showGridLines=False
ar.column_dimensions['A'].width=36; ar.column_dimensions['B'].width=18; ar.column_dimensions['C'].width=118
put(ar,1,1,'ASSUMPTIONS RATIONALE  -  basis for every forecast driver',font=F_HDRW,fill=FILL_HDR)
for c in range(2,4): ar.cell(row=1,column=c).fill=FILL_HDR
put(ar,2,1,'Each driver documents: why reasonable | historical support | management guidance | macro / industry driver | sensitivity.',font=F_NOTE)
for j,h in enumerate(['Driver','Base case','Rationale (historical support | management guidance | macro/industry driver | sensitivity)']):
    put(ar,3,1+j,h,font=F_LBLB,fill=FILL_SUB,border=B_ALL,align=(CEN if j<2 else LFT) if 'LFT' in dir() else None)
WRAP=Alignment(wrap_text=True,vertical='top')
arr=[4]
def aR(driver,val,rat,hdr=False):
    r=arr[0]
    if hdr:
        put(ar,r,1,driver,font=F_SECT)
        for c in range(1,4): ar.cell(row=r,column=c).fill=PatternFill('solid',fgColor='EAF0FA')
        arr[0]+=1; return
    put(ar,r,1,driver,font=F_LBLB); put(ar,r,2,val,font=F_BLACK,align=CEN)
    cell=put(ar,r,3,rat,font=F_BLACK); cell.alignment=WRAP
    ar.row_dimensions[r].height=46
    for c in range(1,4): ar.cell(row=r,column=c).border=B_ALL
    arr[0]+=1
aR('REVENUE & ORDER BOOK','','',hdr=True)
aR('Execution rate','16.5%->19.5%','Revenue = execution rate x opening order book. Implies ~13.5% revenue CAGR - within management guidance of 12-15% CAGR (Q4FY24 concall). BHEL has ~10 GW/yr execution capacity (demonstrated 12 GW); 80+ GW thermal to be ordered by 2032. Historically executed ~15-17% of opening OB.')
aR('Order inflow','Rs78k->96k cr','FY25 record inflow Rs92,535 cr; FY26 actual Rs75,916 cr (Q4FY26 supplementary); FY26 closing order book Rs2,40,000 cr - highest ever (Power 81% / Industry 18% / Export 1%), CONFIRMING the model driver. ~80 GW thermal to be ordered by 2032 (~10-12 GW/yr) plus T&D, railways, defence, nuclear.')
aR('Revenue mix - Power','75%','BHEL core; power was ~75-80% of revenue historically (FY26 power Rs25,407 cr). Thermal ordering remains dominant near-term.')
aR('Revenue mix - Industry','22%','Rising diversification (industrial systems, transportation, defence). Industry segment grew strongly in FY25-26; balance ~3% is exports/renewables.')
aR('COSTS','','',hdr=True)
aR('EBITDA margin','7.5%->12.0%','Audited historical margin (annual reports): FY24 2.6%, FY25 4.4%, FY26 6.9% (rising as revenue scales over a largely fixed cost base). Expansion from operating leverage, better project mix and roll-off of legacy low-margin orders. Management targets low-teens; peer capital-goods margins 10-18%. Sensitivity: 1pp ~ Rs400-800 cr EBITDA.')
aR('Steel (% rev)','12%','Boilers, turbines and structurals are steel-intensive. Key commodity exposure; +10% steel price ~ +1.2pp COGS. Reflects BHEL bill-of-materials.')
aR('Copper (% rev)','5%','Generators, transformers and windings are copper-intensive. Exposed to LME copper prices.')
aR('Electrical & electronic components','10%','Control systems, switchgear, instrumentation and electronics content in power equipment.')
aR('Imported components','8%','High-tech sub-assemblies imported; FX and import-duty sensitive. Localisation drive aims to reduce this over time.')
aR('Other materials & consumables','residual','Balancing item so that materials + subcontracting + opex reconcile to the target EBITDA margin (avoids double-counting).')
aR('Subcontracting & erection','8.5%','Civil works and site erection largely outsourced; scales with project execution. Historical ~8-9%.')
aR('Employee cost (% rev)','18.5%->14.0%','Audited employee benefits (annual reports): Rs5,432 cr (FY20) -> Rs6,468 cr (FY26); as a % of revenue FY24 23.6%, FY25 20.9%, FY26 19.1%. Structurally high & broadly fixed; falls as a % as revenue scales and workforce declines (~33,500->25,500). Forecast starts at the FY26 actual (~19%) and eases toward 14%.')
aR('Selling & distribution','2.5%','Bid/marketing, logistics, liquidated-damages and warranty-related selling costs; consistent with historical SG&A composition.')
aR('Administrative (% rev)','~2.5% (residual)','Corporate overheads; residual within total selling/admin so opex reconciles.')
aR('Research & development','2.5%','BHEL discloses R&D at ~2.5% of turnover; strategic for supercritical/USC, hydrogen/electrolysers, defence and localisation.')
aR('Depreciation rate','9.5% of opening net block','Consistent with historical D&A / net block (~9-10%); plant & machinery weighted asset base.')
aR('Capex','Rs600-950 cr','Modernisation plus new capacity for renewables, defence and electrolysers. Historical capex Rs300-900 cr; internally funded.')
aR('WORKING CAPITAL','','',hdr=True)
aR('Receivable days (vs revenue)','74 -> 68','Audited history 49-121 (FY26 73). Tapered down per management WC-improvement focus (Q4FY24 concall): legacy dues clearing within a quarter of milestone completion.')
aR('Inventory days (vs COGS)','210 -> 192','Audited history 151-232 (FY26 212); long project cycle. Tapered per management guidance on better execution & supply-chain/vendor normalisation (Q4FY24 concall).')
aR('Payable days (vs COGS)','167->160','Audited history 167-230 days (FY26 167); computed on COGS. Held near the FY26 level with a slight tightening.')
aR('Contract assets (total) / Revenue','83% -> 65%','Full sell-side driver derivation in the Contract Accounting Schedule (Ind AS 115): 5 candidate drivers tested by coefficient of variation across FY20-26; CA/Revenue selected (CV 0.139, lowest, and economically sound - CA is unbilled revenue under %-of-completion accounting, so it tracks revenue recognised, not the largely-unexecuted order book). Tapers from the FY26 actual (87%) toward a steady state, continuing the FY24-26 improvement trend (Q4FY24 concall + Q4FY26 supplementary: contract assets flat YoY despite +19% revenue) at a moderated pace. Current/non-current split (52%/48%) = recent 3yr average, reflecting a post-FY23 structural shift.')
aR('Contract liabilities (total) / Order book','9.5% -> 10.5%','Full sell-side driver derivation in the Contract Accounting Schedule (Ind AS 115): CL/Closing-order-book selected (CV 0.178, lowest among order-book/inflow-based measures, and economically sound - CL is customer advances tied to UNEXECUTED backlog, not to revenue already recognised). Continues the FY24-26 rise (5.4%->7.8%->9.4%) per management guidance on improved advance/milestone payment terms, plateauing modestly above the FY26 level. Current/non-current split (40%/60%) = recent 3yr average.')
aR('Other non-current assets (residual)','Rs6,738 cr flat','Non-CONTRACT residual only (deferred tax assets ~Rs3,533 cr, long-term/legacy trade receivables ~Rs2,427 cr, other financial assets/deposits). Non-current CONTRACT assets are modelled separately and explicitly in the Contract Accounting Schedule (see above) and added back on the balance sheet. Residual held flat at the FY26 audited level - no clear multi-year trend in DTA/legacy receivables to support a growth assumption.')
aR('Other non-current liabilities (residual)','Rs4,155 cr flat','Non-CONTRACT residual only (long-term provisions - warranty & employee benefits ~Rs2,355 cr - and other). Non-current CONTRACT liabilities are modelled separately and explicitly in the Contract Accounting Schedule (see above) and added back on the balance sheet. Residual held flat at the FY26 audited level - no clear multi-year trend to support a growth assumption.')
aR('Other current operating assets/liabilities growth','3% p.a.','Remaining minor current operating items (excl. contract assets/liabilities, which now have their own driver) grow slowly, not 1:1 with revenue, to avoid overstating working-capital drag.')
aR('CWIP / Investments','Rs400 cr / Rs310-370 cr','CWIP steady-state low (capex flows quickly to assets); investments = JV/associate stakes, modest growth.')
aR('CAPITAL STRUCTURE & TAX','','',hdr=True)
aR('LT vs ST borrowings','~2% LT / 98% ST','Audited: BHEL borrowings are almost entirely short-term working-capital lines. The only "long-term" borrowing is lease liabilities (~Rs170 cr in FY26 = 2.05% of total; FY20-26 range 0.3-2.1%). Re-anchored from the prior arbitrary 35% split to the audited actual ~2%.')
aR('New borrowings / repayment','held flat','BHEL borrowings are short-term working-capital lines (not a term loan) - no fixed repayment schedule. Held flat: the large cash balance (net cash ~Rs3,700 cr) funds working-capital growth. The prior Rs1,000 cr/yr repayment was an unsupported placeholder. Financing choices do not affect FCFF/EV (pre-financing).')
aR('Interest rate','8.5%','Blended cost on borrowings; consistent with historical finance cost / average debt and the current rate environment.')
aR('Tax rate','25%','BHEL on the new corporate-tax regime (~25.17%); normalised to 25%.')
aR('Dividend payout','30%','CPSE/DIPAM policy (~30% of PAT or 5% of net worth); historical payout 30-33%.')
aR('VALUATION (DCF)','','',hdr=True)
aR('Risk-free rate','6.8%','~India 10-year G-Sec yield.')
aR('Beta','1.25','BHEL is a high-beta cyclical capital-goods PSU; levered beta > 1.')
aR('Equity risk premium','6.5%','India ERP ~6-7%.')
aR('Cost of debt / weights','8.5% / 15% D, 85% E','Low leverage; target structure mostly equity. After-tax Kd used in WACC.')
aR('Terminal growth','4.5%','Below long-run nominal GDP (~10-11%); conservative perpetuity. Terminal FCFF is NORMALIZED so working capital grows at g (steady state), not at the FY33 high-growth rate - otherwise the terminal value understates value because the last explicit year still carries a large growth-driven WC investment.')
aR('WACC (derived)','~13.6%','We*Ke + Wd*Kd*(1-t); high due to elevated equity beta. Tested on the Sensitivity sheet vs terminal growth.')
freeze(ar,'A4')

# ===== PRICE TARGETS (blended valuation + 3/6/12-month scenarios) =====
PT='Price Targets'; pt=newsheet(PT)
put(pt,2,1,'Blended valuation & scenario price targets. Fundamental fair value = 50% DCF + 50% quality-adjusted relative. Horizon targets apply scenario EV/EBITDA multiples to forward EBITDA (3M/6M on FY2027E, 12M rolled to FY2028E) plus net cash, per share. Multiples (gold) are editable. Model output, not investment advice.',font=F_NOTE)
for col,w in {'A':42,'B':13,'C':13,'D':13}.items(): pt.column_dimensions[col].width=w
section(pt,4,'A.  Inputs (live links)',span=4)
def ptin(r,label,formula,fmt=FMT_CR,key=None):
    put(pt,r,1,label,font=F_LBL); put(pt,r,3,formula,font=F_GREEN,fmt=fmt,align=RGT,border=B_ALL)
    if key: reg(pt,key,r)
ptin(5,'FY2027E EBITDA (INR Cr)',"=%s"%FR('ebitda',9),key='eb27')
ptin(6,'FY2028E EBITDA (INR Cr)',"=%s"%FR('ebitda',10),key='eb28')
ptin(7,'Net cash - FY26 (INR Cr)',"=%s-%s"%(FR('cash',8),FR('bor',8)),key='nc')
ptin(8,'Shares outstanding (crore)',"=%s"%XR(ASM,'shares',3),fmt=FMT_CR1,key='sh')
ptin(9,'Current market price (INR)',"=%s"%XR(ASM,'price',3),fmt=FMT_PS,key='px')
ptin(10,'DCF value/share - base (INR)',"=%s"%XR(DCF,'ivps',3),fmt=FMT_PS,key='dcf')
put(pt,11,1,'Relative fair value - Bear/Base/Bull (INR)',font=F_LBL)
for j,col in enumerate(['B','C','D']):
    put(pt,11,2+j,"='%s'!%s%d"%(RV,col,REG[RV]['relfv']),font=F_GREEN,fmt=FMT_PS,align=RGT,border=B_ALL)
NCr=REG[PT]['nc']; SHr=REG[PT]['sh']; EB27=REG[PT]['eb27']; EB28=REG[PT]['eb28']; PXr=REG[PT]['px']; DCFr=REG[PT]['dcf']; RELr=11
section(pt,13,'B.  Fundamental fair value (12-month)',span=4)
for j,h in enumerate(['','Bear','Base','Bull']): put(pt,14,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
put(pt,15,1,'DCF (INR/share)',font=F_LBL)
put(pt,15,2,"=$C$%d*0.81"%DCFr,font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)   # bear = flat-WC sensitivity
put(pt,15,3,"=$C$%d"%DCFr,font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)        # base = live DCF
put(pt,15,4,"=$C$%d*1.20"%DCFr,font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)   # bull = aggressive-WC sensitivity
put(pt,16,1,'Relative (INR/share)',font=F_LBL)
for j,col in enumerate(['B','C','D']): put(pt,16,2+j,"=%s%d"%(col,RELr),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL)
put(pt,17,1,'Blended fair value (50/50)',font=F_LBLB)
for j,col in enumerate(['B','C','D']): put(pt,17,2+j,"=AVERAGE(%s15,%s16)"%(col,col),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=FILL_OK)
reg(pt,'blend',17)
put(pt,18,1,'  Upside/(downside) vs current price',font=F_ITAL)
for j,col in enumerate(['B','C','D']): put(pt,18,2+j,"=%s17/$C$%d-1"%(col,PXr),font=F_BLUE,fmt=FMT_PCT,align=RGT,border=B_ALL)
section(pt,20,'C.  Scenario EV/EBITDA multiple (x) - editable',span=4)
for j,h in enumerate(['Scenario','3-month','6-month','12-month']): put(pt,21,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
for nm,rr,vals in [('Bull',22,[45,46,44]),('Base',23,[43,40,36]),('Bear',24,[38,32,26])]:
    put(pt,rr,1,nm,font=F_LBL)
    for j,v in enumerate(vals): put(pt,rr,2+j,v,font=F_BLACK,fmt=FMT_X,fill=FILL_IN,align=RGT,border=B_ALL)
put(pt,25,1,'EBITDA basis',font=F_ITAL)
for j,v in enumerate(['FY2027E','FY2027E','FY2028E']): put(pt,25,2+j,v,font=F_ITAL,align=CEN,border=B_ALL)
section(pt,27,'D.  Scenario price target (INR/share)',span=4)
for j,h in enumerate(['Scenario','3-month','6-month','12-month']): put(pt,28,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
for nm,rr,mr in [('Bull',29,22),('Base',30,23),('Bear',31,24)]:
    fillb=FILL_OK if nm=='Base' else None
    put(pt,rr,1,nm,font=(F_LBLB if nm=='Base' else F_LBL))
    put(pt,rr,2,"=(B%d*$C$%d+$C$%d)/$C$%d"%(mr,EB27,NCr,SHr),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=fillb)
    put(pt,rr,3,"=(C%d*$C$%d+$C$%d)/$C$%d"%(mr,EB27,NCr,SHr),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=fillb)
    put(pt,rr,4,"=(D%d*$C$%d+$C$%d)/$C$%d"%(mr,EB28,NCr,SHr),font=F_BLUE,fmt=FMT_PS,align=RGT,border=B_ALL,fill=fillb)
section(pt,33,'E.  Upside/(downside) vs current price',span=4)
for j,h in enumerate(['Scenario','3-month','6-month','12-month']): put(pt,34,1+j,h,font=F_LBLB,fill=FILL_SUB,align=CEN,border=B_ALL)
for nm,rr,pr in [('Bull',35,29),('Base',36,30),('Bear',37,31)]:
    put(pt,rr,1,nm,font=F_LBL)
    for j,col in enumerate(['B','C','D']): put(pt,rr,2+j,"=%s%d/$C$%d-1"%(col,pr,PXr),font=F_BLUE,fmt=FMT_PCT,align=RGT,border=B_ALL)
put(pt,39,1,'Verdict: fundamental fair value ~Rs110-190 (base ~Rs145) implies BHEL is expensive vs the market price; the price is sustained by sector-premium multiples. Bull requires strong growth to materialise AND premium multiples to persist; Bear reflects de-rating toward fundamentals.',font=F_NOTE)
freeze(pt,'B5')
print("Price Targets built.")

# ===== finalise: tab order, colours, save =====
order=['Cover Page','Assumptions','Assumptions Rationale','Historical Financial Statements','Historical Ratio Analysis',
 'Operational Drivers','Revenue Build-up','Contract Accounting Schedule','Cost Forecast','Working Capital Schedule','Fixed Asset Schedule',
 'Depreciation Schedule','Debt Schedule','Equity Schedule','Other Assets & Liabilities','Cash Flow Statement',
 'Forecast Financial Statements','Ratio Analysis','DCF Valuation','Relative Valuation','Sensitivity Analysis',
 'Scenario Analysis','Price Targets','Dashboard','Error Checks']
wb._sheets.sort(key=lambda s: order.index(s.title))
for t in ['Cover Page','Dashboard']: wb[t].sheet_properties.tabColor=NAVY
for t in ['DCF Valuation','Relative Valuation','Scenario Analysis','Price Targets']: wb[t].sheet_properties.tabColor='548235'
wb['Error Checks'].sheet_properties.tabColor='C00000'; wb['Assumptions'].sheet_properties.tabColor='BF8F00'
wb.active=wb._sheets.index(wb['Dashboard'])
wb.save('/projects/sandbox/BHEL/BHEL_Financial_Model.xlsx')
print("SAVED.")
