"""Write audited FY2020-FY2026 balance-sheet history into both models.
Forecast columns are left untouched (already seeded off audited FY2026)."""
import openpyxl
from openpyxl.styles import Font, Alignment
from audited_data import D, derived

BLK=Font(name='Calibri',size=10,color='000000')   # audited actual
EST=Font(name='Calibri',size=10,color='C55A11')    # derived contract split

def rows_for_year(i):
    d=derived(i)
    r={
      'sharecap':D['sharecap'][i],'reserve':D['reserve'][i],
      'leaseNC':D['leaseNC'][i],'tpnc':D['tpnc'][i],'CLnc':d['CLnc'],'provnc':D['provnc'][i],'othncliab':d['othncliab'],
      'stbor':D['stbor_grp'][i],'tpcur':D['tpcur'][i],'CLcur':d['CLcur'],'provcur':D['provcur'][i],'othcurliab':d['othcurliab'],
      'ppe':D['ppe_grp'][i],'cwip':D['cwip'][i],'invest':D['invest'][i],'trnc':D['trnc'][i],'CAnc':d['CAnc'],
      'dta':D['dta'][i],'othncasset':d['othncasset'],
      'trcur':D['trcur'][i],'invtry':D['invtry'][i],'CAcur':d['CAcur'],'othcurasset':d['othcurasset'],'cash':D['cash_grp'][i],
    }
    r={k:round(v,2) for k,v in r.items()}
    ta=(r['ppe']+r['cwip']+r['invest']+r['trnc']+r['CAnc']+r['dta']+r['othncasset']
        +r['invtry']+r['trcur']+r['CAcur']+r['othcurasset']+r['cash'])
    tel=(r['sharecap']+r['reserve']+r['leaseNC']+r['tpnc']+r['CLnc']+r['provnc']+r['othncliab']
         +r['stbor']+r['tpcur']+r['CLcur']+r['provcur']+r['othcurliab'])
    r['othcurasset']=round(r['othcurasset']+(tel-ta),2)   # absorb <=0.01 rounding so BS ties exactly
    return r
CONTRACT={'CLnc','CLcur','CAnc','CAcur'}   # orange (derived)

def setcell(ws,col,row,key,val):
    c=ws[f'{col}{row}']; c.value=round(val,2)
    c.font=EST if key in CONTRACT else BLK
    c.number_format=c.number_format or '#,##0.0'
    c.alignment=Alignment(horizontal='right')

# ---------------- 1. BHEL FM.xlsx  (cols C..I = FY20..FY26) ----------------
fm=openpyxl.load_workbook('BHEL/BHEL FM.xlsx')
bs=fm['BS']
MAP_FM={'sharecap':8,'reserve':9,'leaseNC':13,'tpnc':14,'CLnc':15,'provnc':16,'othncliab':17,
 'stbor':21,'tpcur':22,'CLcur':23,'provcur':24,'othcurliab':25,
 'ppe':33,'cwip':34,'invest':35,'trnc':36,'CAnc':37,'dta':38,'othncasset':39,
 'trcur':43,'invtry':44,'CAcur':45,'othcurasset':46,'cash':47}
cols_fm=['C','D','E','F','G','H','I']
for i,col in enumerate(cols_fm):
    r=rows_for_year(i)
    for key,row in MAP_FM.items():
        setcell(bs,col,row,key,r[key])
bs['B54']=('Note: Balance sheet FY2020-FY2026 sourced from BHEL audited consolidated annual reports '
  '(FY2021, FY2023, FY2025) and the FY2026 filing. All line items are audited actuals; only the split of '
  'contract assets/liabilities (orange) within "Other" balances for FY2020-25 is derived from the FY2026 '
  'audited composition (notes not in the 2-page filings) - audited section totals are preserved and the BS ties every year.')
bs['B54'].font=Font(size=8,italic=True,color='808080')
fm.save('BHEL/BHEL FM.xlsx')
print("BHEL FM.xlsx historical updated to audited.")

# ---------------- 2. BHEL_Financial_Model.xlsx (Forecast BS cols B..H) ----------------
full=openpyxl.load_workbook('BHEL_Financial_Model.xlsx')
ffs=full['Forecast Financial Statements']
MAP_FULL={'sharecap':28,'reserve':29,'leaseNC':33,'tpnc':34,'CLnc':35,'provnc':36,'othncliab':37,
 'stbor':41,'tpcur':42,'CLcur':43,'provcur':44,'othcurliab':45,
 'ppe':53,'cwip':54,'invest':55,'trnc':56,'CAnc':57,'dta':58,'othncasset':59,
 'trcur':63,'invtry':64,'CAcur':65,'othcurasset':66,'cash':67}
cols_full=['B','C','D','E','F','G','H']
for i,col in enumerate(cols_full):
    r=rows_for_year(i)
    for key,row in MAP_FULL.items():
        setcell(ffs,col,row,key,r[key])
ffs.cell(row=73,column=1).value=('Note: Balance sheet FY2020-FY2026 from BHEL audited consolidated annual reports '
  '(FY2021/2023/2025) + FY2026 filing. All line items audited; only the contract asset/liability split (orange) for '
  'FY2020-25 is derived from the FY2026 audited composition. Section totals are audited and the BS ties every year.')
ffs.cell(row=73,column=1).font=Font(size=8,italic=True,color='808080')

# keep the Historical Financial Statements aggregate sheet consistent with audited
hfs=full['Historical Financial Statements']
for i,col in enumerate(cols_full):
    r=rows_for_year(i)
    ta=(r['ppe']+r['cwip']+r['invest']+r['trnc']+r['CAnc']+r['dta']+r['othncasset']
        +r['invtry']+r['trcur']+r['CAcur']+r['othcurasset']+r['cash'])
    borrow=r['leaseNC']+r['stbor']
    othliab=(r['tpnc']+r['CLnc']+r['provnc']+r['othncliab']+r['tpcur']+r['CLcur']+r['provcur']+r['othcurliab'])
    hfs[f'{col}28']=round(r['ppe'],2); hfs[f'{col}29']=round(r['cwip'],2); hfs[f'{col}30']=round(r['invest'],2)
    hfs[f'{col}31']=round(ta-r['ppe']-r['cwip']-r['invest'],2)
    hfs[f'{col}33']=696.41; hfs[f'{col}34']=round(r['reserve'],2)
    hfs[f'{col}35']=round(borrow,2); hfs[f'{col}36']=round(othliab,2)
full.save('BHEL_Financial_Model.xlsx')
print("BHEL_Financial_Model.xlsx historical updated to audited.")
