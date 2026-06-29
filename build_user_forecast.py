# -*- coding: utf-8 -*-
"""Fill FY2027-FY2033 forecast into the user's BHEL FM.xlsx (PL & BS) from their
actuals. Assumptions split into VARIABLE (Base/Best/Bear, driven by a Scenario
Switch dropdown) and FIXED. Adds Assumptions, Forecast Engine, Discrepancies."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL
from openpyxl.worksheet.datavalidation import DataValidation

PATH='/projects/sandbox/BHEL/BHEL FM.xlsx'
wb=openpyxl.load_workbook(PATH)
FC=list(range(10,17)); YR=[2027,2028,2029,2030,2031,2032,2033]   # forecast cols J..P
BLUE=Font(color='0000CC'); BLACK=Font(color='000000'); BOLD=Font(bold=True)
GOLD=PatternFill('solid',fgColor='FFF2CC'); SUB=PatternFill('solid',fgColor='D9E1F2')
TOT=PatternFill('solid',fgColor='DDEBF7'); LTBL=PatternFill('solid',fgColor='E7EEF8')
RGT=Alignment(horizontal='right'); CEN=Alignment(horizontal='center')
WRAP=Alignment(wrap_text=True,vertical='top')
thin=Side(style='thin',color='BFBFBF'); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
NF='#,##0.0;(#,##0.0)'; PF='0.0%'; DF='0'

# ===================================================================
# 1. ASSUMPTIONS  (variable scenarios + fixed)
# ===================================================================
if 'Assumptions' in wb.sheetnames: del wb['Assumptions']
asm=wb.create_sheet('Assumptions'); asm.sheet_view.showGridLines=False
asm.column_dimensions['A'].width=2; asm.column_dimensions['B'].width=44; asm.column_dimensions['C'].width=11
for c in FC: asm.column_dimensions[CL(c)].width=10
asm.column_dimensions[CL(17)].width=66
asm.cell(1,2,'FORECAST ASSUMPTIONS  -  variable (scenario-driven) & fixed').font=Font(bold=True,size=13,color='1F3864')
asm.cell(3,2,'Scenario Switch').font=Font(bold=True,size=11)
sw=asm.cell(3,3,'Base'); sw.font=Font(bold=True,color='C00000'); sw.fill=GOLD; sw.alignment=CEN; sw.border=BOX
dv=DataValidation(type='list',formula1='"Base,Best,Bear"',allow_blank=False); asm.add_data_validation(dv); dv.add(sw)
asm.cell(3,5,'<- toggle Base / Best / Bear and the whole PL, BS & engine recalculate').font=Font(italic=True,size=9,color='808080')
# active-scenario quick outputs (live from PL)
asm.cell(4,5,'Active scenario FY2033E:').font=Font(italic=True,size=9,color='1F3864')
asm.cell(4,7,"=\"Sales \"&TEXT('PL'!P5,\"#,##0\")&\"  |  EBITDA \"&TEXT('PL'!P12,\"#,##0\")&\"  |  PAT \"&TEXT('PL'!P20,\"#,##0\")").font=Font(italic=True,size=9,color='1F3864')
SW='$C$3'; A={}
def yhdr(r):
    asm.cell(r,2,'ratios / INR Crore').font=Font(italic=True,size=9)
    for j,c in enumerate(FC):
        cell=asm.cell(r,c,YR[j]); cell.font=BOLD; cell.fill=SUB; cell.alignment=CEN; cell.border=BOX
    asm.cell(r,17,'Rationale').font=BOLD
asm.cell(6,2,'A.  VARIABLE ASSUMPTIONS  (Base / Best / Bear, selected by Scenario Switch)').font=Font(bold=True,color='1F3864')
yhdr(7); R=[8]
def vblock(key,label,base,best,bear,fmt,rat):
    r=R[0]
    asm.cell(r,2,label).font=Font(bold=True,size=10)
    for j,c in enumerate(FC):
        f='=IF(%s="Best",%s%d,IF(%s="Bear",%s%d,%s%d))'%(SW,CL(c),r+2,SW,CL(c),r+3,CL(c),r+1)
        cell=asm.cell(r,c,f); cell.font=Font(bold=True,color='0000CC'); cell.number_format=fmt; cell.alignment=RGT; cell.fill=LTBL; cell.border=BOX
    rc=asm.cell(r,17,rat); rc.font=Font(size=9,italic=True); rc.alignment=WRAP; asm.row_dimensions[r].height=40
    for jj,(nm,vals) in enumerate([('Base',base),('Best',best),('Bear',bear)]):
        r2=r+1+jj; asm.cell(r2,2,'   '+nm).font=Font(size=9,color='808080')
        for j,c in enumerate(FC):
            cell=asm.cell(r2,c,vals[j]); cell.font=BLACK; cell.fill=GOLD; cell.number_format=fmt; cell.alignment=RGT; cell.border=BOX
    A[key]=r; R[0]=r+5
vblock('growth','Revenue growth %',
 [0.15,0.14,0.13,0.12,0.11,0.10,0.09],[0.18,0.17,0.16,0.14,0.13,0.12,0.11],[0.10,0.09,0.08,0.07,0.06,0.06,0.05],PF,
 'Demand driver. Base = order-book-supported double digits tapering; Best = faster execution/new ordering; Bear = slower power capex & execution delays.')
vblock('gross','Gross margin %',
 [0.325,0.330,0.335,0.340,0.340,0.345,0.350],[0.335,0.345,0.355,0.360,0.365,0.370,0.375],[0.310,0.310,0.315,0.315,0.320,0.320,0.325],PF,
 'Margin driver. Actual FY26 32.0%. Best = favourable mix + benign steel/copper; Bear = commodity inflation & competitive pricing.')
vblock('ebitda','EBITDA margin %',
 [0.080,0.090,0.100,0.110,0.115,0.120,0.125],[0.095,0.110,0.120,0.130,0.135,0.140,0.145],[0.065,0.070,0.078,0.085,0.090,0.095,0.100],PF,
 'Operating leverage. Actual FY26 6.9%. Best = strong leverage to low-/mid-teens; Bear = limited leverage if growth disappoints.')
vblock('recv_days','Receivable days',
 [73,73,73,72,72,72,72],[70,68,66,65,64,63,62],[80,82,84,85,86,86,86],DF,
 'Collections cycle. Best = faster realisation from PSU/SEB customers; Bear = stretched receivables in a weak cycle.')
vblock('inv_days','Inventory days (on COGS)',
 [210,205,202,200,198,195,193],[195,188,182,178,175,172,170],[225,225,222,220,218,215,213],DF,
 'Project inventory cycle. Best = tighter planning/execution; Bear = slow-moving project inventory build-up.')
# ---- fixed ----
P=[R[0]+1]
asm.cell(P[0],2,'B.  FIXED ASSUMPTIONS  (structural / policy - constant across scenarios)').font=Font(bold=True,color='1F3864'); P[0]+=1
yhdr(P[0]); P[0]+=1
def frow(key,label,vals,fmt,rat):
    r=P[0]; asm.cell(r,2,label).font=Font(size=10)
    for j,c in enumerate(FC):
        cell=asm.cell(r,c,vals[j]); cell.font=BLACK; cell.fill=GOLD; cell.number_format=fmt; cell.alignment=RGT; cell.border=BOX
    rc=asm.cell(r,17,rat); rc.font=Font(size=9,italic=True); rc.alignment=WRAP; asm.row_dimensions[r].height=28
    A[key]=r; P[0]+=1
frow('emp','Employee cost % of sales',[0.185,0.180,0.175,0.170,0.165,0.160,0.155],PF,'Largely fixed cost base; declines as % as sales scale (FY26 19.1%).')
frow('oth_inc','Other income % of sales',[0.022]*7,PF,'Treasury income on cash (~2.2%).')
frow('pay_days','Payable days (on COGS)',[167]*7,DF,'Supplier credit ~167 days (FY26 actual).')
frow('oca_pct','Other current assets % of sales',[0.57,0.56,0.55,0.54,0.53,0.52,0.51],PF,'Unbilled/contract assets & advances; ~58% of sales declining.')
frow('ocl_pct','Other current liabilities % of sales',[0.40,0.40,0.39,0.39,0.38,0.38,0.37],PF,'Customer advances & provisions; key WC funding (~41%).')
frow('onca_g','Other non-current assets growth %',[0.03]*7,PF,'Deferred tax assets, legacy LT receivables; slow growth.')
frow('oncl_g','Other non-current liabilities growth %',[0.04]*7,PF,'Long-term advances/deferred; moderate growth.')
frow('ltprov_g','LT provisions growth %',[0.03]*7,PF,'Warranty & employee-benefit provisions.')
frow('capex','Capital expenditure (Rs cr)',[600,700,750,800,850,900,950],NF,'Modernisation + renewables/defence capacity; internally funded.')
frow('dep_rate','Depreciation % of opening net block',[0.095]*7,PF,'~ historical D&A / net block (9-10%).')
frow('cwip','CWIP closing (Rs cr)',[400]*7,NF,'Steady-state low.')
frow('invst','Investments closing (Rs cr)',[310,320,330,340,350,360,370],NF,'JV/associate stakes.')
frow('lt_pct','LT borrowings % of total debt',[0.02]*7,PF,'BHEL debt almost entirely short-term (FY26 LT only Rs168 cr).')
frow('repay','Debt repayment (Rs cr)',[500]*7,NF,'Gradual reduction of WC borrowings from strong cash.')
frow('int_rate','Interest rate on opening debt %',[0.085]*7,PF,'Blended ~ FY26 finance cost / debt.')
frow('tax','Effective tax rate %',[0.25]*7,PF,'New corporate-tax regime (~25%).')
frow('payout','Dividend payout % of PAT',[0.30]*7,PF,'CPSE/DIPAM policy ~30%.')
# ---- WACC singles ----
P[0]+=1; asm.cell(P[0],2,'C.  WACC & VALUATION (fixed)').font=Font(bold=True,color='1F3864'); P[0]+=1
def single(key,label,val,fmt):
    r=P[0]; asm.cell(r,2,label).font=Font(size=10)
    cell=asm.cell(r,3,val); cell.font=BLACK; cell.fill=GOLD; cell.number_format=fmt; cell.alignment=RGT; cell.border=BOX
    A[key]=('s',r); P[0]+=1
single('rf','Risk-free rate',0.068,PF); single('beta','Levered beta',1.25,'0.00')
single('erp','Equity risk premium',0.065,PF); single('kd','Pre-tax cost of debt',0.085,PF)
single('wd','Weight of debt',0.10,PF); single('we','Weight of equity',0.90,PF)
single('tg','Terminal growth',0.045,PF)
asm.cell(P[0],2,'Cost of equity').font=Font(size=10,bold=True)
asm.cell(P[0],3,"=C%d+C%d*C%d"%(A['rf'][1],A['beta'][1],A['erp'][1])).number_format=PF; A['ke']=('s',P[0]); P[0]+=1
asm.cell(P[0],2,'WACC').font=Font(size=10,bold=True)
asm.cell(P[0],3,"=C%d*C%d+C%d*C%d*0.75"%(A['we'][1],A['ke'][1],A['wd'][1],A['kd'][1])).number_format=PF; A['wacc']=('s',P[0]); P[0]+=1
single('shares','Shares outstanding (cr)',348.2,NF); single('price','Current price (Rs)',402.7,'0.00')
print("Assumptions built.")

# ===================================================================
# 2. FORECAST ENGINE
# ===================================================================
if 'Forecast Engine' in wb.sheetnames: del wb['Forecast Engine']
fe=wb.create_sheet('Forecast Engine'); fe.sheet_view.showGridLines=False
fe.column_dimensions['A'].width=2; fe.column_dimensions['B'].width=34
for c in range(9,17): fe.column_dimensions[CL(c)].width=11
fe.cell(1,2,'FORECAST ENGINE  (2026 seed = actuals; 2027-2033 driven by Assumptions/Scenario)').font=Font(bold=True,size=12,color='1F3864')
for j,c in enumerate(range(9,17)):
    cell=fe.cell(3,c,2026+j); cell.font=BOLD; cell.fill=SUB; cell.alignment=CEN; cell.border=BOX
def ad(key,c): return "'Assumptions'!%s%d"%(CL(c),A[key])
def E(r,c,f):
    cell=fe.cell(r,c,f); cell.font=BLUE; cell.number_format=NF; cell.alignment=RGT
def lbl(r,t,b=False): fe.cell(r,2,t).font=Font(bold=b,size=10)
rows={'Revenue':4,'COGS':5,'Gross':6,'Employee':7,'EBITDA':8,'SGA':9,'NBopen':11,'Capex':12,'Dep':13,
 'NBclose':14,'EBIT':15,'Dopen':17,'Repay':18,'Dclose':19,'LTbor':20,'STbor':21,'Int':22,'OthInc':23,
 'PBT':24,'Tax':25,'PAT':26,'Div':27,'ResOpen':29,'ResClose':30,'Recv':32,'Inv':33,'Pay':34,'OCA':35,
 'OCL':36,'ONCA':37,'ONCL':38,'LTprov':39,'CWIP':40,'Invst':41,'NWC':42,'dNWC':43,'CFO':45,'CFI':46,
 'CFF':47,'NetCF':48,'CashOpen':49,'CashClose':50}
for k,r in rows.items(): lbl(r,k,b=k in('Revenue','Gross','EBITDA','EBIT','PBT','PAT','NetCF','CashClose'))
for c in range(9,17):
    p=CL(c-1); cc=CL(c)
    E(rows['Revenue'],c, "='PL'!I5" if c==9 else "=%s4*(1+%s)"%(p,ad('growth',c)))
    E(rows['COGS'],c, "=-'PL'!I6" if c==9 else "=%s4*(1-%s)"%(cc,ad('gross',c)))
    E(rows['Gross'],c, "=%s4-%s5"%(cc,cc))
    E(rows['Employee'],c, "=-'PL'!I9" if c==9 else "=%s4*%s"%(cc,ad('emp',c)))
    E(rows['EBITDA'],c, "='PL'!I12" if c==9 else "=%s4*%s"%(cc,ad('ebitda',c)))
    E(rows['SGA'],c, "=%s6-%s7-%s8"%(cc,cc,cc))
    E(rows['NBopen'],c, "='BS'!I30" if c==9 else "=%s14"%p)
    E(rows['Capex'],c, 0 if c==9 else "=%s"%ad('capex',c))
    E(rows['Dep'],c, "=-'PL'!I13" if c==9 else "=%s11*%s"%(cc,ad('dep_rate',c)))
    E(rows['NBclose'],c, "='BS'!I30" if c==9 else "=%s11+%s12-%s13"%(cc,cc,cc))
    E(rows['EBIT'],c, "=%s8-%s13"%(cc,cc))
    E(rows['Dopen'],c, "='BS'!I14+'BS'!I19" if c==9 else "=%s19"%p)
    E(rows['Repay'],c, 0 if c==9 else "=%s"%ad('repay',c))
    E(rows['Dclose'],c, "=%s17-%s18"%(cc,cc))
    E(rows['LTbor'],c, "='BS'!I14" if c==9 else "=%s19*%s"%(cc,ad('lt_pct',c)))
    E(rows['STbor'],c, "=%s19-%s20"%(cc,cc))
    E(rows['Int'],c, "=-'PL'!I16" if c==9 else "=%s17*%s"%(cc,ad('int_rate',c)))
    E(rows['OthInc'],c, "='PL'!I15" if c==9 else "=%s4*%s"%(cc,ad('oth_inc',c)))
    E(rows['PBT'],c, "=%s15+%s23-%s22"%(cc,cc,cc))
    E(rows['Tax'],c, "=-'PL'!I18" if c==9 else "=%s24*%s"%(cc,ad('tax',c)))
    E(rows['PAT'],c, "=%s24-%s25"%(cc,cc))
    E(rows['Div'],c, 0 if c==9 else "=%s26*%s"%(cc,ad('payout',c)))
    E(rows['ResOpen'],c, "='BS'!I9" if c==9 else "=%s30"%p)
    E(rows['ResClose'],c, "='BS'!I9" if c==9 else "=%s29+%s26-%s27"%(cc,cc,cc))
    E(rows['Recv'],c, "='BS'!I38" if c==9 else "=%s/365*%s4"%(ad('recv_days',c),cc))
    E(rows['Inv'],c, "='BS'!I39" if c==9 else "=%s/365*%s5"%(ad('inv_days',c),cc))
    E(rows['Pay'],c, "='BS'!I20" if c==9 else "=%s/365*%s5"%(ad('pay_days',c),cc))
    E(rows['OCA'],c, "='BS'!I40" if c==9 else "=%s4*%s"%(cc,ad('oca_pct',c)))
    E(rows['OCL'],c, "='BS'!I21" if c==9 else "=%s4*%s"%(cc,ad('ocl_pct',c)))
    E(rows['ONCA'],c, "='BS'!I34" if c==9 else "=%s37*(1+%s)"%(p,ad('onca_g',c)))
    E(rows['ONCL'],c, "='BS'!I16" if c==9 else "=%s38*(1+%s)"%(p,ad('oncl_g',c)))
    E(rows['LTprov'],c, "='BS'!I15" if c==9 else "=%s39*(1+%s)"%(p,ad('ltprov_g',c)))
    E(rows['CWIP'],c, "='BS'!I31" if c==9 else "=%s"%ad('cwip',c))
    E(rows['Invst'],c, "='BS'!I32" if c==9 else "=%s"%ad('invst',c))
    E(rows['NWC'],c, "=(%s32+%s33+%s35+%s37)-(%s34+%s36+%s38+%s39)"%(cc,cc,cc,cc,cc,cc,cc,cc))
    if c>=10:
        E(rows['dNWC'],c, "=-(%s42-%s42)"%(cc,p))
        E(rows['CFO'],c, "=%s26+%s13+%s43"%(cc,cc,cc))
        E(rows['CFI'],c, "=-%s12-(%s40-%s40)-(%s41-%s41)"%(cc,cc,p,cc,p))
        E(rows['CFF'],c, "=-%s18-%s27"%(cc,cc))
        E(rows['NetCF'],c, "=%s45+%s46+%s47"%(cc,cc,cc))
    E(rows['CashOpen'],c, "='BS'!I41" if c==9 else "=%s50"%p)
    E(rows['CashClose'],c, "='BS'!I41" if c==9 else "=%s49+%s48"%(cc,cc))

# ===================================================================
# 3. Fill PL & BS forecast cols J..P
# ===================================================================
pl=wb['PL']; bs=wb['BS']
def setf(ws,r,c,f,bold=False,fill=None):
    cell=ws.cell(r,c,f); cell.font=Font(color='0000CC',bold=bold); cell.number_format=NF; cell.alignment=RGT
    if fill: cell.fill=fill
for c in FC:
    cc=CL(c); g="='Forecast Engine'!%s"%cc
    setf(pl,5,c, g+"4", True); setf(pl,6,c, "=-'Forecast Engine'!%s5"%cc)
    setf(pl,7,c, "=%s5+%s6"%(cc,cc), True)
    setf(pl,9,c, "=-'Forecast Engine'!%s7"%cc); setf(pl,10,c, "=-'Forecast Engine'!%s9"%cc)
    setf(pl,11,c, "=%s9+%s10"%(cc,cc), True); setf(pl,12,c, "=%s7+%s11"%(cc,cc), True, TOT)
    setf(pl,13,c, "=-'Forecast Engine'!%s13"%cc); setf(pl,14,c, "=%s12+%s13"%(cc,cc), True)
    setf(pl,15,c, g+"23"); setf(pl,16,c, "=-'Forecast Engine'!%s22"%cc)
    setf(pl,17,c, "=%s14+%s15+%s16"%(cc,cc,cc), True); setf(pl,18,c, "=-'Forecast Engine'!%s25"%cc)
    setf(pl,19,c, 0); setf(pl,20,c, "=%s17+%s18+%s19"%(cc,cc,cc), True, TOT)
    setf(bs,8,c, 696.4); setf(bs,9,c, g+"30"); setf(bs,10,c, "=%s8+%s9"%(cc,cc), True)
    setf(bs,14,c, g+"20"); setf(bs,15,c, g+"39"); setf(bs,16,c, g+"38")
    setf(bs,17,c, "=%s14+%s15+%s16"%(cc,cc,cc), True)
    setf(bs,19,c, g+"21"); setf(bs,20,c, g+"34"); setf(bs,21,c, g+"36")
    setf(bs,22,c, "=%s19+%s20+%s21"%(cc,cc,cc), True); setf(bs,23,c, "=%s17+%s22"%(cc,cc), True)
    setf(bs,25,c, "=%s10+%s23"%(cc,cc), True, TOT)
    setf(bs,30,c, g+"14"); setf(bs,31,c, g+"40"); setf(bs,32,c, g+"41")
    setf(bs,33,c, "=%s30+%s31+%s32"%(cc,cc,cc), True); setf(bs,34,c, g+"37")
    setf(bs,35,c, "=%s33+%s34"%(cc,cc), True)
    setf(bs,38,c, g+"32"); setf(bs,39,c, g+"33"); setf(bs,40,c, g+"35"); setf(bs,41,c, g+"50")
    setf(bs,42,c, "=%s38+%s39+%s40+%s41"%(cc,cc,cc,cc), True); setf(bs,44,c, "=%s35+%s42"%(cc,cc), True, TOT)
    setf(bs,47,c, "=%s44-%s25"%(cc,cc))
bs.cell(47,2,'Balance check (TA - TE&L)').font=Font(italic=True,size=9)

# ===================================================================
# 4. DISCREPANCIES
# ===================================================================
if 'Discrepancies' in wb.sheetnames: del wb['Discrepancies']
dq=wb.create_sheet('Discrepancies'); dq.sheet_view.showGridLines=False
dq.column_dimensions['A'].width=2; dq.column_dimensions['B'].width=30
for c in range(3,11): dq.column_dimensions[CL(c)].width=11
dq.column_dimensions[CL(11)].width=60
dq.cell(1,2,'DATA DISCREPANCIES  (BS sheet vs Data Sheet/Screener vs Moneycontrol)').font=Font(bold=True,size=12,color='C00000')
for j,h in enumerate(['Item','FY20','FY21','FY22','FY23','FY24','FY25','FY26','Comment']):
    cell=dq.cell(3,2+j,h); cell.font=BOLD; cell.fill=SUB; cell.alignment=CEN; cell.border=BOX
def drow(r,item,vals,comment):
    dq.cell(r,2,item).font=Font(size=10)
    for j,v in enumerate(vals):
        cell=dq.cell(r,3+j,v); cell.number_format=NF; cell.alignment=RGT; cell.border=BOX
    dq.cell(r,11,comment).font=Font(size=9,italic=True); dq.cell(r,11).alignment=WRAP; dq.row_dimensions[r].height=30
drow(4,'BS-sheet total',[59757.6,55251.9,56243.8,56920.0,59005.5,68083.2,76185.6],'Sum of your Moneycontrol BS line items.')
drow(5,'Data Sheet total (Screener)',[60291.0,55959.7,56990.8,57663.8,59784.1,68848.9,76185.6],'Screener aggregated total.')
drow(6,'Difference',[533.4,707.8,747.0,743.8,778.6,765.7,0.0],'FY20-25 gap ~Rs530-780 cr; FY26 reconciles.')
drow(7,'Inventory (BS sheet)',[8908.2,7194.4,6560.2,6755.9,7220.6,9869.5,13334.6],'Moneycontrol inventory.')
drow(8,'Inventory (Screener)',[9450.7,7914.0,7307.3,7499.7,8002.7,10635.3,13334.6],'Screener inventory.')
drow(9,'Inventory difference',[542.5,719.6,747.1,743.8,782.1,765.8,0.0],'ROOT CAUSE: equals the total gap. Screener likely includes stores/spares Moneycontrol nets elsewhere.')
drow(10,'Borrowings (BS LT+ST)',[4947.9,4849.3,4745.0,5385.0,8808.0,8795.0,8187.0],'Moneycontrol.')
drow(11,'Borrowings (Screener)',[5080.0,4950.9,4829.9,5453.5,8856.5,9014.6,8186.9],'Screener; Rs50-220 cr higher FY20-25; FY26 matches.')
dq.cell(13,2,'Other findings').font=Font(bold=True,color='C00000')
notes=[
 "PL 'Selling and admin' is a residual (Gross - Employee - EBITDA): turns POSITIVE in FY22 (+510.6) & FY23 (+351.1) as it absorbs Screener's volatile 'Other Expenses'. EBITDA/EBIT/PBT/PAT are unaffected and correct.",
 "FY23 Cash differs by Rs55 cr (BS 6,642.6 vs Data Sheet 6,698.1).",
 "FY26 cash Rs11,866.6 cr vs debt Rs8,187 cr => BHEL is NET CASH ~Rs3,680 cr (used in forecast).",
 "RESOLUTION: forecast built on your BS-sheet (Moneycontrol) line items - internally balanced and reconciling to Screener in FY26 (the seed year).",
]
rn=14
for t in notes:
    c=dq.cell(rn,2,t); c.font=Font(size=9,italic=True); c.alignment=WRAP
    dq.merge_cells(start_row=rn,start_column=2,end_row=rn,end_column=11); dq.row_dimensions[rn].height=42; rn+=1

order=['Data Sheet','Sheet1','Assumptions','Forecast Engine','PL','BS','Discrepancies']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save(PATH)
print("SAVED with scenario switch.")
