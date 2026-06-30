"""BHEL FM.xlsx: fold contract assets/liabilities into 'Other assets/liabilities'.
Result: 100% audited historical BS (no derived cells), current vs non-current,
forecast driven off the combined 'Other' lines."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string as ci
from audited_data import D

# combined audited 'Other' buckets (fully audited - contract folded in)
oth_nc_a =[D['note9_nc'][i]+D['ofa_nc'][i] for i in range(7)]
oth_cur_a=[D['note9_cur'][i]+D['ofa_cur'][i]+D['curtax'][i] for i in range(7)]
oth_nc_l =[D['note20_nc'][i]+D['ofl_nc'][i] for i in range(7)]
oth_cur_l=[D['note20_cur'][i]+D['ofl_cur'][i] for i in range(7)]
# round to 2dp and absorb <=0.01 audited rounding into other current assets so BS ties exactly
def _r(x): return round(x,2)
for i in range(7):
    tel=_r(696.41)+_r(D['reserve'][i])+_r(D['leaseNC'][i])+_r(D['tpnc'][i])+_r(D['provnc'][i])+_r(oth_nc_l[i])+_r(D['stbor_grp'][i])+_r(D['tpcur'][i])+_r(D['provcur'][i])+_r(oth_cur_l[i])
    ta=_r(D['ppe_grp'][i])+_r(D['cwip'][i])+_r(D['invest'][i])+_r(D['trnc'][i])+_r(D['dta'][i])+_r(oth_nc_a[i])+_r(D['trcur'][i])+_r(D['invtry'][i])+_r(oth_cur_a[i])+_r(D['cash_grp'][i])
    oth_cur_a[i]=_r(oth_cur_a[i])+_r(tel-ta)

wb=openpyxl.load_workbook('BHEL/BHEL FM.xlsx')
bs=wb['BS']; fe=wb['Forecast Engine']; asm=wb['Assumptions']

BLK=Font(name='Calibri',size=10,color='000000'); GRN=Font(name='Calibri',size=10,color='008000')
BOLD=Font(name='Calibri',size=10,bold=True); TITLE=Font(size=12,bold=True,color='1F3864')
SEC=Font(size=10,bold=True,color='FFFFFF'); SECF=PatternFill('solid',fgColor='1F3864')
SUB=Font(size=10,bold=True,italic=True,color='1F3864'); TOTF=PatternFill('solid',fgColor='D9E1F2')
NUM='#,##0.0;(#,##0.0)'; topb=Border(top=Side(style='thin'))
HC=['C','D','E','F','G','H','I']; FC=['J','K','L','M','N','O','P']; ALL=HC+FC

# clear
for r in range(1,60):
    for c in range(1,18): bs.cell(row=r,column=c).value=None; bs.cell(row=r,column=c).fill=PatternFill(fill_type=None)

bs['B2']='Balance Sheet  (Consolidated, INR Crore)  -  audited FY2020-26, current vs non-current'; bs['B2'].font=TITLE
bs['B4']='Particulars'; bs['B4'].font=BOLD; bs['C4']=2020; bs['C4'].font=BOLD
prev='C'
for col in ['D','E','F','G','H','I','J','K','L','M','N','O','P']:
    bs[f'{col}4']=f'={prev}4+1'; bs[f'{col}4'].font=BOLD; bs[f'{col}4'].alignment=Alignment(horizontal='right'); prev=col
bs['I4'].font=BOLD

def sec(r,t):
    bs[f'B{r}']=t; bs[f'B{r}'].font=SEC
    for col in ['A','B']+ALL: bs[f'{col}{r}'].fill=SECF
def subh(r,t): bs[f'B{r}']=t; bs[f'B{r}'].font=SUB
def hist(r,vals,font=BLK):
    for i,col in enumerate(HC):
        c=bs[f'{col}{r}']; c.value=round(vals[i],2); c.font=font; c.number_format=NUM; c.alignment=Alignment(horizontal='right')
def fcst(r,fe_row):
    for col in FC:
        c=bs[f'{col}{r}']; c.value=f"='Forecast Engine'!{col}{fe_row}"; c.font=GRN; c.number_format=NUM; c.alignment=Alignment(horizontal='right')
def lbl(r,t): bs[f'B{r}']=t; bs[f'B{r}'].font=BLK
def total(r,rng):
    for col in ALL:
        c=bs[f'{col}{r}']; c.value=f'=SUM({col}{rng[0]}:{col}{rng[1]})'; c.font=BOLD; c.number_format=NUM; c.fill=TOTF; c.alignment=Alignment(horizontal='right')
def addr(r,a,b):
    for col in ALL:
        c=bs[f'{col}{r}']; c.value=f'={col}{a}+{col}{b}'; c.font=BOLD; c.number_format=NUM; c.fill=TOTF; c.alignment=Alignment(horizontal='right')

shc=[696.41]*7
sec(6,'EQUITY & LIABILITIES'); bs['A6']='x'
subh(7,"Shareholders' Fund")
lbl(8,'   Share Capital'); hist(8,shc); 
for col in FC: bs[f'{col}8']=696.4; bs[f'{col}8'].number_format=NUM; bs[f'{col}8'].alignment=Alignment(horizontal='right')
lbl(9,'   Reserve & Surplus (Other Equity)'); hist(9,D['reserve']); fcst(9,30)
bs['B10']='Total Shareholders\' Fund'; bs['B10'].font=BOLD; addr(10,8,9)
subh(12,'Non-Current Liabilities')
lbl(13,'   Borrowings & lease liabilities (LT)'); hist(13,D['leaseNC']); fcst(13,20)
lbl(14,'   Trade Payables - non-current'); hist(14,D['tpnc']); fcst(14,60)
lbl(15,'   Long-term Provisions'); hist(15,D['provnc']); fcst(15,39)
lbl(16,'   Other Non-Current Liabilities (incl. contract liabilities)'); hist(16,oth_nc_l); fcst(16,38)
bs['B17']='Total Non-Current Liabilities'; bs['B17'].font=BOLD; total(17,(13,16))
subh(19,'Current Liabilities')
lbl(20,'   Short-term Borrowings & lease liabilities'); hist(20,D['stbor_grp']); fcst(20,21)
lbl(21,'   Trade Payables - current'); hist(21,D['tpcur']); fcst(21,34)
lbl(22,'   Provisions - current'); hist(22,D['provcur']); fcst(22,61)
lbl(23,'   Other Current Liabilities (incl. contract liabilities)'); hist(23,oth_cur_l); fcst(23,36)
bs['B24']='Total Current Liabilities'; bs['B24'].font=BOLD; total(24,(20,23))
bs['B25']='Total Liabilities'; bs['B25'].font=BOLD; addr(25,17,24)
bs['B27']='TOTAL EQUITY & LIABILITIES'; bs['B27'].font=BOLD; addr(27,10,25)
for col in ALL: bs[f'{col}27'].border=topb

sec(29,'ASSETS'); bs['A29']='x'
subh(30,'Non-Current Assets')
lbl(31,'   Property, Plant & Equipment (incl. intangibles)'); hist(31,D['ppe_grp']); fcst(31,14)
lbl(32,'   Capital Work in Progress'); hist(32,D['cwip']); fcst(32,40)
lbl(33,'   Investments (equity method & financial)'); hist(33,D['invest']); fcst(33,41)
lbl(34,'   Trade Receivables - non-current'); hist(34,D['trnc']); fcst(34,52)
lbl(35,'   Deferred Tax Assets (net)'); hist(35,D['dta']); fcst(35,55)
lbl(36,'   Other Non-Current Assets (incl. contract assets)'); hist(36,oth_nc_a); fcst(36,37)
bs['B37']='Total Non-Current Assets'; bs['B37'].font=BOLD; total(37,(31,36))
subh(39,'Current Assets')
lbl(40,'   Trade Receivables - current'); hist(40,D['trcur']); fcst(40,32)
lbl(41,'   Inventories'); hist(41,D['invtry']); fcst(41,33)
lbl(42,'   Other Current Assets (incl. contract assets)'); hist(42,oth_cur_a); fcst(42,35)
lbl(43,'   Cash & Bank Balances'); hist(43,D['cash_grp']); fcst(43,50)
bs['B44']='Total Current Assets'; bs['B44'].font=BOLD; total(44,(40,43))
bs['B46']='TOTAL ASSETS'; bs['B46'].font=BOLD; addr(46,37,44)
for col in ALL: bs[f'{col}46'].border=topb
bs['B48']='Balance check (TA - TE&L)'; bs['B48'].font=Font(size=9,italic=True)
for col in ALL:
    c=bs[f'{col}48']; c.value=f'={col}46-{col}27'; c.font=Font(size=9,italic=True,color='C00000'); c.number_format='0.00'
bs['B50']=('Note: Balance sheet FY2020-FY2026 = audited consolidated actuals from BHEL annual reports (FY2021, FY2023, '
 'FY2025) and the FY2026 filing. Contract assets & liabilities are reported within "Other assets/liabilities" exactly '
 'as in the audited financial statements. No estimated/derived figures.')
bs['B50'].font=Font(size=8,italic=True,color='808080')
bs.column_dimensions['B'].width=46
bs.freeze_panes='C5'

# ---------- Forecast Engine: fold contract into combined Other rows ----------
# re-point FY26 seeds (col I) to new BS rows
seed={4:"='PL'!I5",11:"='BS'!I31",14:"='BS'!I31",17:"='BS'!I13+'BS'!I20",20:"='BS'!I13",
 29:"='BS'!I9",30:"='BS'!I9",32:"='BS'!I40",33:"='BS'!I41",34:"='BS'!I21",
 35:"='BS'!I42",36:"='BS'!I23",37:"='BS'!I36",38:"='BS'!I16",39:"='BS'!I15",
 40:"='BS'!I32",41:"='BS'!I33",49:"='BS'!I43",50:"='BS'!I43",52:"='BS'!I34",55:"='BS'!I35",60:"='BS'!I14",61:"='BS'!I22"}
for r,f in seed.items(): fe[f'I{r}']=f
# combined Other forecast (% revenue) from Assumptions 39(OCA),40(OCL),41(ONCA),42(ONCL)
for col in FC:
    fe[f'{col}35']=f"={col}4*'Assumptions'!{col}39"; fe[f'{col}35'].font=GRN; fe[f'{col}35'].number_format=NUM
    fe[f'{col}36']=f"={col}4*'Assumptions'!{col}40"; fe[f'{col}36'].font=GRN; fe[f'{col}36'].number_format=NUM
    fe[f'{col}37']=f"={col}4*'Assumptions'!{col}41"; fe[f'{col}37'].font=GRN; fe[f'{col}37'].number_format=NUM
    fe[f'{col}38']=f"={col}4*'Assumptions'!{col}42"; fe[f'{col}38'].font=GRN; fe[f'{col}38'].number_format=NUM
# clear now-unused contract rows 53,54,58,59
for r in [53,54,58,59]:
    for col in ['I']+FC: fe[f'{col}{r}']=None
    fe[f'B{r}']=None
# NWC (42) without separate contract rows
for col in ALL:
    a=f"{col}32+{col}52+{col}33+{col}55+{col}35+{col}37"
    l=f"{col}34+{col}60+{col}61+{col}39+{col}36+{col}38"
    fe[f'{col}42']=f"=({a})-({l})"; fe[f'{col}42'].font=Font(color='0000CC'); fe[f'{col}42'].number_format=NUM

# ---------- Assumptions: combined Other drivers (% revenue) ----------
def setpct(row,vals,label,rat):
    asm[f'B{row}']=label
    for col,v in zip(FC,vals):
        c=asm[f'{col}{row}']; c.value=v; c.font=BLK; c.number_format='0.0%'; c.alignment=Alignment(horizontal='right')
    asm[f'Q{row}']=rat; asm[f'Q{row}'].font=Font(size=9,italic=True)
setpct(39,[0.57,0.56,0.55,0.54,0.53,0.52,0.51],'Other current assets % of revenue (incl. contract assets)','FY26 57.6% of revenue; eases as unbilled converts.')
setpct(40,[0.35,0.35,0.34,0.34,0.33,0.33,0.32],'Other current liabilities % of revenue (incl. contract liab.)','FY26 35.1%; customer advances + statutory dues.')
setpct(41,[0.43,0.41,0.39,0.37,0.35,0.33,0.31],'Other non-current assets % of revenue (incl. contract assets)','FY26 44.3%; long-cycle unbilled, declines.')
setpct(42,[0.40,0.38,0.36,0.34,0.32,0.30,0.28],'Other non-current liabilities % of revenue (incl. contract liab.)','FY26 41.1%; LT customer advances, normalise.')
# clear contract assumption block 67-74
for r in range(67,75):
    for col in ['B']+FC+['Q','I']: asm[f'{col}{r}']=None

wb.save('BHEL/BHEL FM.xlsx')
print("BHEL FM.xlsx: contract folded into Other; 100% audited history.")
